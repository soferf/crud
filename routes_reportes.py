"""
routes_reportes.py — Reportes, PDF, Excel exports.
"""
import io
import json
from collections import defaultdict
from datetime import date, datetime, timedelta

from flask import request, session, redirect, url_for, render_template, Response as FlaskResp, send_file

from extensions import app
from db import get_db_connection
from config import MAX_GASTO_POR_HA, TOTAL_HA, MIN_CARGAS_POR_HA, KG_POR_CARGA
from budget import build_budget_movements
from session_service import auth_redirect


@app.route('/reportes')
def reportes():
    if 'user_id' not in session:
        return auth_redirect('login', 'Inicia sesion.', 'warning')
    if not session.get('lote_id'):
        return redirect(url_for('select_lote'))
    lote_id = session['lote_id']

    conn   = get_db_connection()
    cursor = conn.cursor(dictionary=True)

    cursor.execute("SELECT COALESCE(SUM(neto_a_pagar),0) as total FROM recibos WHERE lote_id=%s", (lote_id,))
    total_gastado = float(cursor.fetchone()['total'])

    cursor.execute("""
        SELECT proveedor, SUM(neto_a_pagar) as total
        FROM recibos WHERE neto_a_pagar IS NOT NULL AND lote_id=%s
        GROUP BY proveedor ORDER BY total DESC LIMIT 10
    """, (lote_id,))
    por_trabajador = cursor.fetchall()

    cursor.execute("""
        SELECT YEARWEEK(fecha, 1) AS semana_key, MIN(fecha) AS semana_inicio, SUM(neto_a_pagar) AS total
        FROM recibos WHERE fecha IS NOT NULL AND neto_a_pagar IS NOT NULL AND lote_id=%s
        GROUP BY semana_key ORDER BY semana_key DESC LIMIT 16
    """, (lote_id,))
    por_semana_raw = cursor.fetchall()

    cursor.execute("SELECT SUM(cargas) as total_cargas, SUM(kg_total) as total_kg FROM cosechas WHERE lote_id=%s", (lote_id,))
    produccion   = cursor.fetchone()
    total_cargas = int(produccion['total_cargas'] or 0)
    total_kg     = float(produccion['total_kg'] or 0)

    cursor.execute("SELECT * FROM recibos WHERE lote_id=%s ORDER BY serial DESC LIMIT 5", (lote_id,))
    ultimos_recibos = cursor.fetchall()

    cursor.execute("SELECT hectareas, meta_cargas_ha FROM lotes WHERE id=%s", (lote_id,))
    lote_row   = cursor.fetchone() or {}
    lote_ha    = float(lote_row.get('hectareas') or session.get('lote_ha') or TOTAL_HA)
    meta_cargas = int(lote_ha * (lote_row.get('meta_cargas_ha') or MIN_CARGAS_POR_HA))

    cursor.execute(
        "SELECT id_worker, CONCAT(name,' ',lastname) AS nombre_completo, cc AS nit, trabajo_desarrolla AS cargo "
        "FROM workers WHERE lote_id=%s AND activo=1 ORDER BY name",
        (lote_id,)
    )
    workers_list = cursor.fetchall()

    cursor.execute("SELECT COALESCE(SUM(monto),0) as ti FROM presupuesto_recargas WHERE lote_id=%s", (lote_id,))
    total_ingresado = float(cursor.fetchone()['ti'])
    pres_saldo = total_ingresado - total_gastado
    pres_pct   = round(total_gastado / total_ingresado * 100, 1) if total_ingresado > 0 else 0

    cursor.close(); conn.close()

    ledger              = build_budget_movements(lote_id)
    movimientos_reporte = ledger['movimientos']
    max_gasto           = lote_ha * MAX_GASTO_POR_HA

    for p in por_trabajador:
        p['total'] = float(p['total'])

    por_semana = []
    for row in reversed(por_semana_raw):
        row['total'] = float(row['total'])
        ini = row['semana_inicio']
        row['semana'] = 'Sem. ' + ini.strftime('%d/%m') if hasattr(ini, 'strftime') else str(ini)
        por_semana.append(row)

    pct_gasto      = min(100, round(total_gastado / max_gasto * 100, 1)) if max_gasto else 0
    pct_produccion = min(100, round(total_cargas / meta_cargas * 100, 1)) if meta_cargas else 0

    return render_template('reportes/index.html',
        total_gastado=total_gastado, max_gasto=max_gasto, pct_gasto=pct_gasto,
        por_trabajador=por_trabajador, por_semana=por_semana,
        total_cargas=total_cargas, total_kg=total_kg,
        min_cargas=meta_cargas, pct_produccion=pct_produccion,
        ultimos_recibos=ultimos_recibos, total_ha=lote_ha, max_gasto_ha=MAX_GASTO_POR_HA,
        workers_list=workers_list,
        total_ingresado=total_ingresado, pres_saldo=pres_saldo, pres_pct=pres_pct,
        movimientos_reporte=movimientos_reporte,
        saldo_inicial_reporte=ledger['saldo_inicial'],
        saldo_final_reporte=ledger['saldo_final'],
    )


@app.route('/reportes/semana')
def reporte_semana():
    if 'user_id' not in session:
        return auth_redirect('login', 'Inicia sesion.', 'warning')
    lote_id   = session.get('lote_id')
    fecha_str = request.args.get('fecha', date.today().isoformat())
    try:
        fecha_ref = date.fromisoformat(fecha_str)
    except Exception:
        fecha_ref = date.today()
    inicio = fecha_ref - timedelta(days=fecha_ref.weekday())
    fin    = inicio + timedelta(days=6)

    conn   = get_db_connection()
    cursor = conn.cursor(dictionary=True)
    cursor.execute("SELECT * FROM recibos WHERE fecha BETWEEN %s AND %s AND lote_id=%s ORDER BY serial",
                   (inicio, fin, lote_id))
    recibos = cursor.fetchall()
    cursor.execute("SELECT COALESCE(SUM(neto_a_pagar),0) as total FROM recibos WHERE fecha BETWEEN %s AND %s AND lote_id=%s",
                   (inicio, fin, lote_id))
    total = float(cursor.fetchone()['total'])
    cursor.execute("SELECT COALESCE(SUM(monto),0) as total FROM presupuesto_recargas WHERE fecha BETWEEN %s AND %s AND lote_id=%s",
                   (inicio, fin, lote_id))
    ingresos_semana = float(cursor.fetchone()['total'])
    cursor.close(); conn.close()

    for r in recibos:
        r['neto_a_pagar']    = float(r['neto_a_pagar'] or 0)
        r['valor_operacion'] = float(r['valor_operacion'] or 0)

    ledger = build_budget_movements(lote_id, inicio, fin)
    movimientos_por_serial = {
        m['serial']: m
        for m in ledger['movimientos']
        if m.get('origen') == 'recibo' and m.get('serial') is not None
    }

    return render_template('reportes/semana.html',
        recibos=recibos, total=total, ingresos_semana=ingresos_semana,
        movimientos_semana=ledger['movimientos'],
        movimientos_por_serial=movimientos_por_serial,
        saldo_inicial_semana=ledger['saldo_inicial'],
        saldo_final_semana=ledger['saldo_final'],
        inicio=inicio, fin=fin, fecha_str=fecha_str)


@app.route('/reportes/exportar_txt')
def exportar_txt():
    return redirect(url_for('generar_pdf', tipo='completo'))


@app.route('/reportes/pdf')
def generar_pdf():
    if 'user_id' not in session:
        return auth_redirect('login', 'Inicia sesion.', 'warning')
    if not session.get('lote_id'):
        return redirect(url_for('select_lote'))

    lote_id     = session['lote_id']
    lote_nombre = session.get('lote_nombre', 'Arrocera')
    from fpdf import FPDF

    tipo      = request.args.get('tipo', 'gastos')
    desde_str = request.args.get('desde', '')
    hasta_str = request.args.get('hasta', '')

    C_DEEP   = (27,  67,  50)
    C_FOREST = (45, 106,  79)
    C_SAGE   = (82, 183, 136)
    C_SMINT  = (238, 247, 242)
    C_MINT   = (216, 243, 220)
    C_GOLD   = (233, 168,   0)
    C_WHITE  = (255, 255, 255)
    C_DARK   = ( 27,  45,  30)
    C_MUTED  = (100, 100, 100)

    TIPO_LABELS = {
        'gastos':       'Reporte de Gastos',
        'trabajadores': 'Reporte por Trabajador',
        'produccion':   'Reporte de Produccion',
        'semana':       'Reporte Semanal',
        'completo':     'Recibos Detallados',
        'rango':        'Reporte por Rango de Fechas',
    }

    def fmt_cop(v):
        try:    return '$ {:,.0f}'.format(float(v or 0)).replace(',', '.')
        except: return '$ 0'

    class ReportePDF(FPDF):
        PAGE_W = 210
        MARGIN = 10

        def header(self):
            if self.page_no() == 1:
                return
            self.set_fill_color(*C_DEEP)
            self.rect(0, 0, self.PAGE_W, 16, 'F')
            self.set_fill_color(*C_GOLD)
            self.rect(0, 16, self.PAGE_W, 2, 'F')
            self.set_xy(self.MARGIN, 3)
            self.set_font('Helvetica', 'B', 10)
            self.set_text_color(*C_WHITE)
            self.cell(130, 10, lote_nombre + '  -  Contabilidad Interna', align='L')
            self.set_xy(self.PAGE_W - 60, 3)
            self.set_font('Helvetica', '', 8)
            self.set_text_color(216, 243, 220)
            self.cell(55, 10, f'Gen: {datetime.now().strftime("%d/%m/%Y")}', align='R')
            self.set_y(22)

        def footer(self):
            if self.page_no() == 1:
                return
            self.set_y(-13)
            self.set_font('Helvetica', '', 8)
            self.set_text_color(*C_MUTED)
            self.cell(0, 10, f'Pag. {self.page_no()}  |  {lote_nombre} - Contabilidad Interna', align='C')

        def cover(self, tipo_label, desde_s='', hasta_s=''):
            self.set_fill_color(*C_DEEP)
            self.rect(0, 0, self.PAGE_W, 297, 'F')
            self.set_fill_color(45, 106, 79)
            for row in range(30):
                for col in range(22):
                    self.rect(col * 10 + 4.5, row * 10 + 4.5, 1, 1, 'F')
            self.set_fill_color(*C_GOLD)
            self.rect(0, 292, self.PAGE_W, 5, 'F')
            cw, ch = 160, 90
            cxs = (self.PAGE_W - cw) / 2
            cys = 98
            self.set_fill_color(15, 40, 28)
            self.rect(cxs + 3, cys + 3, cw, ch, 'F')
            self.set_fill_color(*C_WHITE)
            self.rect(cxs, cys, cw, ch, 'F')
            self.set_fill_color(*C_GOLD)
            self.rect(cxs, cys, cw, 5, 'F')
            self.set_xy(cxs, cys + 9)
            self.set_font('Helvetica', 'B', 20)
            self.set_text_color(*C_DEEP)
            self.cell(cw, 12, lote_nombre, align='C')
            self.set_draw_color(*C_SAGE)
            self.set_line_width(0.6)
            self.line(cxs + 25, cys + 23, cxs + cw - 25, cys + 23)
            self.set_xy(cxs, cys + 26)
            self.set_font('Helvetica', 'B', 14)
            self.set_text_color(*C_FOREST)
            self.cell(cw, 10, tipo_label, align='C')
            if desde_s and hasta_s:
                self.set_xy(cxs, cys + 38)
                self.set_font('Helvetica', '', 10)
                self.set_text_color(*C_MUTED)
                self.cell(cw, 8, f'Periodo: {desde_s}  al  {hasta_s}', align='C')
            self.set_xy(cxs, cys + 52)
            self.set_font('Helvetica', '', 9)
            self.set_text_color(150, 150, 150)
            self.cell(cw, 7, str(datetime.now().year), align='C')
            self.set_xy(0, 268)
            self.set_font('Helvetica', 'I', 9)
            self.set_text_color(*C_MINT)
            self.cell(self.PAGE_W, 8, 'Documento de uso exclusivo administrativo  -  Confidencial', align='C')

        def section_title(self, text):
            self.set_font('Helvetica', 'B', 13)
            self.set_text_color(*C_DEEP)
            self.cell(0, 9, text, new_x='LMARGIN', new_y='NEXT')
            self.set_draw_color(*C_SAGE)
            self.set_line_width(0.8)
            self.line(self.MARGIN, self.get_y(), self.PAGE_W - self.MARGIN, self.get_y())
            self.ln(4)
            self.set_text_color(*C_DARK)
            self.set_draw_color(*C_FOREST)
            self.set_line_width(0.3)

        def stat_row(self, stats):
            sw  = (self.PAGE_W - 2 * self.MARGIN) / len(stats)
            y0  = self.get_y()
            bgs = [C_DEEP, C_FOREST, C_SAGE]
            txs = [C_WHITE, C_WHITE, C_WHITE]
            for i, (lbl, val) in enumerate(stats):
                x = self.MARGIN + i * sw
                self.set_fill_color(*bgs[i % 3])
                self.rect(x, y0, sw - 2, 22, 'F')
                self.set_xy(x, y0 + 2)
                self.set_font('Helvetica', 'B', 14)
                self.set_text_color(*txs[i % 3])
                self.cell(sw - 2, 9, str(val), align='C')
                self.set_xy(x, y0 + 13)
                self.set_font('Helvetica', '', 7.5)
                self.cell(sw - 2, 6, lbl, align='C')
            self.set_y(y0 + 27)
            self.set_text_color(*C_DARK)

        def tbl_header(self, headers, widths):
            self.set_fill_color(*C_FOREST)
            self.set_text_color(*C_WHITE)
            self.set_font('Helvetica', 'B', 8)
            self.set_draw_color(*C_DEEP)
            self.set_line_width(0.3)
            for h, w in zip(headers, widths):
                self.cell(w, 7, h, border=1, fill=True)
            self.ln()

        def tbl_row(self, vals, widths, idx=0):
            self.set_fill_color(*(C_SMINT if idx % 2 == 0 else C_WHITE))
            self.set_text_color(*C_DARK)
            self.set_font('Helvetica', '', 7.5)
            self.set_draw_color(200, 220, 208)
            self.set_line_width(0.2)
            for v, w in zip(vals, widths):
                self.cell(w, 6, str(v), border=1, fill=True)
            self.ln()

        def tbl_total(self, label, valor, label_w, valor_w):
            self.set_fill_color(*C_DEEP)
            self.set_text_color(*C_WHITE)
            self.set_font('Helvetica', 'B', 8.5)
            self.set_draw_color(*C_DEEP)
            self.cell(label_w, 7, label, border=1, fill=True)
            self.cell(valor_w, 7, valor, border=1, fill=True, align='R')
            self.ln(3)

    conn   = get_db_connection()
    cursor = conn.cursor(dictionary=True)
    tipo_label = TIPO_LABELS.get(tipo, 'Reporte')

    pdf = ReportePDF(orientation='P', unit='mm', format='A4')
    pdf.set_auto_page_break(auto=True, margin=18)
    pdf.add_page()
    pdf.cover(tipo_label, desde_str, hasta_str)
    pdf.add_page()

    if tipo == 'gastos':
        if desde_str and hasta_str:
            cursor.execute("SELECT * FROM recibos WHERE lote_id=%s AND fecha BETWEEN %s AND %s ORDER BY serial",
                           (lote_id, desde_str, hasta_str))
            subtitle = f'Periodo: {desde_str}  al  {hasta_str}'
        else:
            cursor.execute("SELECT * FROM recibos WHERE lote_id=%s ORDER BY serial", (lote_id,))
            subtitle = 'Todos los registros'
        recibos = cursor.fetchall()
        total   = sum(float(r.get('neto_a_pagar') or 0) for r in recibos)

        pdf.section_title('Reporte de Gastos')
        pdf.set_font('Helvetica', '', 9)
        pdf.set_text_color(*C_MUTED)
        pdf.cell(0, 6, subtitle, new_x='LMARGIN', new_y='NEXT')
        pdf.ln(2)
        pdf.stat_row([('Total Recibos', str(len(recibos))), ('Total Gastado', fmt_cop(total))])

        hdrs = ['#', 'Fecha', 'Proveedor', 'NIT', 'Dirección / Ciudad', 'Concepto', 'Neto']
        wids = [14, 22, 38, 22, 44, 26, 24]
        pdf.tbl_header(hdrs, wids)
        for i, r in enumerate(recibos):
            neto      = float(r.get('neto_a_pagar') or 0)
            fecha_fmt = r['fecha'].strftime('%d/%m/%Y') if r.get('fecha') else '-'
            dir_raw   = (r.get('direccion') or '').strip()
            ciudad_r  = (r.get('ciudad') or '').strip()
            dir_ciudad = (dir_raw + (', ' + ciudad_r if ciudad_r else '')).strip(', ') or '-'
            pdf.tbl_row([str(r['serial']), fecha_fmt,
                         (r.get('proveedor') or '')[:22], (r.get('nit') or ''),
                         dir_ciudad[:26], (r.get('concepto') or '')[:16], fmt_cop(neto)], wids, i)
        pdf.tbl_total(f'  TOTAL -- {len(recibos)} recibos', fmt_cop(total),
                      sum(wids[:-1]), wids[-1])

    elif tipo == 'trabajadores':
        workers_param = request.args.get('workers', '')
        nits_filter   = [n.strip() for n in workers_param.split(',') if n.strip()] if workers_param else []

        if nits_filter:
            fmt_in = ','.join(['%s'] * len(nits_filter))
            cursor.execute(f"""
                SELECT proveedor, nit, COUNT(*) as num_recibos, SUM(neto_a_pagar) as total_pagado
                FROM recibos WHERE lote_id=%s AND nit IN ({fmt_in})
                GROUP BY proveedor, nit ORDER BY total_pagado DESC
            """, (lote_id, *nits_filter))
        else:
            cursor.execute("""
                SELECT proveedor, nit, COUNT(*) as num_recibos, SUM(neto_a_pagar) as total_pagado
                FROM recibos WHERE lote_id=%s GROUP BY proveedor, nit ORDER BY total_pagado DESC
            """, (lote_id,))
        rows  = cursor.fetchall()
        grand = sum(float(r.get('total_pagado') or 0) for r in rows)

        pdf.section_title('Reporte por Trabajador / Proveedor')
        pdf.set_font('Helvetica', '', 9)
        pdf.set_text_color(*C_MUTED)
        subtitle_w = (f'Filtrado: {len(nits_filter)} trabajador(es) seleccionado(s)'
                      if nits_filter else 'Consolidado total de pagos por persona')
        pdf.cell(0, 6, subtitle_w, new_x='LMARGIN', new_y='NEXT')
        pdf.ln(2)
        pdf.stat_row([('Trabajadores', str(len(rows))), ('Total General', fmt_cop(grand))])

        hdrs = ['Proveedor / Trabajador', 'NIT', 'Recibos', 'Total Pagado']
        wids = [90, 30, 20, 50]
        pdf.tbl_header(hdrs, wids)
        for i, r in enumerate(rows):
            t = float(r.get('total_pagado') or 0)
            pdf.tbl_row([(r.get('proveedor') or '')[:43], (r.get('nit') or ''),
                          str(r['num_recibos']), fmt_cop(t)], wids, i)
        pdf.tbl_total(f'  TOTAL -- {len(rows)} proveedores', fmt_cop(grand),
                      sum(wids[:-1]), wids[-1])

        for w in rows:
            nit_w    = w.get('nit') or ''
            nombre_w = (w.get('proveedor') or 'Proveedor sin nombre')[:60]
            if nit_w:
                cursor.execute("""
                    SELECT serial, fecha, concepto, direccion, ciudad, telefono,
                           valor_operacion AS subtotal, rte_fte AS deducciones, neto_a_pagar
                    FROM recibos WHERE lote_id=%s AND nit=%s ORDER BY fecha, serial
                """, (lote_id, nit_w))
            else:
                cursor.execute("""
                    SELECT serial, fecha, concepto, direccion, ciudad, telefono,
                           valor_operacion AS subtotal, rte_fte AS deducciones, neto_a_pagar
                    FROM recibos WHERE lote_id=%s AND proveedor=%s AND (nit IS NULL OR nit='')
                    ORDER BY fecha, serial
                """, (lote_id, w.get('proveedor') or ''))
            det = cursor.fetchall()
            if not det:
                continue

            dir_w    = next((d.get('direccion') or '' for d in det if d.get('direccion')), '')
            ciudad_w = next((d.get('ciudad')    or '' for d in det if d.get('ciudad')),    '')
            tel_w    = next((d.get('telefono')  or '' for d in det if d.get('telefono')),  '')

            pdf.add_page()
            pdf.set_fill_color(*C_FOREST)
            pdf.set_text_color(*C_WHITE)
            pdf.set_font('Helvetica', 'B', 12)
            pdf.rect(pdf.MARGIN, pdf.get_y(), 190, 10, 'F')
            pdf.set_xy(pdf.MARGIN + 2, pdf.get_y() + 1.5)
            pdf.cell(186, 7, f'{nombre_w}  |  NIT: {nit_w or "-"}', align='L')
            pdf.ln(11)
            pdf.set_fill_color(*C_SMINT)
            pdf.rect(pdf.MARGIN, pdf.get_y(), 190, 6, 'F')
            pdf.set_xy(pdf.MARGIN + 2, pdf.get_y() + 0.8)
            pdf.set_font('Helvetica', '', 8)
            pdf.set_text_color(*C_MUTED)
            addr_parts = []
            if dir_w:    addr_parts.append('Dir: ' + dir_w)
            if ciudad_w: addr_parts.append('Ciudad: ' + ciudad_w)
            if tel_w:    addr_parts.append('Tel: ' + tel_w)
            pdf.cell(186, 5, '   '.join(addr_parts) or 'Sin datos de contacto', align='L')
            pdf.ln(8)

            total_w   = sum(float(d.get('neto_a_pagar') or 0) for d in det)
            total_sub = sum(float(d.get('subtotal')     or 0) for d in det)
            pdf.stat_row([('Recibos', str(len(det))),
                          ('Subtotal', fmt_cop(total_sub)),
                          ('Neto Total', fmt_cop(total_w))])

            hdrs_d = ['Serial', 'Fecha', 'Concepto', 'Subtotal', 'Deduc.', 'Neto']
            wids_d = [18, 24, 68, 28, 24, 28]
            pdf.tbl_header(hdrs_d, wids_d)
            for i, d in enumerate(det):
                fecha_d = d['fecha'].strftime('%d/%m/%Y') if d.get('fecha') else '-'
                pdf.tbl_row([str(d['serial']), fecha_d,
                             (d.get('concepto') or '-')[:35],
                             fmt_cop(d.get('subtotal')), fmt_cop(d.get('deducciones')),
                             fmt_cop(d.get('neto_a_pagar'))], wids_d, i)
            pdf.tbl_total(f'  TOTAL -- {len(det)} recibos', fmt_cop(total_w),
                          sum(wids_d[:-1]), wids_d[-1])

    elif tipo == 'produccion':
        cursor.execute("SELECT * FROM cosechas WHERE lote_id=%s ORDER BY fecha DESC", (lote_id,))
        cosechas       = cursor.fetchall()
        tot_cargas     = sum(int(c.get('cargas') or 0) for c in cosechas)
        tot_val        = sum(float(c.get('valor_total') or 0) for c in cosechas)
        cosechas_count = sum(1 for c in cosechas if (c.get('fase') or 'cosecha') == 'cosecha')
        siembras_count = sum(1 for c in cosechas if (c.get('fase') or 'cosecha') == 'siembra')

        pdf.section_title('Reporte de Produccion')
        pdf.set_font('Helvetica', '', 9)
        pdf.set_text_color(*C_MUTED)
        pdf.cell(0, 6, 'Historial de cosechas y siembras registradas', new_x='LMARGIN', new_y='NEXT')
        pdf.ln(2)
        pdf.stat_row([('Cosechas', str(cosechas_count)), ('Siembras', str(siembras_count)),
                      ('Total Cargas', str(tot_cargas)), ('Valor Total', fmt_cop(tot_val))])

        MET  = {'al_voleo': 'Al voleo', 'sembradora': 'Sembradora',
                'labranza_minima': 'Labranza min.', 'otro': 'Otro'}
        hdrs = ['Fase', 'Fecha', 'Lote', 'Variedad', 'Metodo', 'Bultos', 'Cargas', 'Valor']
        wids = [18, 22, 28, 30, 24, 18, 18, 32]
        pdf.tbl_header(hdrs, wids)
        for i, c in enumerate(cosechas):
            fase_label = 'Cosecha' if (c.get('fase') or 'cosecha') == 'cosecha' else 'Siembra'
            cargas    = int(c.get('cargas') or 0)
            val       = float(c.get('valor_total') or 0)
            fecha_fmt = c['fecha'].strftime('%d/%m/%Y') if c.get('fecha') else '-'
            met       = MET.get(c.get('metodo_siembra') or '', '-')
            bultos    = str(float(c.get('total_bultos') or 0) or '-')
            pdf.tbl_row([fase_label, fecha_fmt, (c.get('lote') or '')[:12],
                         (c.get('variedad_semilla') or '-')[:14],
                         met, bultos, str(cargas) if cargas else '-', fmt_cop(val)], wids, i)
        pdf.tbl_total(f'  TOTALES -- {len(cosechas)} registros', fmt_cop(tot_val),
                      sum(wids[:-1]), wids[-1])

    elif tipo == 'semana':
        fecha_ref_str = request.args.get('fecha', date.today().isoformat())
        try:
            fecha_ref = date.fromisoformat(fecha_ref_str)
        except Exception:
            fecha_ref = date.today()
        inicio = fecha_ref - timedelta(days=fecha_ref.weekday())
        fin    = inicio + timedelta(days=6)
        cursor.execute("SELECT * FROM recibos WHERE lote_id=%s AND fecha BETWEEN %s AND %s ORDER BY serial",
                       (lote_id, inicio, fin))
        recibos_s = cursor.fetchall()
        total_s   = sum(float(r.get('neto_a_pagar') or 0) for r in recibos_s)

        pdf.section_title('Reporte Semanal')
        pdf.set_font('Helvetica', '', 9)
        pdf.set_text_color(*C_MUTED)
        pdf.cell(0, 6, f'Semana: {inicio.strftime("%d/%m/%Y")}  al  {fin.strftime("%d/%m/%Y")}',
                 new_x='LMARGIN', new_y='NEXT')
        pdf.ln(2)
        pdf.stat_row([('Recibos', str(len(recibos_s))), ('Total Semana', fmt_cop(total_s))])

        hdrs = ['#', 'Fecha', 'Proveedor', 'NIT', 'Dirección / Ciudad', 'Concepto', 'Neto']
        wids = [14, 22, 38, 22, 44, 26, 24]
        pdf.tbl_header(hdrs, wids)
        for i, r in enumerate(recibos_s):
            neto      = float(r.get('neto_a_pagar') or 0)
            fecha_fmt = r['fecha'].strftime('%d/%m/%Y') if r.get('fecha') else '-'
            dir_raw   = (r.get('direccion') or '').strip()
            ciudad_s  = (r.get('ciudad') or '').strip()
            dir_ciudad = (dir_raw + (', ' + ciudad_s if ciudad_s else '')).strip(', ') or '-'
            pdf.tbl_row([str(r['serial']), fecha_fmt,
                         (r.get('proveedor') or '')[:22], (r.get('nit') or ''),
                         dir_ciudad[:26], (r.get('concepto') or '')[:16], fmt_cop(neto)], wids, i)
        pdf.tbl_total(f'  TOTAL SEMANA -- {len(recibos_s)} recibos', fmt_cop(total_s),
                      sum(wids[:-1]), wids[-1])

    elif tipo == 'completo':
        if desde_str and hasta_str:
            cursor.execute(
                "SELECT * FROM recibos WHERE lote_id=%s AND fecha BETWEEN %s AND %s ORDER BY serial",
                (lote_id, desde_str, hasta_str))
            subtitle = 'Periodo: ' + desde_str + '  al  ' + hasta_str
        else:
            cursor.execute("SELECT * FROM recibos WHERE lote_id=%s ORDER BY serial", (lote_id,))
            subtitle = 'Todos los recibos registrados'
        recibos_c   = cursor.fetchall()
        grand_total = sum(float(r.get('neto_a_pagar') or 0) for r in recibos_c)

        pdf.section_title('Recibos Detallados - Conceptos Completos')
        pdf.set_font('Helvetica', '', 9)
        pdf.set_text_color(*C_MUTED)
        pdf.cell(0, 6, subtitle, new_x='LMARGIN', new_y='NEXT')
        pdf.ln(2)
        pdf.stat_row([('Total Recibos', str(len(recibos_c))), ('Total Neto Pagado', fmt_cop(grand_total))])

        W = 190
        for r in recibos_c:
            if pdf.get_y() > 240:
                pdf.add_page()
            neto      = float(r.get('neto_a_pagar') or 0)
            rte       = float(r.get('rte_fte') or 0)
            subtot    = neto + rte
            fecha_fmt = r['fecha'].strftime('%d/%m/%Y') if r.get('fecha') else '-'
            lineas = []
            if r.get('conceptos_json'):
                try:
                    lineas = json.loads(r['conceptos_json'])
                except Exception:
                    pass
            if not lineas and r.get('concepto'):
                lineas = [{'concepto': r['concepto'], 'valor': float(r.get('valor_operacion') or 0)}]
            y0 = pdf.get_y()
            pdf.set_fill_color(*C_DEEP)
            pdf.rect(pdf.MARGIN, y0, W, 7, 'F')
            pdf.set_xy(pdf.MARGIN + 1, y0)
            pdf.set_font('Helvetica', 'B', 8)
            pdf.set_text_color(*C_WHITE)
            pdf.cell(18, 7, '# ' + str(r.get('serial', '')), fill=False)
            pdf.cell(26, 7, fecha_fmt, fill=False)
            pdf.cell(80, 7, (r.get('proveedor') or '')[:45], fill=False)
            pdf.cell(30, 7, 'NIT: ' + str(r.get('nit') or '-'), fill=False)
            pdf.set_text_color(*C_GOLD)
            pdf.cell(0, 7, 'Neto: ' + fmt_cop(neto), align='R', fill=False)
            pdf.ln(7)
            pdf.set_fill_color(*C_SMINT)
            pdf.rect(pdf.MARGIN, pdf.get_y(), W, 5, 'F')
            pdf.set_xy(pdf.MARGIN + 1, pdf.get_y())
            pdf.set_font('Helvetica', '', 7)
            pdf.set_text_color(*C_MUTED)
            pdf.cell(0, 5,
                     '  Dir: '     + (r.get('direccion') or '-') +
                     '   Ciudad: ' + (r.get('ciudad') or '-') +
                     '   Tel: '    + (r.get('telefono') or '-'),
                     fill=False)
            pdf.ln(5)
            pdf.set_font('Helvetica', 'B', 7)
            pdf.set_fill_color(*C_FOREST)
            pdf.set_text_color(*C_WHITE)
            pdf.set_draw_color(*C_DEEP)
            pdf.set_line_width(0.2)
            pdf.cell(W - 30, 5, '  Concepto', border=1, fill=True)
            pdf.cell(30,     5, 'Valor',      border=1, fill=True, align='R')
            pdf.ln(5)
            for li, ln in enumerate(lineas):
                pdf.set_fill_color(*(C_SMINT if li % 2 == 0 else C_WHITE))
                pdf.set_text_color(*C_DARK)
                pdf.set_font('Helvetica', '', 7)
                pdf.cell(W - 30, 5, '  ' + str(ln.get('concepto') or '')[:85], border=1, fill=True)
                pdf.cell(30,     5, fmt_cop(float(ln.get('valor') or 0)), border=1, fill=True, align='R')
                pdf.ln(5)
            pdf.set_font('Helvetica', '', 7)
            pdf.set_text_color(*C_MUTED)
            pdf.cell(W - 30, 4.5, '')
            pdf.cell(30, 4.5, 'Subtotal: ' + fmt_cop(subtot), align='R')
            pdf.ln(4.5)
            if rte:
                pdf.cell(W - 30, 4.5, '')
                pdf.cell(30, 4.5, 'RTE/FTE: -' + fmt_cop(rte), align='R')
                pdf.ln(4.5)
            pdf.set_fill_color(*C_MINT)
            pdf.set_text_color(*C_DEEP)
            pdf.set_font('Helvetica', 'B', 7.5)
            pdf.cell(W - 30, 5, '', fill=False)
            pdf.cell(30, 5, 'NETO: ' + fmt_cop(neto), fill=True, align='R')
            pdf.ln(7)

    elif tipo == 'rango':
        if not (desde_str and hasta_str):
            desde_str = date.today().replace(day=1).isoformat()
            hasta_str = date.today().isoformat()

        cursor.execute(
            "SELECT * FROM recibos WHERE lote_id=%s AND fecha BETWEEN %s AND %s ORDER BY serial",
            (lote_id, desde_str, hasta_str))
        recibos_r = cursor.fetchall()

        worker_totals = {}
        for r in recibos_r:
            key = (r.get('proveedor') or 'Sin nombre', r.get('nit') or '')
            if key not in worker_totals:
                worker_totals[key] = {'count': 0, 'total': 0.0}
            worker_totals[key]['count'] += 1
            worker_totals[key]['total'] += float(r.get('neto_a_pagar') or 0)
        worker_rows = sorted(worker_totals.items(), key=lambda x: -x[1]['total'])

        week_totals = defaultdict(float)
        for r in recibos_r:
            if r.get('fecha'):
                fd = r['fecha']
                week_start = fd - timedelta(days=fd.weekday())
                week_totals[week_start] += float(r.get('neto_a_pagar') or 0)

        grand_total = sum(float(r.get('neto_a_pagar') or 0) for r in recibos_r)
        avg_recibo  = grand_total / len(recibos_r) if recibos_r else 0
        rte_total   = sum(float(r.get('rte_fte') or 0) for r in recibos_r)

        pdf.section_title('Estadisticas Generales')
        pdf.set_font('Helvetica', '', 9)
        pdf.set_text_color(*C_MUTED)
        pdf.cell(0, 6, 'Periodo: ' + desde_str + '  al  ' + hasta_str, new_x='LMARGIN', new_y='NEXT')
        pdf.ln(2)
        pdf.stat_row([
            ('Recibos', str(len(recibos_r))),
            ('Trabajadores', str(len(worker_rows))),
            ('Total Neto', fmt_cop(grand_total)),
            ('Promedio/recibo', fmt_cop(avg_recibo)),
        ])
        if rte_total:
            pdf.set_font('Helvetica', '', 8)
            pdf.set_text_color(*C_MUTED)
            pdf.ln(2)
            pdf.cell(0, 5, '  RTE/FTE total descontado en el periodo: ' + fmt_cop(rte_total),
                     new_x='LMARGIN', new_y='NEXT')
        pdf.ln(4)

        pdf.section_title('Gastos por Trabajador / Proveedor')
        hdrs_w = ['Trabajador / Proveedor', 'NIT', 'Recibos', 'Total Pagado', '%']
        wids_w = [74, 28, 18, 38, 32]
        pdf.tbl_header(hdrs_w, wids_w)
        for i, ((nombre, nit), vals) in enumerate(worker_rows):
            pct      = (vals['total'] / grand_total * 100) if grand_total else 0
            bar_chars = int(pct / 5)
            bar_str   = '#' * bar_chars + '.' * (20 - bar_chars)
            pdf.tbl_row([nombre[:38], nit[:16], str(vals['count']),
                         fmt_cop(vals['total']), '{:.1f}% {}'.format(pct, bar_str[:10])], wids_w, i)
        pdf.tbl_total('  TOTAL -- ' + str(len(worker_rows)) + ' proveedores',
                      fmt_cop(grand_total), sum(wids_w[:-1]), wids_w[-1])

        if week_totals:
            pdf.ln(6)
            if pdf.get_y() > 220:
                pdf.add_page()
            pdf.section_title('Evolucion Semanal de Gastos')
            hdrs_wk = ['Semana (inicio)', 'Total Pagado', 'Barra de progreso']
            wids_wk = [45, 40, 105]
            pdf.tbl_header(hdrs_wk, wids_wk)
            max_week = max(week_totals.values()) if week_totals else 1
            for i, (wdate, wtotal) in enumerate(sorted(week_totals.items())):
                bar_len = int(wtotal / max_week * 50) if max_week else 0
                pdf.tbl_row([wdate.strftime('%d/%m/%Y'), fmt_cop(wtotal), '|' * bar_len], wids_wk, i)
            pdf.tbl_total('  TOTAL DEL PERIODO', fmt_cop(grand_total),
                          sum(wids_wk[:-1]), wids_wk[-1])

        if recibos_r:
            pdf.ln(6)
            if pdf.get_y() > 200:
                pdf.add_page()
            pdf.section_title('Detalle de Recibos')
            hdrs_d = ['Serial', 'Fecha', 'Proveedor', 'NIT', 'Dirección / Ciudad', 'Concepto', 'RTE/FTE', 'Neto']
            wids_d = [14, 20, 36, 20, 38, 22, 18, 22]
            pdf.tbl_header(hdrs_d, wids_d)
            for i, r in enumerate(recibos_r):
                neto_r    = float(r.get('neto_a_pagar') or 0)
                rte_r     = float(r.get('rte_fte') or 0)
                fecha_fmt = r['fecha'].strftime('%d/%m/%Y') if r.get('fecha') else '-'
                concepto_str = ''
                if r.get('conceptos_json'):
                    try:
                        cjs = json.loads(r['conceptos_json'])
                        if cjs:
                            concepto_str = str(cjs[0].get('concepto') or '')
                    except Exception:
                        pass
                if not concepto_str:
                    concepto_str = r.get('concepto') or ''
                dir_raw    = (r.get('direccion') or '').strip()
                ciudad_rr  = (r.get('ciudad') or '').strip()
                dir_ciudad_r = (dir_raw + (', ' + ciudad_rr if ciudad_rr else '')).strip(', ') or '-'
                pdf.tbl_row([str(r.get('serial', '')), fecha_fmt,
                             (r.get('proveedor') or '')[:20], (r.get('nit') or '')[:12],
                             dir_ciudad_r[:22], concepto_str[:14],
                             fmt_cop(rte_r) if rte_r else '-', fmt_cop(neto_r)], wids_d, i)
            pdf.tbl_total('  TOTAL -- ' + str(len(recibos_r)) + ' recibos',
                          fmt_cop(grand_total), sum(wids_d[:-1]), wids_d[-1])

    cursor.close(); conn.close()
    nombre_archivo = 'reporte_' + tipo + '_' + date.today().isoformat() + '.pdf'
    pdf_bytes = pdf.output()
    return FlaskResp(bytes(pdf_bytes), mimetype='application/pdf',
                     headers={'Content-Disposition': f'inline; filename={nombre_archivo}'})


@app.route('/reportes/excel')
def generar_excel():
    if 'user_id' not in session:
        return auth_redirect('login', 'Inicia sesion.', 'warning')
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    except ImportError:
        return "openpyxl no instalado. Ejecuta: pip install openpyxl", 500

    if not session.get('lote_id'):
        return redirect(url_for('select_lote'))
    lote_id_xl = session['lote_id']

    tipo      = request.args.get('tipo', 'recibos')
    desde_str = request.args.get('desde', '')
    hasta_str = request.args.get('hasta', '')

    wb = openpyxl.Workbook()
    ws = wb.active

    clr_header  = PatternFill('solid', fgColor='2D6A4F')
    clr_alt     = PatternFill('solid', fgColor='F0F8F4')
    fnt_header  = Font(bold=True, color='FFFFFF', size=10)
    fnt_title   = Font(bold=True, color='1B4332', size=13)
    fnt_bold    = Font(bold=True, size=9)
    fnt_norm    = Font(size=9)
    aln_center  = Alignment(horizontal='center', vertical='center')
    aln_right   = Alignment(horizontal='right',  vertical='center')
    thin_side   = Side(style='thin', color='CCCCCC')
    thin_border = Border(left=thin_side, right=thin_side, top=thin_side, bottom=thin_side)

    def hdr_cell(cell, text):
        cell.value = text; cell.font = fnt_header; cell.fill = clr_header
        cell.alignment = aln_center; cell.border = thin_border

    def data_cell(cell, text, alt=False, bold=False, align='left'):
        cell.value = text
        cell.font  = fnt_bold if bold else fnt_norm
        if alt: cell.fill = clr_alt
        cell.alignment = Alignment(horizontal=align, vertical='center')
        cell.border = thin_border

    conn   = get_db_connection()
    cursor = conn.cursor(dictionary=True)

    if tipo == 'recibos':
        ws.title = 'Recibos'
        ws.merge_cells('A1:J1')
        ws['A1'].value = 'Contabilidad Arroceras - Listado de Recibos'
        ws['A1'].font = fnt_title; ws['A1'].alignment = aln_center
        if desde_str and hasta_str:
            cursor.execute("SELECT * FROM recibos WHERE lote_id=%s AND fecha BETWEEN %s AND %s ORDER BY serial",
                           (lote_id_xl, desde_str, hasta_str))
        else:
            cursor.execute("SELECT * FROM recibos WHERE lote_id=%s ORDER BY serial", (lote_id_xl,))
        recibos = cursor.fetchall()
        headers = ['Serial','Fecha','Proveedor','NIT','Dirección','Teléfono','Ciudad','Concepto','Valor Operación','Neto a Pagar']
        for col, h in enumerate(headers, 1):
            hdr_cell(ws.cell(3, col), h)
        ws.row_dimensions[3].height = 20
        for i, r in enumerate(recibos, 4):
            alt      = (i % 2 == 0)
            fecha_val = r['fecha'].strftime('%d/%m/%Y') if r.get('fecha') else ''
            row = [r['serial'], fecha_val, r.get('proveedor',''), r.get('nit',''),
                   r.get('direccion',''), r.get('telefono',''), r.get('ciudad',''),
                   r.get('concepto',''), float(r.get('valor_operacion') or 0), float(r.get('neto_a_pagar') or 0)]
            for col, val in enumerate(row, 1):
                data_cell(ws.cell(i, col), val, alt=alt, align='right' if col >= 9 else 'left')
        tot  = sum(float(r.get('neto_a_pagar') or 0) for r in recibos)
        trow = len(recibos) + 4
        ws.cell(trow, 8).value = f'TOTAL ({len(recibos)} recibos)'; ws.cell(trow, 8).font = fnt_bold
        ws.cell(trow, 10).value = tot
        ws.cell(trow, 10).font  = Font(bold=True, color='FFFFFF', size=9)
        ws.cell(trow, 10).fill  = clr_header
        ws.column_dimensions['C'].width = 28; ws.column_dimensions['H'].width = 40
        for c in ['A','B','D','E','F','G','I','J']: ws.column_dimensions[c].width = 16

    elif tipo == 'trabajadores':
        workers_param_xl = request.args.get('workers', '')
        nits_filter_xl   = [n.strip() for n in workers_param_xl.split(',') if n.strip()] if workers_param_xl else []
        ws.title = 'Por Trabajador'
        ws.merge_cells('A1:D1')
        ws['A1'].value = 'Gastos por Trabajador / Proveedor'
        ws['A1'].font = fnt_title; ws['A1'].alignment = aln_center

        if nits_filter_xl:
            fmt_in_xl = ','.join(['%s'] * len(nits_filter_xl))
            cursor.execute(
                f"SELECT proveedor, nit, COUNT(*) as num_recibos, SUM(neto_a_pagar) as total_pagado "
                f"FROM recibos WHERE lote_id=%s AND nit IN ({fmt_in_xl}) "
                f"GROUP BY proveedor, nit ORDER BY total_pagado DESC",
                (lote_id_xl, *nits_filter_xl))
        else:
            cursor.execute(
                "SELECT proveedor, nit, COUNT(*) as num_recibos, SUM(neto_a_pagar) as total_pagado "
                "FROM recibos WHERE lote_id=%s GROUP BY proveedor, nit ORDER BY total_pagado DESC",
                (lote_id_xl,))
        rows = cursor.fetchall()
        for col, h in enumerate(['Proveedor','NIT','Num. Recibos','Total Pagado'], 1):
            hdr_cell(ws.cell(3, col), h)
        grand = 0
        for i, r in enumerate(rows, 4):
            t = float(r.get('total_pagado') or 0); grand += t
            for col, val in enumerate([r.get('proveedor',''), r.get('nit',''), r['num_recibos'], t], 1):
                data_cell(ws.cell(i, col), val, alt=(i % 2 == 0), align='right' if col >= 3 else 'left')
        trow2 = len(rows) + 4
        ws.cell(trow2, 3).value = 'TOTAL'; ws.cell(trow2, 3).font = fnt_bold
        ws.cell(trow2, 4).value = grand
        ws.cell(trow2, 4).font  = Font(bold=True, color='FFFFFF', size=9)
        ws.cell(trow2, 4).fill  = clr_header
        ws.column_dimensions['A'].width = 35; ws.column_dimensions['B'].width = 18
        ws.column_dimensions['C'].width = 15; ws.column_dimensions['D'].width = 20

        for w in rows:
            nit_w_xl    = w.get('nit') or ''
            nombre_w_xl = (w.get('proveedor') or 'Sin nombre')[:30]
            if nit_w_xl:
                cursor.execute("""SELECT serial, fecha, concepto, direccion, ciudad, telefono,
                                         valor_operacion AS subtotal, rte_fte AS deducciones, neto_a_pagar
                                   FROM recibos WHERE lote_id=%s AND nit=%s ORDER BY fecha, serial""",
                               (lote_id_xl, nit_w_xl))
            else:
                cursor.execute("""SELECT serial, fecha, concepto, direccion, ciudad, telefono,
                                         valor_operacion AS subtotal, rte_fte AS deducciones, neto_a_pagar
                                   FROM recibos WHERE lote_id=%s AND proveedor=%s AND (nit IS NULL OR nit='')
                                   ORDER BY fecha, serial""",
                               (lote_id_xl, w.get('proveedor') or ''))
            det_xl = cursor.fetchall()
            if not det_xl:
                continue
            ws2 = wb.create_sheet(title=nombre_w_xl[:28])
            ws2.merge_cells('A1:I1')
            ws2['A1'].value = nombre_w_xl + (f'  |  NIT: {nit_w_xl}' if nit_w_xl else '')
            ws2['A1'].font  = fnt_title; ws2['A1'].alignment = aln_center
            for col, h in enumerate(['Serial','Fecha','Concepto','Dirección','Ciudad','Teléfono','Subtotal','Deducciones','Neto a Pagar'], 1):
                hdr_cell(ws2.cell(3, col), h)
            tot_w = 0
            for i, d in enumerate(det_xl, 4):
                neto_v = float(d.get('neto_a_pagar') or 0); tot_w += neto_v
                fecha_d = d['fecha'].strftime('%d/%m/%Y') if d.get('fecha') else ''
                row_d = [d['serial'], fecha_d, d.get('concepto','')[:40],
                         d.get('direccion',''), d.get('ciudad',''), d.get('telefono',''),
                         float(d.get('subtotal') or 0), float(d.get('deducciones') or 0), neto_v]
                for col, val in enumerate(row_d, 1):
                    data_cell(ws2.cell(i, col), val, alt=(i % 2 == 0),
                              align='right' if col >= 7 else 'left')
            trow_w = len(det_xl) + 4
            ws2.cell(trow_w, 8).value = 'NETO TOTAL'; ws2.cell(trow_w, 8).font = fnt_bold
            ws2.cell(trow_w, 9).value = tot_w
            ws2.cell(trow_w, 9).font  = Font(bold=True, color='FFFFFF', size=9)
            ws2.cell(trow_w, 9).fill  = clr_header
            ws2.column_dimensions['C'].width = 40
            for c2 in ['A','B','D','E','F','G','H','I']: ws2.column_dimensions[c2].width = 18

    elif tipo == 'produccion':
        ws.title = 'Produccion'
        ws.merge_cells('A1:G1')
        ws['A1'].value = 'Registro de Cosechas y Producción'
        ws['A1'].font  = fnt_title; ws['A1'].alignment = aln_center
        cursor.execute("SELECT * FROM cosechas WHERE lote_id=%s ORDER BY fecha DESC", (lote_id_xl,))
        cosechas = cursor.fetchall()
        for col, h in enumerate(['Fecha','Lote','Hectáreas','Cargas','Kg Total','Precio/Carga','Valor Total','Observaciones'], 1):
            hdr_cell(ws.cell(3, col), h)
        tot_c = tot_k = tot_v = 0
        for i, c in enumerate(cosechas, 4):
            cargas = int(c.get('cargas') or 0)
            kg     = float(c.get('kg_total') or 0)
            val    = float(c.get('valor_total') or 0)
            tot_c += cargas; tot_k += kg; tot_v += val
            fecha_val = c['fecha'].strftime('%d/%m/%Y') if c.get('fecha') else ''
            row_c = [fecha_val, c.get('lote',''), float(c.get('hectareas') or 20), cargas, kg,
                     float(c.get('precio_carga') or 0), val, c.get('observaciones','')]
            for col, val_c in enumerate(row_c, 1):
                data_cell(ws.cell(i, col), val_c, alt=(i % 2 == 0),
                          align='right' if col in [3,4,5,6,7] else 'left')
        trow3 = len(cosechas) + 4
        ws.cell(trow3, 3).value = 'TOTALES'; ws.cell(trow3, 3).font = fnt_bold
        for col, val_t in zip([4,5,7], [tot_c, tot_k, tot_v]):
            ws.cell(trow3, col).value = val_t
            ws.cell(trow3, col).font  = Font(bold=True, color='FFFFFF', size=9)
            ws.cell(trow3, col).fill  = clr_header
        for c2 in ['A','B','C','D','E','F','G']: ws.column_dimensions[c2].width = 18
        ws.column_dimensions['H'].width = 35

    cursor.close(); conn.close()
    buf = io.BytesIO()
    wb.save(buf); buf.seek(0)
    nombre = f'reporte_{tipo}_{date.today().isoformat()}.xlsx'
    return send_file(buf, as_attachment=True, download_name=nombre,
                     mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
