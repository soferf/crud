import os
import uuid
import json
import re
import secrets
import smtplib
import ssl
from email.mime.multipart import MIMEMultipart
from email.mime.text import MIMEText
from datetime import date, timedelta, datetime
from flask import Flask, render_template, render_template_string, request, redirect, url_for, session, jsonify, send_from_directory
import mysql.connector
from werkzeug.security import generate_password_hash, check_password_hash
from werkzeug.utils import secure_filename
from email_utils import (
    render_verify_email,
    render_reset_email,
    render_password_changed_email,
    render_login_alert_email,
    render_signup_code_email,
    render_reset_code_email,
)


# ── .env loader (no external deps) ───────────────────────────────────────────
def _load_dotenv():
    env_path = os.path.join(os.path.dirname(__file__), '.env')
    if os.path.exists(env_path):
        with open(env_path, encoding='utf-8') as _f:
            for _line in _f:
                _line = _line.strip()
                if _line and not _line.startswith('#') and '=' in _line:
                    _k, _, _v = _line.partition('=')
                    os.environ.setdefault(_k.strip(), _v.strip())

_load_dotenv()

# ── Email config (set via .env or environment variables) ─────────────────────
MAIL_SERVER    = os.environ.get('MAIL_SERVER',    'smtp.gmail.com')
MAIL_PORT      = int(os.environ.get('MAIL_PORT',  '587'))
MAIL_USE_TLS   = os.environ.get('MAIL_USE_TLS',   'true').lower() == 'true'
MAIL_USE_SSL   = os.environ.get('MAIL_USE_SSL',   'false').lower() == 'true'
MAIL_TIMEOUT   = int(os.environ.get('MAIL_TIMEOUT', '20'))
AUTH_SEND_LOGIN_ALERT = os.environ.get('AUTH_SEND_LOGIN_ALERT', 'true').lower() == 'true'
MAIL_USERNAME  = os.environ.get('MAIL_USERNAME',  '')
MAIL_PASSWORD  = os.environ.get('MAIL_PASSWORD',  '')
MAIL_FROM_NAME = os.environ.get('MAIL_FROM_NAME', 'Contabilidad Arroceras')
APP_URL        = os.environ.get('APP_URL',         'http://localhost:5000')

app = Flask(__name__)
app.secret_key = os.environ.get('SECRET_KEY', 'cambia-esta-clave-en-produccion')
app.permanent_session_lifetime = timedelta(days=30)

UPLOAD_FOLDER = os.path.join(os.path.dirname(__file__), 'uploads')
os.makedirs(UPLOAD_FOLDER, exist_ok=True)
ALLOWED_EXTENSIONS = {'png', 'jpg', 'jpeg', 'webp'}

DB_CONFIG = {
    'host': 'localhost',
    'user': 'root',
    'password': '',
    'database': 'arrocera_db',
}

WORKER_OPTIONS = [
    'administrador', 'agronomo', 'bombero', 'despalillador',
    'fumigador', 'operario', 'operario_maquinas', 'polivalente',
    'regador', 'transportador', 'versatil'
]
EMAIL_REGEX = re.compile(r'^[^@\s]+@[^@\s]+\.[^@\s]+$')

# Business constants
MAX_GASTO_POR_HA  = 11_000_000
TOTAL_HA          = 20
MAX_GASTO_TOTAL   = 220_000_000
MIN_CARGAS        = 2000
MIN_CARGAS_POR_HA = 100
KG_POR_CARGA      = 62.5

CONCEPTOS_PROHIBIDOS = ['aceite', 'acpm']
AUTH_CODE_TTL_MINUTES = 10
AUTH_CODE_MAX_ATTEMPTS = 5

# =========================
# CONEXIÓN
# =========================
def get_db_connection():
    return mysql.connector.connect(**DB_CONFIG)

def allowed_file(filename):
    return '.' in filename and filename.rsplit('.', 1)[1].lower() in ALLOWED_EXTENSIONS

def is_valid_password(password):
    if len(password) < 8:
        return False
    return any(c.isupper() for c in password) and any(c.islower() for c in password) and any(c.isdigit() for c in password)


def generate_6_digit_code():
    return f"{secrets.randbelow(1000000):06d}"


def save_auth_code(email, purpose, code, payload=None):
    expires_at = datetime.utcnow() + timedelta(minutes=AUTH_CODE_TTL_MINUTES)
    payload_json = json.dumps(payload or {}, ensure_ascii=False)
    conn = get_db_connection()
    cursor = conn.cursor()
    cursor.execute(
        """
        INSERT INTO auth_email_codes (email, purpose, code, payload_json, expires_at, max_attempts)
        VALUES (%s, %s, %s, %s, %s, %s)
        """,
        (email, purpose, code, payload_json, expires_at, AUTH_CODE_MAX_ATTEMPTS)
    )
    conn.commit()
    cursor.close()
    conn.close()


def consume_auth_code(email, purpose, code):
    """Validate and consume a one-time 6-digit code."""
    conn = get_db_connection()
    cursor = conn.cursor(dictionary=True)
    cursor.execute(
        """
        SELECT id, code, payload_json, expires_at, used, attempts, max_attempts
        FROM auth_email_codes
        WHERE email = %s AND purpose = %s
        ORDER BY id DESC
        LIMIT 1
        """,
        (email, purpose)
    )
    row = cursor.fetchone()

    if not row:
        cursor.close(); conn.close()
        return False, 'No se encontró un código activo para este correo.', None

    if row['used']:
        cursor.close(); conn.close()
        return False, 'Este código ya fue utilizado. Solicita uno nuevo.', None

    if datetime.utcnow() > row['expires_at']:
        cursor.close(); conn.close()
        return False, 'El código expiró. Solicita uno nuevo.', None

    if row['attempts'] >= row['max_attempts']:
        cursor.close(); conn.close()
        return False, 'Se alcanzó el número máximo de intentos. Solicita un nuevo código.', None

    if (code or '').strip() != row['code']:
        cursor.execute(
            "UPDATE auth_email_codes SET attempts = attempts + 1 WHERE id = %s",
            (row['id'],)
        )
        conn.commit()
        attempts_left = max(0, row['max_attempts'] - (row['attempts'] + 1))
        cursor.close(); conn.close()
        return False, f'Código incorrecto. Intentos restantes: {attempts_left}.', None

    cursor.execute(
        "UPDATE auth_email_codes SET used = TRUE WHERE id = %s",
        (row['id'],)
    )
    conn.commit()
    payload = json.loads(row['payload_json'] or '{}')
    cursor.close(); conn.close()
    return True, None, payload


# ── Email sender ──────────────────────────────────────────────────────────────
def send_email(to_addr: str, subject: str, html_body: str):
    """Send an HTML email. Returns (ok, error_message)."""
    if not MAIL_USERNAME or not MAIL_PASSWORD:
        msg = 'Email auth not configured. Set MAIL_USERNAME and MAIL_PASSWORD in .env'
        app.logger.error(msg)
        return False, msg
    try:
        msg = MIMEMultipart('alternative')
        msg['Subject'] = subject
        msg['From']    = f'{MAIL_FROM_NAME} <{MAIL_USERNAME}>'
        msg['To']      = to_addr
        msg.attach(MIMEText(html_body, 'html', 'utf-8'))

        if MAIL_USE_SSL:
            with smtplib.SMTP_SSL(MAIL_SERVER, MAIL_PORT, timeout=MAIL_TIMEOUT) as srv:
                srv.login(MAIL_USERNAME, MAIL_PASSWORD)
                srv.sendmail(MAIL_USERNAME, to_addr, msg.as_string())
        else:
            with smtplib.SMTP(MAIL_SERVER, MAIL_PORT, timeout=MAIL_TIMEOUT) as srv:
                srv.ehlo()
                if MAIL_USE_TLS:
                    srv.starttls(context=ssl.create_default_context())
                    srv.ehlo()
                srv.login(MAIL_USERNAME, MAIL_PASSWORD)
                srv.sendmail(MAIL_USERNAME, to_addr, msg.as_string())

        return True, None
    except Exception as exc:
        err_msg = str(exc)
        app.logger.error(f'send_email failed -> {err_msg}')
        return False, err_msg

def auth_redirect(form, message, message_type='info'):
    return redirect(url_for('home', form=form, message=message, type=message_type))


def render_auth_page(form='login', message=None, message_type='info', form_data=None):
    if form not in {'login', 'signup', 'forgot', 'signup_code', 'reset_code'}:
        form = 'login'
    safe_form_data = {
        'login': {'email': '', 'remember_me': False},
        'signup': {'full_name': '', 'email': ''},
        'forgot': {'email': ''},
        'signup_code': {'email': '', 'code': ''},
        'reset_code': {'email': '', 'code': ''}
    }
    if form_data:
        for key, values in form_data.items():
            if key in safe_form_data and isinstance(values, dict):
                safe_form_data[key].update(values)
    return render_template(
        'index.html',
        active_form=form,
        auth_message=message,
        auth_message_type=message_type,
        logged_in='user_id' in session,
        user_name=session.get('user_name'),
        form_data=safe_form_data
    )

def load_trabajadores():
    path = os.path.join(os.path.dirname(__file__), 'data', 'trabajadores.json')
    with open(path, 'r', encoding='utf-8') as f:
        return json.load(f)

def es_concepto_prohibido(concepto):
    """Returns True if concepto is a prohibited purchase (aceite/ACPM direct, not transport)."""
    c = concepto.lower()
    if 'transporte' in c and 'acpm' in c:
        return False  # transporte de acpm is allowed
    for palabra in CONCEPTOS_PROHIBIDOS:
        if palabra in c:
            return True
    return False

def get_serial_inicial(lote_id=None):
    try:
        conn = get_db_connection()
        cursor = conn.cursor(dictionary=True)
        if lote_id is not None:
            cursor.execute("""
                SELECT valor
                FROM config
                WHERE clave = 'serial_inicial'
                  AND (lote_id = %s OR lote_id IS NULL)
                ORDER BY CASE WHEN lote_id = %s THEN 0 ELSE 1 END
                LIMIT 1
            """, (lote_id, lote_id))
        else:
            cursor.execute("SELECT valor FROM config WHERE clave = 'serial_inicial'")
        row = cursor.fetchone()
        cursor.close()
        conn.close()
        return int(row['valor']) if row else 1
    except:
        return 1

def get_next_serial(lote_id=None):
    try:
        serial_configurado = get_serial_inicial(lote_id)
        conn = get_db_connection()
        cursor = conn.cursor()
        if lote_id:
            cursor.execute("SELECT MAX(serial) FROM recibos WHERE lote_id = %s", (lote_id,))
        else:
            cursor.execute("SELECT MAX(serial) FROM recibos")
        max_serial = cursor.fetchone()[0]
        cursor.close()
        conn.close()
        if max_serial is None:
            return serial_configurado
        return max(max_serial + 1, serial_configurado)
    except:
        return get_serial_inicial(lote_id)


def format_currency(value):
    try:
        return '$ {:,.0f}'.format(float(value or 0)).replace(',', '.')
    except Exception:
        return '$ 0'


def build_budget_movements(lote_id, start_date=None, end_date=None):
    """Builds a chronological ledger of budget top-ups and expenses for a lote."""
    conn = get_db_connection()
    cursor = conn.cursor(dictionary=True)

    saldo_inicial = 0.0
    if start_date is not None:
        cursor.execute(
            "SELECT COALESCE(SUM(monto),0) AS total FROM presupuesto_recargas WHERE lote_id=%s AND fecha < %s",
            (lote_id, start_date)
        )
        ingresos_previos = float(cursor.fetchone()['total'] or 0)
        cursor.execute(
            "SELECT COALESCE(SUM(neto_a_pagar),0) AS total FROM recibos WHERE lote_id=%s AND fecha < %s",
            (lote_id, start_date)
        )
        gastos_previos = float(cursor.fetchone()['total'] or 0)
        saldo_inicial = ingresos_previos - gastos_previos

    where_fecha = []
    params = [lote_id]
    if start_date is not None:
        where_fecha.append("fecha >= %s")
        params.append(start_date)
    if end_date is not None:
        where_fecha.append("fecha <= %s")
        params.append(end_date)
    where_sql = f" AND {' AND '.join(where_fecha)}" if where_fecha else ""

    cursor.execute(
        f"""
        SELECT id, fecha, descripcion, monto
        FROM presupuesto_recargas
        WHERE lote_id=%s{where_sql}
        ORDER BY fecha ASC, id ASC
        """,
        tuple(params)
    )
    recargas = cursor.fetchall()

    cursor.execute(
        f"""
        SELECT serial, fecha, proveedor, concepto, neto_a_pagar
        FROM recibos
        WHERE lote_id=%s{where_sql}
        ORDER BY fecha ASC, serial ASC
        """,
        tuple(params)
    )
    recibos = cursor.fetchall()
    cursor.close()
    conn.close()

    movimientos = []
    for recarga in recargas:
        monto = float(recarga.get('monto') or 0)
        movimientos.append({
            'fecha': recarga.get('fecha'),
            'tipo': 'ingreso',
            'tipo_label': 'Ingreso',
            'referencia': f"REC-{recarga.get('id')}",
            'detalle': recarga.get('descripcion') or 'Recarga de presupuesto',
            'ingreso': monto,
            'gasto': 0.0,
            'orden': int(recarga.get('id') or 0),
            'origen': 'presupuesto',
        })

    for recibo in recibos:
        gasto = float(recibo.get('neto_a_pagar') or 0)
        movimientos.append({
            'fecha': recibo.get('fecha'),
            'tipo': 'gasto',
            'tipo_label': 'Gasto',
            'referencia': f"RCB-{recibo.get('serial')}",
            'detalle': recibo.get('proveedor') or recibo.get('concepto') or 'Recibo',
            'detalle_secundario': recibo.get('concepto') or '',
            'ingreso': 0.0,
            'gasto': gasto,
            'orden': int(recibo.get('serial') or 0),
            'origen': 'recibo',
            'serial': recibo.get('serial'),
        })

    movimientos.sort(key=lambda item: (item.get('fecha') or date.min, 0 if item['tipo'] == 'ingreso' else 1, item.get('orden', 0)))

    gasto_acumulado = 0.0
    saldo_actual = saldo_inicial
    for movimiento in movimientos:
        gasto_acumulado += movimiento['gasto']
        saldo_actual += movimiento['ingreso'] - movimiento['gasto']
        movimiento['gasto_acumulado'] = gasto_acumulado
        movimiento['saldo_despues'] = saldo_actual
        movimiento['ingreso_fmt'] = format_currency(movimiento['ingreso']) if movimiento['ingreso'] else '—'
        movimiento['gasto_fmt'] = format_currency(movimiento['gasto']) if movimiento['gasto'] else '—'
        movimiento['gasto_acumulado_fmt'] = format_currency(movimiento['gasto_acumulado'])
        movimiento['saldo_despues_fmt'] = format_currency(movimiento['saldo_despues'])

    return {
        'saldo_inicial': saldo_inicial,
        'saldo_final': saldo_actual,
        'movimientos': movimientos,
    }

# =========================
# INIT DB
# =========================
def init_database():
    conn = mysql.connector.connect(host='localhost', user='root', password='')
    cursor = conn.cursor()
    cursor.execute("CREATE DATABASE IF NOT EXISTS arrocera_db")
    cursor.close()
    conn.close()

    conn = get_db_connection()
    cursor = conn.cursor()

    datadir_cache = None

    def _fix_orphan_tablespace(table_name):
        """Rename orphan .ibd file when table metadata is missing."""
        nonlocal datadir_cache
        try:
            cursor.execute(
                """
                SELECT COUNT(*)
                FROM information_schema.tables
                WHERE table_schema = %s AND table_name = %s
                """,
                (DB_CONFIG['database'], table_name),
            )
            metadata_exists = cursor.fetchone()[0] > 0
            if metadata_exists:
                return False

            if datadir_cache is None:
                cursor.execute("SHOW VARIABLES LIKE 'datadir'")
                datadir_cache = cursor.fetchone()[1]

            db_path = os.path.join(datadir_cache, DB_CONFIG['database'])
            ibd_path = os.path.join(db_path, f"{table_name}.ibd")
            if not os.path.exists(ibd_path):
                return False

            stamp = datetime.now().strftime('%Y%m%d_%H%M%S')
            backup_path = os.path.join(db_path, f"{table_name}.ibd.orphan.{stamp}.bak")
            os.replace(ibd_path, backup_path)
            print(f"[db-repair] tablespace huerfano movido: {backup_path}")
            return True
        except Exception:
            return False

    def ensure_table(table_name, create_sql):
        """Create table and self-heal when metadata/engine are out of sync."""
        try:
            cursor.execute(create_sql)
            cursor.execute(f"SELECT 1 FROM {table_name} LIMIT 1")
            cursor.fetchone()
        except mysql.connector.Error as e:
            errno = getattr(e, 'errno', None)
            if errno in (1813, 1932):
                if errno == 1813:
                    _fix_orphan_tablespace(table_name)
                cursor.execute(f"DROP TABLE IF EXISTS {table_name}")
                try:
                    cursor.execute(create_sql.replace("IF NOT EXISTS ", "", 1))
                    cursor.execute(f"SELECT 1 FROM {table_name} LIMIT 1")
                    cursor.fetchone()
                except mysql.connector.Error as e2:
                    if getattr(e2, 'errno', None) == 1813 and _fix_orphan_tablespace(table_name):
                        cursor.execute(f"DROP TABLE IF EXISTS {table_name}")
                        cursor.execute(create_sql.replace("IF NOT EXISTS ", "", 1))
                        cursor.execute(f"SELECT 1 FROM {table_name} LIMIT 1")
                        cursor.fetchone()
                    else:
                        raise
            else:
                raise

    ensure_table("workers", """
        CREATE TABLE IF NOT EXISTS workers (
            id_worker INT AUTO_INCREMENT PRIMARY KEY,
            name VARCHAR(100) NOT NULL,
            lastname VARCHAR(100) NOT NULL,
            cc VARCHAR(20) UNIQUE NOT NULL,
            phone_number VARCHAR(20) NOT NULL,
            email VARCHAR(100),
            trabajo_desarrolla ENUM('fumigador','agronomo','administrador','operario') NOT NULL,
            fecha_ingreso DATE,
            activo BOOLEAN DEFAULT TRUE,
            observaciones TEXT,
            created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP
        )
    """)

    # Add new columns if they don't exist
    new_cols = [
        ("foto", "VARCHAR(255)"),
        ("alias", "VARCHAR(100)"),
        ("direccion", "VARCHAR(100)"),
        ("ciudad", "VARCHAR(50)"),
        ("concepto_habitual", "TEXT"),
        ("valor_habitual", "DECIMAL(12,2)"),
    ]
    for col_name, col_type in new_cols:
        try:
            cursor.execute(f"ALTER TABLE workers ADD COLUMN {col_name} {col_type}")
        except mysql.connector.Error:
            pass  # Column already exists

    ensure_table("users", """
        CREATE TABLE IF NOT EXISTS users (
            id_user INT AUTO_INCREMENT PRIMARY KEY,
            full_name VARCHAR(120) NOT NULL,
            email VARCHAR(120) UNIQUE NOT NULL,
            password_hash VARCHAR(255) NOT NULL,
            is_active BOOLEAN DEFAULT TRUE,
            created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP
        )
    """)

    # Add email verification columns if they don't exist
    for col_def in [
        ("email_verified", "BOOLEAN DEFAULT FALSE"),
        ("verify_token",   "VARCHAR(100)"),
    ]:
        try:
            cursor.execute(f"ALTER TABLE users ADD COLUMN {col_def[0]} {col_def[1]}")
        except Exception:
            pass

    # Mark pre-existing users (created before email auth) as already verified
    # so they are not locked out after this upgrade
    try:
        cursor.execute("""
            UPDATE users
            SET email_verified = TRUE
            WHERE email_verified = FALSE AND verify_token IS NULL
        """)
    except Exception:
        pass

    ensure_table("password_reset_tokens", """
        CREATE TABLE IF NOT EXISTS password_reset_tokens (
            id         INT AUTO_INCREMENT PRIMARY KEY,
            user_id    INT NOT NULL,
            token      VARCHAR(100) UNIQUE NOT NULL,
            expires_at DATETIME NOT NULL,
            used       BOOLEAN DEFAULT FALSE,
            created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
            FOREIGN KEY (user_id) REFERENCES users(id_user) ON DELETE CASCADE
        )
    """)

    ensure_table("auth_email_codes", """
        CREATE TABLE IF NOT EXISTS auth_email_codes (
            id           INT AUTO_INCREMENT PRIMARY KEY,
            email        VARCHAR(120) NOT NULL,
            purpose      ENUM('signup','reset') NOT NULL,
            code         VARCHAR(6) NOT NULL,
            payload_json TEXT,
            expires_at   DATETIME NOT NULL,
            used         BOOLEAN DEFAULT FALSE,
            attempts     INT DEFAULT 0,
            max_attempts INT DEFAULT 5,
            created_at   TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
            INDEX idx_auth_codes_email_purpose (email, purpose)
        )
    """)

    ensure_table("recibos", """
        CREATE TABLE IF NOT EXISTS recibos (
            serial INT PRIMARY KEY,
            fecha DATE,
            proveedor VARCHAR(100) NOT NULL,
            nit VARCHAR(20),
            direccion VARCHAR(100),
            telefono VARCHAR(20),
            ciudad VARCHAR(50),
            concepto TEXT,
            valor_operacion DECIMAL(12,2),
            neto_a_pagar DECIMAL(12,2),
            tipo VARCHAR(20) DEFAULT 'normal',
            creado_en TIMESTAMP DEFAULT CURRENT_TIMESTAMP
        )
    """)

    ensure_table("config", """
        CREATE TABLE IF NOT EXISTS config (
            id INT AUTO_INCREMENT PRIMARY KEY,
            clave VARCHAR(50) NOT NULL,
            valor VARCHAR(200),
            lote_id INT DEFAULT NULL,
            UNIQUE KEY uq_config_clave_lote (clave, lote_id)
        )
    """)
    # Migrate: add lote_id if the table existed before multi-lote support
    try:
        cursor.execute("ALTER TABLE config ADD COLUMN lote_id INT DEFAULT NULL")
        cursor.execute("ALTER TABLE config ADD UNIQUE KEY uq_config_clave_lote (clave, lote_id)")
        conn.commit()
    except mysql.connector.Error as _e:
        if getattr(_e, 'errno', None) not in (1060, 1061):  # 1060=dup column, 1061=dup key
            raise

    # Legacy migration: config used to have PRIMARY KEY(clave), which breaks
    # per-lote settings because ON DUPLICATE KEY updates another lote's row.
    try:
        cursor.execute("SHOW COLUMNS FROM config LIKE 'id'")
        has_id = cursor.fetchone() is not None
        cursor.execute("SHOW INDEX FROM config WHERE Key_name = 'PRIMARY'")
        primary_rows = cursor.fetchall()
        primary_cols = [row[4] for row in primary_rows]

        if primary_cols == ['clave']:
            cursor.execute("ALTER TABLE config DROP PRIMARY KEY")
            if not has_id:
                cursor.execute("ALTER TABLE config ADD COLUMN id INT AUTO_INCREMENT PRIMARY KEY FIRST")
            else:
                cursor.execute("ALTER TABLE config ADD PRIMARY KEY (id)")
            conn.commit()
        elif not has_id:
            cursor.execute("ALTER TABLE config ADD COLUMN id INT AUTO_INCREMENT PRIMARY KEY FIRST")
            conn.commit()
    except mysql.connector.Error as _e:
        if getattr(_e, 'errno', None) not in (1060, 1068, 1091):
            raise

    try:
        cursor.execute("ALTER TABLE config DROP INDEX clave")
        conn.commit()
    except mysql.connector.Error as _e:
        if getattr(_e, 'errno', None) != 1091:
            raise

    try:
        cursor.execute("ALTER TABLE config ADD UNIQUE KEY uq_config_clave_lote (clave, lote_id)")
        conn.commit()
    except mysql.connector.Error as _e:
        if getattr(_e, 'errno', None) != 1061:
            raise

    ensure_table("cosechas", """
        CREATE TABLE IF NOT EXISTS cosechas (
            id INT AUTO_INCREMENT PRIMARY KEY,
            fecha DATE NOT NULL,
            lote VARCHAR(100) DEFAULT 'El Mangon',
            hectareas DECIMAL(6,2) DEFAULT 20.00,
            cargas INT NOT NULL,
            kg_total DECIMAL(10,2),
            precio_carga DECIMAL(10,2),
            valor_total DECIMAL(14,2),
            observaciones TEXT,
            creado_en TIMESTAMP DEFAULT CURRENT_TIMESTAMP
        )
    """)

    try:
        cursor.execute("""ALTER TABLE workers MODIFY trabajo_desarrolla
            ENUM('fumigador','agronomo','administrador','operario','regador',
                 'bombero','despalillador','operario_maquinas','transportador',
                 'versatil','polivalente')""")
    except Exception:
        pass

    # Add roles_adicionales for secondary/versatile roles
    try:
        cursor.execute("ALTER TABLE workers ADD COLUMN roles_adicionales TEXT")
    except Exception:
        pass

    # Add telefono column to workers if missing
    try:
        cursor.execute("ALTER TABLE workers ADD COLUMN telefono VARCHAR(20)")
    except Exception:
        pass

    # Add rol and conceptos_pago to workers
    for col, typedef in [("rol", "VARCHAR(150)"), ("conceptos_pago", "TEXT")]:
        try:
            cursor.execute(f"ALTER TABLE workers ADD COLUMN {col} {typedef}")
        except Exception:
            pass

    # Add conceptos_json column to recibos if missing (multi-concept support)
    try:
        cursor.execute("ALTER TABLE recibos ADD COLUMN conceptos_json TEXT")
    except Exception:
        pass

    # Add rte_fte column to recibos if missing
    try:
        cursor.execute("ALTER TABLE recibos ADD COLUMN rte_fte DECIMAL(12,2) DEFAULT 0")
    except Exception:
        pass

    # Seed & sowing fields for cosechas
    for col_def in [
        ("variedad_semilla",  "VARCHAR(100)"),
        ("origen_semilla",    "VARCHAR(100)"),
        ("bultos_ha",         "DECIMAL(8,2)"),
        ("total_bultos",      "DECIMAL(8,2)"),
        ("metodo_siembra",    "ENUM('al_voleo','sembradora','labranza_minima','otro') DEFAULT 'al_voleo'"),
        ("fase",              "ENUM('siembra','cosecha') DEFAULT 'cosecha'"),
        ("fecha_siembra",     "DATE"),
        ("precio_carga",      "DECIMAL(10,2)"),
        ("valor_total",       "DECIMAL(14,2)"),
    ]:
        try:
            cursor.execute(f"ALTER TABLE cosechas ADD COLUMN {col_def[0]} {col_def[1]}")
        except Exception:
            pass

    # ── Multi-tenant: lotes ───────────────────────────────────────────────────
    ensure_table("lotes", """
        CREATE TABLE IF NOT EXISTS lotes (
            id                    INT AUTO_INCREMENT PRIMARY KEY,
            nombre                VARCHAR(120) NOT NULL,
            propietario_nombre    VARCHAR(120),
            hectareas             DECIMAL(8,2) DEFAULT 20.00,
            municipio             VARCHAR(100),
            departamento          VARCHAR(100),
            cultivo_principal     VARCHAR(80) DEFAULT 'Arroz',
            fecha_inicio_operacion DATE,
            moneda                VARCHAR(10) DEFAULT 'COP',
            meta_cargas_ha        INT DEFAULT 100,
            limite_gasto_ha       DECIMAL(14,2) DEFAULT 11000000,
            estado                ENUM('activo','inactivo') DEFAULT 'activo',
            created_at            TIMESTAMP DEFAULT CURRENT_TIMESTAMP
        )
    """)

    # ── Roles ─────────────────────────────────────────────────────────────────
    ensure_table("roles", """
        CREATE TABLE IF NOT EXISTS roles (
            id          INT AUTO_INCREMENT PRIMARY KEY,
            nombre      VARCHAR(60) UNIQUE NOT NULL,
            descripcion VARCHAR(200),
            es_global   BOOLEAN DEFAULT FALSE
        )
    """)

    # ── Permisos ──────────────────────────────────────────────────────────────
    ensure_table("permissions", """
        CREATE TABLE IF NOT EXISTS permissions (
            id          INT AUTO_INCREMENT PRIMARY KEY,
            clave       VARCHAR(60) UNIQUE NOT NULL,
            descripcion VARCHAR(200)
        )
    """)

    # ── Relación rol-permisos ─────────────────────────────────────────────────
    ensure_table("role_permissions", """
        CREATE TABLE IF NOT EXISTS role_permissions (
            role_id       INT NOT NULL,
            permission_id INT NOT NULL,
            PRIMARY KEY (role_id, permission_id),
            FOREIGN KEY (role_id)       REFERENCES roles(id)       ON DELETE CASCADE,
            FOREIGN KEY (permission_id) REFERENCES permissions(id)  ON DELETE CASCADE
        )
    """)

    # ── Relación usuario-lote-rol ─────────────────────────────────────────────
    ensure_table("user_lote_roles", """
        CREATE TABLE IF NOT EXISTS user_lote_roles (
            id         INT AUTO_INCREMENT PRIMARY KEY,
            user_id    INT NOT NULL,
            lote_id    INT NOT NULL,
            role_id    INT NOT NULL,
            invited_by INT,
            created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
            UNIQUE KEY uq_user_lote (user_id, lote_id),
            FOREIGN KEY (user_id)    REFERENCES users(id_user) ON DELETE CASCADE,
            FOREIGN KEY (lote_id)    REFERENCES lotes(id)      ON DELETE CASCADE,
            FOREIGN KEY (role_id)    REFERENCES roles(id)      ON DELETE RESTRICT
        )
    """)

    # ── Roles globales (superadmin) ───────────────────────────────────────────
    ensure_table("user_global_roles", """
        CREATE TABLE IF NOT EXISTS user_global_roles (
            user_id    INT NOT NULL,
            role_id    INT NOT NULL,
            PRIMARY KEY (user_id, role_id),
            FOREIGN KEY (user_id)  REFERENCES users(id_user) ON DELETE CASCADE,
            FOREIGN KEY (role_id)  REFERENCES roles(id)      ON DELETE CASCADE
        )
    """)

    # ── Catálogo global de labores ────────────────────────────────────────────
    ensure_table("labor_catalog_global", """
        CREATE TABLE IF NOT EXISTS labor_catalog_global (
            id          INT AUTO_INCREMENT PRIMARY KEY,
            nombre      VARCHAR(120) NOT NULL,
            descripcion TEXT,
            valor_base  DECIMAL(12,2),
            unidad      VARCHAR(40) DEFAULT 'jornal',
            activo      BOOLEAN DEFAULT TRUE,
            created_at  TIMESTAMP DEFAULT CURRENT_TIMESTAMP
        )
    """)

    # ── Overrides de labores por lote ─────────────────────────────────────────
    ensure_table("labor_lote_overrides", """
        CREATE TABLE IF NOT EXISTS labor_lote_overrides (
            lote_id    INT NOT NULL,
            labor_id   INT NOT NULL,
            valor_override DECIMAL(12,2) NOT NULL,
            PRIMARY KEY (lote_id, labor_id),
            FOREIGN KEY (lote_id)  REFERENCES lotes(id)                ON DELETE CASCADE,
            FOREIGN KEY (labor_id) REFERENCES labor_catalog_global(id) ON DELETE CASCADE
        )
    """)

    # ── Invitaciones pendientes ───────────────────────────────────────────────
    ensure_table("lote_invitations", """
        CREATE TABLE IF NOT EXISTS lote_invitations (
            id          INT AUTO_INCREMENT PRIMARY KEY,
            lote_id     INT NOT NULL,
            email       VARCHAR(120) NOT NULL,
            role_id     INT NOT NULL,
            token       VARCHAR(100) UNIQUE NOT NULL,
            invited_by  INT NOT NULL,
            expires_at  DATETIME NOT NULL,
            used        BOOLEAN DEFAULT FALSE,
            created_at  TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
            FOREIGN KEY (lote_id)     REFERENCES lotes(id)      ON DELETE CASCADE,
            FOREIGN KEY (role_id)     REFERENCES roles(id)      ON DELETE RESTRICT,
            FOREIGN KEY (invited_by)  REFERENCES users(id_user) ON DELETE CASCADE
        )
    """)

    # ── Sesiones IA para onboarding ───────────────────────────────────────────
    ensure_table("ai_sessions", """
        CREATE TABLE IF NOT EXISTS ai_sessions (
            id         INT AUTO_INCREMENT PRIMARY KEY,
            user_id    INT NOT NULL,
            status     ENUM('pending','confirmed','cancelled') DEFAULT 'pending',
            created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
            updated_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP ON UPDATE CURRENT_TIMESTAMP,
            FOREIGN KEY (user_id) REFERENCES users(id_user) ON DELETE CASCADE
        )
    """)

    ensure_table("ai_messages", """
        CREATE TABLE IF NOT EXISTS ai_messages (
            id         INT AUTO_INCREMENT PRIMARY KEY,
            session_id INT NOT NULL,
            role       ENUM('user','assistant','system') NOT NULL,
            content    TEXT NOT NULL,
            created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
            FOREIGN KEY (session_id) REFERENCES ai_sessions(id) ON DELETE CASCADE
        )
""")

    ensure_table("presupuesto_recargas", """
        CREATE TABLE IF NOT EXISTS presupuesto_recargas (
            id          INT AUTO_INCREMENT PRIMARY KEY,
            lote_id     INT NOT NULL,
            monto       DECIMAL(15,2) NOT NULL,
            descripcion VARCHAR(255) DEFAULT '',
            fecha       DATE NOT NULL,
            created_at  TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
            INDEX idx_prec_lote (lote_id)
        )
    """)

    ensure_table("ai_form_state", """
        CREATE TABLE IF NOT EXISTS ai_form_state (
            session_id   INT PRIMARY KEY,
            payload_json TEXT,
            step         VARCHAR(40) DEFAULT 'recopilando',
            updated_at   TIMESTAMP DEFAULT CURRENT_TIMESTAMP ON UPDATE CURRENT_TIMESTAMP,
            FOREIGN KEY (session_id) REFERENCES ai_sessions(id) ON DELETE CASCADE
        )
    """)

    # ── Agregar lote_id a tablas operativas (migración segura) ────────────────
    for tbl in ['recibos', 'workers', 'cosechas']:
        try:
            cursor.execute(f"ALTER TABLE {tbl} ADD COLUMN lote_id INT DEFAULT NULL")
        except Exception:
            pass  # Ya existe

    # FK lote_id (solo si lotes ya existe — se hace con try)
    for tbl in ['recibos', 'workers', 'cosechas']:
        try:
            cursor.execute(f"""
                ALTER TABLE {tbl}
                ADD CONSTRAINT fk_{tbl}_lote
                FOREIGN KEY (lote_id) REFERENCES lotes(id) ON DELETE SET NULL
            """)
        except Exception:
            pass  # Ya existe o lotes no tiene datos aún

    conn.commit()
    cursor.close()
    conn.close()

    # Seed roles, permisos y lote inicial
    _seed_roles_and_permissions()
    # Auto-import workers from JSON BEFORE migration so lote_id gets assigned
    _import_trabajadores_from_json()
    _migrate_existing_data_to_initial_lote()


def _seed_roles_and_permissions():
    """Inserta roles, permisos y asignaciones iniciales si no existen."""
    try:
        conn = get_db_connection()
        cursor = conn.cursor()

        # ── Roles ─────────────────────────────────────────────────────────────
        roles = [
            ('superadmin',    'Acceso total al sistema',                          True),
            ('duenio_lote',   'Propietario del lote — acceso completo al lote',   False),
            ('admin_lote',    'Administrador operativo del lote',                  False),
            ('operador_lote', 'Puede crear/editar recibos, trabajadores, cosechas',False),
            ('consulta_lote', 'Solo lectura',                                      False),
        ]
        for nombre, desc, es_global in roles:
            cursor.execute(
                "INSERT IGNORE INTO roles (nombre, descripcion, es_global) VALUES (%s,%s,%s)",
                (nombre, desc, es_global)
            )

        # ── Permisos ──────────────────────────────────────────────────────────
        perms = [
            ('lote.view',          'Ver información del lote'),
            ('lote.manage',        'Gestionar configuración del lote'),
            ('worker.view',        'Ver trabajadores'),
            ('worker.create',      'Registrar trabajadores'),
            ('worker.edit',        'Editar trabajadores'),
            ('worker.toggle',      'Activar/desactivar trabajadores'),
            ('recibo.view',        'Ver recibos'),
            ('recibo.create',      'Crear recibos'),
            ('recibo.edit',        'Editar recibos'),
            ('recibo.delete',      'Eliminar recibos'),
            ('produccion.view',    'Ver registros de producción'),
            ('produccion.create',  'Registrar producción/cosechas'),
            ('produccion.edit',    'Editar registros de producción'),
            ('report.view',        'Ver reportes y estadísticas'),
            ('user.invite',        'Invitar usuarios al lote'),
            ('user.assign_role',   'Asignar roles a usuarios del lote'),
            ('config.manage',      'Gestionar configuración de la aplicación'),
        ]
        for clave, desc in perms:
            cursor.execute(
                "INSERT IGNORE INTO permissions (clave, descripcion) VALUES (%s,%s)",
                (clave, desc)
            )
        conn.commit()

        # ── Asignación de permisos a roles ────────────────────────────────────
        role_perm_map = {
            'superadmin':    [p[0] for p in perms],  # todos
            'duenio_lote':   [p[0] for p in perms],  # todos
            'admin_lote':    [
                'lote.view','lote.manage','worker.view','worker.create','worker.edit',
                'worker.toggle','recibo.view','recibo.create','recibo.edit','recibo.delete',
                'produccion.view','produccion.create','produccion.edit','report.view',
                'user.invite','config.manage',
            ],
            'operador_lote': [
                'lote.view','worker.view','worker.create','worker.edit',
                'recibo.view','recibo.create','recibo.edit',
                'produccion.view','produccion.create','report.view',
            ],
            'consulta_lote': [
                'lote.view','worker.view','recibo.view','produccion.view','report.view',
            ],
        }
        for rol_nombre, perm_claves in role_perm_map.items():
            cursor.execute("SELECT id FROM roles WHERE nombre=%s", (rol_nombre,))
            row = cursor.fetchone()
            if not row:
                continue
            role_id = row[0]
            for clave in perm_claves:
                cursor.execute("SELECT id FROM permissions WHERE clave=%s", (clave,))
                prow = cursor.fetchone()
                if prow:
                    cursor.execute(
                        "INSERT IGNORE INTO role_permissions (role_id, permission_id) VALUES (%s,%s)",
                        (role_id, prow[0])
                    )
        conn.commit()

        # ── Lote inicial "El Mangon" ──────────────────────────────────────────
        cursor.execute("SELECT id FROM lotes WHERE nombre='El Mangon'")
        if not cursor.fetchone():
            cursor.execute("""
                INSERT INTO lotes
                  (nombre, propietario_nombre, hectareas, municipio, departamento,
                   cultivo_principal, moneda, meta_cargas_ha, limite_gasto_ha, estado)
                VALUES ('El Mangon','Fernando',20,'','','Arroz','COP',100,11000000,'activo')
            """)
            conn.commit()

        # ── Catálogo global de labores con tarifas base ───────────────────────
        labores = [
            ('Despalillada',   'Labor de despalille del arroz',          40000, 'jornal'),
            ('Desagüe',        'Labor de desagüe del lote',              45000, 'jornal'),
            ('Abonada',        'Aplicación de abono al cultivo',         40000, 'jornal'),
            ('Fumigación',     'Aplicación de agroquímicos',             50000, 'jornal'),
            ('Regada',         'Labor de riego del cultivo',             40000, 'jornal'),
            ('Cosecha',        'Labor de cosecha del arroz',             55000, 'jornal'),
            ('Transporte',     'Transporte de insumos o cosecha',        80000, 'viaje'),
            ('Maquinaria',     'Uso de maquinaria agrícola',            200000, 'hora'),
            ('Administración', 'Gestión administrativa del lote',       800000, 'mes'),
        ]
        for nombre, desc, valor, unidad in labores:
            cursor.execute(
                "INSERT IGNORE INTO labor_catalog_global (nombre, descripcion, valor_base, unidad) VALUES (%s,%s,%s,%s)",
                (nombre, desc, valor, unidad)
            )
        conn.commit()
        cursor.close(); conn.close()
        print('[seed] Roles, permisos y lote inicial configurados.')
    except Exception as e:
        print(f'[seed] Error: {e}')


def _migrate_existing_data_to_initial_lote():
    """Asigna lote_id=1 (El Mangon) a todos los registros que no tienen lote_id."""
    try:
        conn = get_db_connection()
        cursor = conn.cursor()
        cursor.execute("SELECT id FROM lotes WHERE nombre='El Mangon' LIMIT 1")
        row = cursor.fetchone()
        if not row:
            cursor.close(); conn.close()
            return
        lote_id = row[0]
        for tbl in ['recibos', 'workers', 'cosechas']:
            cursor.execute(f"UPDATE {tbl} SET lote_id=%s WHERE lote_id IS NULL", (lote_id,))
        cursor.execute("""
            DELETE c_null
            FROM config c_null
            JOIN config c_lote
              ON c_null.clave = c_lote.clave
             AND c_null.lote_id IS NULL
             AND c_lote.lote_id = %s
        """, (lote_id,))
        cursor.execute("""
            UPDATE config c
            LEFT JOIN config c_existing
              ON c_existing.clave = c.clave
             AND c_existing.lote_id = %s
            SET c.lote_id = %s
            WHERE c.lote_id IS NULL
              AND c_existing.clave IS NULL
        """, (lote_id, lote_id))
        conn.commit()
        cursor.close(); conn.close()
        print(f'[migration] Datos existentes asignados al lote_id={lote_id} (El Mangon).')
    except Exception as e:
        print(f'[migration] Error: {e}')


def _assign_user_to_initial_lote(user_id: int, role_nombre: str = 'admin_lote'):
    """
    Asigna un usuario recién registrado al lote inicial si es el primer usuario,
    o crea un lote vacío para configurar con el asistente IA.
    Devuelve True si se asignó al lote inicial, False si necesita setup.
    """
    try:
        conn = get_db_connection()
        cursor = conn.cursor(dictionary=True)

        # Verificar si ya tiene lotes asignados
        cursor.execute("SELECT COUNT(*) as cnt FROM user_lote_roles WHERE user_id=%s", (user_id,))
        if cursor.fetchone()['cnt'] > 0:
            cursor.close(); conn.close()
            return True  # Ya tiene lotes

        # Buscar lote inicial
        cursor.execute("SELECT id FROM lotes WHERE nombre='El Mangon' LIMIT 1")
        lote_row = cursor.fetchone()

        if lote_row:
            # ¿Ya hay alguien asignado a este lote con rol admin/dueño?
            cursor.execute("""
                SELECT COUNT(*) as cnt FROM user_lote_roles ulr
                JOIN roles r ON ulr.role_id = r.id
                WHERE ulr.lote_id=%s AND r.nombre IN ('admin_lote','duenio_lote')
            """, (lote_row['id'],))
            tiene_admin = cursor.fetchone()['cnt'] > 0

            if not tiene_admin:
                # Primer usuario → asignar como admin del lote inicial
                cursor.execute("SELECT id FROM roles WHERE nombre=%s", (role_nombre,))
                role_row = cursor.fetchone()
                if role_row:
                    cursor.execute(
                        "INSERT IGNORE INTO user_lote_roles (user_id, lote_id, role_id) VALUES (%s,%s,%s)",
                        (user_id, lote_row['id'], role_row['id'])
                    )
                    conn.commit()
                    cursor.close(); conn.close()
                    return True

        cursor.close(); conn.close()
        return False  # Necesita crear/configurar su propio lote
    except Exception as e:
        print(f'[assign_lote] Error: {e}')
        return False


def _get_user_lotes(user_id: int) -> list:
    """Retorna la lista de lotes del usuario con info del rol."""
    try:
        from auth_middleware import load_user_lotes
        return load_user_lotes(user_id)
    except Exception:
        return []


def _set_active_lote_session(user_id: int, lote_id: int):
    """Actualiza la sesión Flask con el lote activo y sus permisos."""
    from flask import session as _session
    try:
        conn = get_db_connection()
        cursor = conn.cursor(dictionary=True)
        cursor.execute("SELECT nombre, hectareas FROM lotes WHERE id=%s", (lote_id,))
        lote = cursor.fetchone()

        # Rol del usuario en este lote
        cursor.execute("""
            SELECT r.nombre as rol_nombre FROM user_lote_roles ulr
            JOIN roles r ON ulr.role_id = r.id
            WHERE ulr.user_id=%s AND ulr.lote_id=%s
        """, (user_id, lote_id))
        rol_row = cursor.fetchone()
        cursor.close(); conn.close()

        if not lote:
            return False

        _session['lote_id']     = lote_id
        _session['lote_nombre'] = lote['nombre']
        _session['lote_ha']     = float(lote['hectareas'] or 20)
        _session['rol_lote']    = rol_row['rol_nombre'] if rol_row else 'operador_lote'

        from auth_middleware import load_user_lote_perms
        perms = load_user_lote_perms(user_id, lote_id)
        _session['user_perms'] = perms

        # Superadmin check
        conn2 = get_db_connection()
        c2 = conn2.cursor(dictionary=True)
        c2.execute("""
            SELECT 1 FROM user_global_roles ugr
            JOIN roles r ON ugr.role_id = r.id
            WHERE ugr.user_id=%s AND r.nombre='superadmin'
        """, (user_id,))
        _session['is_superadmin'] = c2.fetchone() is not None
        c2.close(); conn2.close()
        return True
    except Exception as e:
        print(f'[set_active_lote] Error: {e}')
        return False


def _import_trabajadores_from_json():
    """Import/update workers from data/trabajadores_arrocera.json into DB."""
    import os as _os
    json_path = _os.path.join(_os.path.dirname(__file__), 'data', 'trabajadores_arrocera.json')
    if not _os.path.exists(json_path):
        return
    try:
        with open(json_path, encoding='utf-8') as f:
            workers_data = json.load(f)
    except Exception as e:
        print(f"[import] Error reading JSON: {e}")
        return

    ROL_TO_ENUM = {
        'agronomo': 'agronomo', 'ingeniero': 'agronomo',
        'regador': 'regador', 'bombero': 'bombero',
        'despalillador': 'despalillador', 'fumigador': 'fumigador',
        'transportador': 'transportador', 'transporte': 'transportador',
        'maquinas': 'operario_maquinas', 'motosierra': 'operario',
        'administrador': 'administrador', 'arrendador': 'administrador',
        'arrendatario': 'administrador', 'propietario': 'operario',
    }

    try:
        conn = get_db_connection()
        cursor = conn.cursor(dictionary=True)
        imported = 0
        updated = 0
        for w in workers_data:
            nit = (w.get('nit') or w.get('cedula') or '').strip()
            if not nit:
                continue
            nombre_completo = w.get('nombre_completo', '').strip()
            parts = nombre_completo.split(' ', 1)
            name = parts[0]
            lastname = parts[1] if len(parts) > 1 else ''
            alias_str = ','.join(w.get('alias', []))
            cpago = w.get('conceptos_pago', [])
            cpago_json = json.dumps(cpago, ensure_ascii=False)
            concepto_h = cpago[0]['descripcion'] if cpago else ''
            valor_h = cpago[0].get('valor_base') if cpago else None
            rol_str = w.get('rol', '').lower()
            trabajo = 'operario'
            for key, val in ROL_TO_ENUM.items():
                if key in rol_str:
                    trabajo = val
                    break

            cursor.execute("SELECT id_worker FROM workers WHERE cc = %s", (nit,))
            existing = cursor.fetchone()
            if existing:
                cursor.execute("""
                    UPDATE workers SET rol = %s, conceptos_pago = %s,
                        concepto_habitual = COALESCE(NULLIF(concepto_habitual,''), %s)
                    WHERE cc = %s
                """, (w.get('rol', ''), cpago_json, concepto_h, nit))
                updated += 1
            else:
                cursor.execute("""
                    INSERT INTO workers
                        (name, lastname, cc, phone_number, alias, direccion, ciudad,
                         trabajo_desarrolla, activo, concepto_habitual, valor_habitual,
                         rol, conceptos_pago, fecha_ingreso)
                    VALUES (%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,'2025-01-01')
                """, (name, lastname, nit, w.get('telefono','') or '',
                      alias_str, w.get('direccion',''), w.get('ciudad',''),
                      trabajo, w.get('activo', True),
                      concepto_h, valor_h,
                      w.get('rol',''), cpago_json))
                imported += 1
        conn.commit()
        cursor.close()
        conn.close()
        if imported or updated:
            print(f"[import] Trabajadores: {imported} nuevos, {updated} actualizados.")
    except Exception as e:
        print(f"[import] Error: {e}")

# =========================
# STATIC UPLOADS
# =========================
@app.route('/uploads/<filename>')
def uploaded_file(filename):
    return send_from_directory(UPLOAD_FOLDER, filename)

# =========================
# HOMEPAGE / AUTH
# =========================
@app.route('/')
def home():
    if 'user_id' in session:
        return redirect(url_for('dashboard'))
    form = request.args.get('form', 'login')
    if form not in {'login', 'signup', 'forgot', 'signup_code', 'reset_code'}:
        form = 'login'

    return render_auth_page(
        form=form,
        message=request.args.get('message'),
        message_type=request.args.get('type', 'info')
    )


@app.route('/dashboard')
def dashboard():
    if 'user_id' not in session:
        return auth_redirect('login', 'Inicia sesion para continuar.', 'warning')
    if not session.get('lote_id'):
        return redirect(url_for('select_lote'))

    lote_id = session['lote_id']
    conn = get_db_connection()
    cursor = conn.cursor(dictionary=True)

    # ── Datos del lote activo ──────────────────────────────────────────────
    cursor.execute("SELECT * FROM lotes WHERE id=%s", (lote_id,))
    lote_info = cursor.fetchone() or {}
    HECTAREAS   = float(lote_info.get('hectareas') or 20)
    LIMITE_HA   = float(lote_info.get('limite_gasto_ha') or 11_000_000)
    META_CARGAS = int(HECTAREAS * (lote_info.get('meta_cargas_ha') or 100))

    # ── Recibos stats ──────────────────────────────────────────────────────
    cursor.execute("SELECT COUNT(*) as cnt, COALESCE(SUM(neto_a_pagar),0) as total FROM recibos WHERE lote_id=%s", (lote_id,))
    rec_stats = cursor.fetchone()
    total_recibos = int(rec_stats['cnt'])
    total_gastos  = float(rec_stats['total'])

    from datetime import date as _date, timedelta as _td
    hoy = _date.today()
    inicio_semana = hoy - _td(days=hoy.weekday())
    fin_semana    = inicio_semana + _td(days=6)
    cursor.execute(
        "SELECT COUNT(*) as cnt, COALESCE(SUM(neto_a_pagar),0) as tot FROM recibos WHERE lote_id=%s AND fecha BETWEEN %s AND %s",
        (lote_id, inicio_semana, fin_semana)
    )
    sem = cursor.fetchone()
    recibos_semana = int(sem['cnt'])
    gasto_semana   = float(sem['tot'])

    cursor.execute(
        "SELECT COALESCE(SUM(neto_a_pagar),0) as tot FROM recibos WHERE lote_id=%s AND YEAR(fecha)=%s AND MONTH(fecha)=%s",
        (lote_id, hoy.year, hoy.month)
    )
    gasto_mes = float(cursor.fetchone()['tot'])

    cursor.execute("SELECT COUNT(*) as cnt FROM workers WHERE lote_id=%s AND activo=1", (lote_id,))
    total_workers = int(cursor.fetchone()['cnt'])

    cursor.execute("SELECT COALESCE(SUM(cargas),0) as tot FROM cosechas WHERE lote_id=%s", (lote_id,))
    total_cargas = int(cursor.fetchone()['tot'])

    cursor.execute("SELECT serial, fecha, proveedor, neto_a_pagar FROM recibos WHERE lote_id=%s ORDER BY serial DESC LIMIT 8", (lote_id,))
    recibos_recientes = cursor.fetchall()

    cursor.execute("SELECT COALESCE(SUM(monto),0) as ti FROM presupuesto_recargas WHERE lote_id=%s", (lote_id,))
    total_ingresado_dash = float(cursor.fetchone()['ti'])
    cursor.close(); conn.close()

    presupuesto_saldo   = total_ingresado_dash - total_gastos
    presupuesto_pct     = round(total_gastos / total_ingresado_dash * 100, 1) if total_ingresado_dash > 0 else 0
    presupuesto_activo  = total_ingresado_dash > 0

    gasto_x_ha = total_gastos / HECTAREAS if HECTAREAS else 0
    gasto_pct  = round(gasto_x_ha / LIMITE_HA * 100, 1) if LIMITE_HA else 0
    alerta_gasto = gasto_x_ha > LIMITE_HA
    cargas_pct  = round(total_cargas / META_CARGAS * 100, 1) if META_CARGAS else 0

    def fmt(v):
        try:
            return '$ {:,.0f}'.format(float(v)).replace(',', '.')
        except Exception:
            return '$ 0'

    stats = {
        'total_recibos':    total_recibos,
        'total_gastos_fmt': fmt(total_gastos),
        'total_workers':    total_workers,
        'total_cargas':     total_cargas,
        'recibos_semana':   recibos_semana,
        'gasto_semana_fmt': fmt(gasto_semana),
        'gasto_mes_fmt':    fmt(gasto_mes),
        'gasto_x_ha_fmt':   fmt(gasto_x_ha),
        'gasto_pct':        gasto_pct,
        'alerta_gasto':     alerta_gasto,
        'cargas_pct':       cargas_pct,
        'meta_cargas':      META_CARGAS,
        'limite_ha_fmt':    fmt(LIMITE_HA),
        'hectareas':        HECTAREAS,
    }

    from auth_middleware import has_permission
    return render_template('dashboard.html',
        user_name=session.get('user_name', 'Usuario'),
        lote_nombre=session.get('lote_nombre', ''),
        rol_lote=session.get('rol_lote', ''),
        today=hoy.strftime('%d/%m/%Y'),
        stats=stats,
        presupuesto_activo=presupuesto_activo,
        presupuesto_saldo=presupuesto_saldo,
        presupuesto_pct=presupuesto_pct,
        total_ingresado_dash=total_ingresado_dash,
        recibos_recientes=recibos_recientes,
        user_perms=session.get('user_perms', []),
    )


@app.route('/auth/signup', methods=['POST'])
def signup():
    full_name = (request.form.get('full_name') or '').strip()
    email = (request.form.get('email') or '').strip().lower()
    password = request.form.get('password') or ''
    confirm_password = request.form.get('confirm_password') or ''

    current_form_data = {
        'signup': {
            'full_name': full_name,
            'email': email
        }
    }

    if not full_name or not email or not password or not confirm_password:
        return render_auth_page('signup', 'Completa todos los campos obligatorios.', 'warning', current_form_data)

    if not EMAIL_REGEX.match(email):
        return render_auth_page('signup', 'Ingresa un correo valido.', 'warning', current_form_data)

    if password != confirm_password:
        return render_auth_page('signup', 'Las contrasenas no coinciden.', 'warning', current_form_data)

    if not is_valid_password(password):
        return render_auth_page(
            'signup',
            'La contrasena debe tener minimo 8 caracteres, mayuscula, minuscula y numero.',
            'warning',
            current_form_data
        )

    try:
        conn = get_db_connection()
        cursor = conn.cursor(dictionary=True)
        cursor.execute("SELECT id_user, email_verified FROM users WHERE email = %s", (email,))
        existing = cursor.fetchone()
        cursor.close(); conn.close()
    except mysql.connector.Error as err:
        return render_auth_page('signup', f'Error al crear la cuenta: {err}', 'danger', current_form_data)

    if existing and existing.get('email_verified'):
        return render_auth_page('signup', 'Ese correo ya esta registrado y verificado.', 'warning', current_form_data)

    code = generate_6_digit_code()
    payload = {
        'full_name': full_name,
        'email': email,
        'password_hash': generate_password_hash(password),
    }

    try:
        save_auth_code(email, 'signup', code, payload)
    except mysql.connector.Error as err:
        return render_auth_page('signup', f'No se pudo generar el codigo: {err}', 'danger', current_form_data)

    html_body = render_signup_code_email(full_name, code)
    sent_ok, send_err = send_email(email, 'Codigo de verificacion - Contabilidad Arroceras', html_body)

    if sent_ok:
        return render_auth_page(
            'signup_code',
            f'Revisamos tu correo {email}. Ingresa el codigo de 6 digitos para activar la cuenta.',
            'success',
            {'signup_code': {'email': email, 'code': ''}}
        )

    return render_auth_page(
        'signup',
        (
            f'No se pudo enviar el codigo de verificacion a {email}. '
            f'Verifica tu SMTP de Gmail e intenta nuevamente. '
            f'Detalle tecnico: {send_err}'
        ),
        'warning',
        current_form_data
    )


@app.route('/auth/verify-signup-code', methods=['POST'])
def verify_signup_code():
    email = (request.form.get('email') or '').strip().lower()
    code = re.sub(r'\D', '', request.form.get('code') or '')

    current_form_data = {'signup_code': {'email': email, 'code': code}}

    if not EMAIL_REGEX.match(email):
        return render_auth_page('signup_code', 'Correo no valido.', 'warning', current_form_data)
    if len(code) != 6:
        return render_auth_page('signup_code', 'El codigo debe tener 6 digitos.', 'warning', current_form_data)

    ok, err_msg, payload = consume_auth_code(email, 'signup', code)
    if not ok:
        return render_auth_page('signup_code', err_msg, 'warning', current_form_data)

    try:
        conn = get_db_connection()
        cursor = conn.cursor(dictionary=True)
        cursor.execute("SELECT id_user, email_verified FROM users WHERE email = %s", (email,))
        user = cursor.fetchone()

        if user and user.get('email_verified'):
            cursor.close(); conn.close()
            return render_auth_page('login', 'La cuenta ya esta verificada. Inicia sesion.', 'info')

        if user:
            cursor2 = conn.cursor()
            new_name = payload.get('full_name') or ''
            new_hash = payload.get('password_hash') or None
            if new_hash:
                cursor2.execute(
                    """
                    UPDATE users
                    SET full_name = %s, password_hash = %s, email_verified = TRUE, verify_token = NULL, is_active = TRUE
                    WHERE id_user = %s
                    """,
                    (new_name, new_hash, user['id_user'])
                )
            else:
                cursor2.execute(
                    """
                    UPDATE users
                    SET full_name = %s, email_verified = TRUE, verify_token = NULL, is_active = TRUE
                    WHERE id_user = %s
                    """,
                    (new_name or user.get('full_name') or 'Usuario', user['id_user'])
                )
            cursor2.close()
        else:
            cursor2 = conn.cursor()
            cursor2.execute(
                """
                INSERT INTO users (full_name, email, password_hash, email_verified, verify_token, is_active)
                VALUES (%s, %s, %s, TRUE, NULL, TRUE)
                """,
                (payload.get('full_name', ''), email, payload.get('password_hash', ''))
            )
            cursor2.close()

        conn.commit()
        cursor.close(); conn.close()
    except mysql.connector.Error as err:
        return render_auth_page('signup_code', f'Error al verificar la cuenta: {err}', 'danger', current_form_data)

    return render_auth_page('login', 'Cuenta verificada correctamente. Ya puedes iniciar sesion.', 'success')


# =========================
# SELECCIÓN DE LOTE
# =========================
@app.route('/select-lote')
def select_lote():
    if 'user_id' not in session:
        return auth_redirect('login', 'Inicia sesion para continuar.', 'warning')
    lotes = _get_user_lotes(session['user_id'])
    if len(lotes) == 0:
        return redirect(url_for('setup_lote_nuevo'))
    if len(lotes) == 1:
        _set_active_lote_session(session['user_id'], lotes[0]['id'])
        return redirect(url_for('dashboard'))
    return render_template('select_lote.html',
                           lotes=lotes,
                           user_name=session.get('user_name', 'Usuario'))


@app.route('/select-lote/<int:lote_id>')
def cambiar_lote(lote_id):
    if 'user_id' not in session:
        return auth_redirect('login', 'Inicia sesion para continuar.', 'warning')
    lotes = _get_user_lotes(session['user_id'])
    lote_ids = [l['id'] for l in lotes]
    if lote_id not in lote_ids and not session.get('is_superadmin'):
        return auth_redirect('login', 'Acceso denegado a ese lote.', 'danger')
    _set_active_lote_session(session['user_id'], lote_id)
    return redirect(url_for('dashboard'))


# =========================
# SETUP LOTE NUEVO (wizard IA)
# =========================
@app.route('/setup/lote', methods=['GET'])
def setup_lote_nuevo():
    if 'user_id' not in session:
        return auth_redirect('login', 'Inicia sesion para continuar.', 'warning')
    from ai_service import AI_ENABLED, OLLAMA_MODEL, ollama as _ollama
    if AI_ENABLED:
        ok_ai, ai_msg = _ollama.health_check()
    else:
        ok_ai, ai_msg = False, 'IA deshabilitada'
    return render_template('setup_lote.html',
                           user_name=session.get('user_name', 'Usuario'),
                           ai_available=ok_ai,
                           ai_message=ai_msg,
                           ai_model=OLLAMA_MODEL)


@app.route('/setup/lote/chat', methods=['POST'])
def setup_lote_chat():
    if 'user_id' not in session:
        return jsonify({'error': 'No autenticado'}), 401

    data = request.get_json(force=True) or {}
    user_message = (data.get('message') or '').strip()[:1000]

    if not user_message:
        return jsonify({'error': 'Mensaje vacío'}), 400

    user_id = session['user_id']

    try:
        conn = get_db_connection()
        cursor = conn.cursor(dictionary=True)

        cursor.execute(
            "SELECT id FROM ai_sessions WHERE user_id=%s AND status='pending' ORDER BY id DESC LIMIT 1",
            (user_id,)
        )
        ai_sess = cursor.fetchone()
        if not ai_sess:
            cursor2 = conn.cursor()
            cursor2.execute(
                "INSERT INTO ai_sessions (user_id, status) VALUES (%s, 'pending')", (user_id,)
            )
            conn.commit()
            ai_sess_id = cursor2.lastrowid
            cursor2.execute(
                "INSERT INTO ai_form_state (session_id, payload_json, step) VALUES (%s,'{}','recopilando')",
                (ai_sess_id,)
            )
            conn.commit()
            cursor2.close()
        else:
            ai_sess_id = ai_sess['id']

        cursor.execute("SELECT payload_json FROM ai_form_state WHERE session_id=%s", (ai_sess_id,))
        state = cursor.fetchone()
        current_payload = json.loads(state['payload_json'] or '{}') if state else {}

        cursor.execute(
            "SELECT role, content FROM ai_messages WHERE session_id=%s ORDER BY id",
            (ai_sess_id,)
        )
        history = cursor.fetchall()
        cursor.close(); conn.close()

        from ai_service import LoteOnboardingSession
        ai_session = LoteOnboardingSession(ai_sess_id, current_payload)
        for msg in history:
            ai_session.messages.append({'role': msg['role'], 'content': msg['content']})

        result = ai_session.process(user_message)

        conn2 = get_db_connection()
        c2 = conn2.cursor()
        c2.execute(
            "INSERT INTO ai_messages (session_id, role, content) VALUES (%s,'user',%s)",
            (ai_sess_id, user_message)
        )
        c2.execute(
            "INSERT INTO ai_messages (session_id, role, content) VALUES (%s,'assistant',%s)",
            (ai_sess_id, result.get('mensaje', ''))
        )
        c2.execute(
            "UPDATE ai_form_state SET payload_json=%s, step=%s WHERE session_id=%s",
            (json.dumps(ai_session.payload, ensure_ascii=False),
             result.get('paso_actual', 'recopilando'), ai_sess_id)
        )
        conn2.commit()
        c2.close(); conn2.close()

        return jsonify({
            'mensaje':            result.get('mensaje', ''),
            'paso_actual':        result.get('paso_actual', 'recopilando'),
            'campos_faltantes':   result.get('campos_faltantes', []),
            'payload_actual':     ai_session.payload,
            'listo_para_guardar': result.get('listo_para_guardar', False),
            'ai_session_id':      ai_sess_id,
            'degraded':           result.get('degraded', False),
        })
    except Exception as e:
        return jsonify({'error': f'Error interno: {e}', 'mensaje': f'Ocurrió un error: {e}'}), 500


@app.route('/setup/lote/confirmar', methods=['POST'])
def setup_lote_confirmar():
    if 'user_id' not in session:
        return jsonify({'error': 'No autenticado'}), 401

    # Handle both JSON (AJAX) and form-encoded (HTML form fallback) submissions.
    # Flask consumes form-encoded bodies into request.form; calling get_json()
    # afterwards would see an empty body and raise "Expecting value: line 1 col 1".
    is_json = 'application/json' in (request.content_type or '')
    if is_json:
        data = request.get_json(force=True, silent=True) or {}
        payload = data.get('payload') or {}
        ai_session_id = data.get('ai_session_id')
    else:
        # HTML fallback form submission
        ubicacion_raw = request.form.get('ubicacion', '')
        municipio_val  = request.form.get('municipio', ubicacion_raw.split(',')[0].strip())
        departamento_val = request.form.get('departamento',
                          ubicacion_raw.split(',')[1].strip() if ',' in ubicacion_raw else '')
        payload = {
            'nombre_lote':            request.form.get('nombre_lote', '').strip(),
            'propietario':            request.form.get('propietario', '').strip(),
            'hectareas':              request.form.get('hectareas', ''),
            'municipio':              municipio_val,
            'departamento':           departamento_val,
            'cultivo_principal':      request.form.get('cultivo_principal', 'Arroz').strip(),
            'fecha_inicio_operacion': request.form.get('fecha_inicio_operacion', '').strip() or None,
            'moneda':                 request.form.get('moneda', 'COP').strip(),
            'meta_cargas_ha':         request.form.get('meta_cargas_ha', ''),
            'limite_gasto_ha':        request.form.get('limite_gasto_ha', ''),
        }
        ai_session_id = None

    from ai_service import validate_lote_payload, apply_field_defaults
    payload = apply_field_defaults(payload)
    valid, errors = validate_lote_payload(payload)
    if not valid:
        return jsonify({'error': 'Datos inválidos', 'errores': errors}), 400

    user_id = session['user_id']
    try:
        conn = get_db_connection()
        cursor = conn.cursor()
        cursor.execute("""
            INSERT INTO lotes
              (nombre, propietario_nombre, hectareas, municipio, departamento,
               cultivo_principal, fecha_inicio_operacion, moneda,
               meta_cargas_ha, limite_gasto_ha, estado)
            VALUES (%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,'activo')
        """, (
            str(payload.get('nombre_lote', 'Mi Lote')),
            str(payload.get('propietario', '')),
            float(payload.get('hectareas', 20)),
            str(payload.get('municipio', '')),
            str(payload.get('departamento', '')),
            str(payload.get('cultivo_principal', 'Arroz')),
            payload.get('fecha_inicio_operacion') or None,
            str(payload.get('moneda', 'COP')),
            int(payload.get('meta_cargas_ha', 100)),
            float(payload.get('limite_gasto_ha', 11000000)),
        ))
        new_lote_id = cursor.lastrowid
        cursor.execute(
            "INSERT IGNORE INTO config (clave, valor, lote_id) VALUES ('serial_inicial','1',%s)",
            (new_lote_id,)
        )
        cursor.execute("SELECT id FROM roles WHERE nombre='duenio_lote'")
        role_row = cursor.fetchone()
        if role_row:
            cursor.execute(
                "INSERT IGNORE INTO user_lote_roles (user_id, lote_id, role_id) VALUES (%s,%s,%s)",
                (user_id, new_lote_id, role_row[0])
            )
        if ai_session_id:
            cursor.execute(
                "UPDATE ai_sessions SET status='confirmed' WHERE id=%s AND user_id=%s",
                (ai_session_id, user_id)
            )
        conn.commit()
        cursor.close(); conn.close()
        _set_active_lote_session(user_id, new_lote_id)
        if not is_json:
            return redirect(url_for('dashboard'))
        return jsonify({'success': True, 'lote_id': new_lote_id, 'redirect': url_for('dashboard')})
    except Exception as e:
        return jsonify({'error': f'Error al guardar: {e}'}), 500


@app.route('/setup/lote/cancelar', methods=['POST'])
def setup_lote_cancelar():
    if 'user_id' not in session:
        return jsonify({'error': 'No autenticado'}), 401
    data = request.get_json(force=True) or {}
    ai_session_id = data.get('ai_session_id')
    if ai_session_id:
        try:
            conn = get_db_connection()
            cursor = conn.cursor()
            cursor.execute(
                "UPDATE ai_sessions SET status='cancelled' WHERE id=%s AND user_id=%s",
                (ai_session_id, session['user_id'])
            )
            conn.commit()
            cursor.close(); conn.close()
        except Exception:
            pass
    lotes = _get_user_lotes(session['user_id'])
    if lotes:
        return jsonify({'success': True, 'redirect': url_for('select_lote')})
    return jsonify({'success': True, 'redirect': url_for('home')})


@app.route('/lote/invitar', methods=['POST'])
def lote_invitar():
    from auth_middleware import has_permission
    if 'user_id' not in session:
        return jsonify({'error': 'No autenticado'}), 401
    if not session.get('lote_id'):
        return jsonify({'error': 'Sin lote activo'}), 400
    if not has_permission('user.invite'):
        return jsonify({'error': 'Sin permiso para invitar'}), 403

    data = request.get_json(force=True) or {}
    email_inv = (data.get('email') or '').strip().lower()
    rol_nombre = data.get('rol', 'operador_lote')

    if not email_inv or not EMAIL_REGEX.match(email_inv):
        return jsonify({'error': 'Correo inválido'}), 400

    lote_id = session['lote_id']
    user_id = session['user_id']
    try:
        conn = get_db_connection()
        cursor = conn.cursor(dictionary=True)
        cursor.execute("SELECT id FROM roles WHERE nombre=%s", (rol_nombre,))
        role_row = cursor.fetchone()
        if not role_row:
            cursor.close(); conn.close()
            return jsonify({'error': 'Rol no válido'}), 400
        token = secrets.token_urlsafe(32)
        expires_at = datetime.utcnow() + timedelta(days=7)
        cursor2 = conn.cursor()
        cursor2.execute("""
            INSERT INTO lote_invitations (lote_id, email, role_id, token, invited_by, expires_at)
            VALUES (%s,%s,%s,%s,%s,%s)
        """, (lote_id, email_inv, role_row['id'], token, user_id, expires_at))
        conn.commit()
        cursor.close(); cursor2.close(); conn.close()
        invite_url = f"{APP_URL}/invitacion/{token}"
        return jsonify({'success': True, 'invite_url': invite_url})
    except Exception as e:
        return jsonify({'error': str(e)}), 500


@app.route('/invitacion/<token>', methods=['GET', 'POST'])
def aceptar_invitacion(token):
    try:
        conn = get_db_connection()
        cursor = conn.cursor(dictionary=True)
        cursor.execute("""
            SELECT li.*, l.nombre as lote_nombre, r.nombre as rol_nombre
            FROM lote_invitations li
            JOIN lotes l ON li.lote_id = l.id
            JOIN roles r ON li.role_id = r.id
            WHERE li.token = %s
        """, (token,))
        inv = cursor.fetchone()
        cursor.close(); conn.close()
    except Exception as e:
        return render_auth_page('login', f'Error al procesar invitación: {e}', 'danger')

    if not inv:
        return render_auth_page('login', 'Invitación no válida.', 'danger')
    if inv['used']:
        return render_auth_page('login', 'Esta invitación ya fue utilizada.', 'warning')
    if datetime.utcnow() > inv['expires_at']:
        return render_auth_page('login', 'Esta invitación ha expirado.', 'warning')

    if request.method == 'GET':
        return render_template('invitacion.html', inv=inv, token=token)

    if 'user_id' not in session:
        return render_auth_page('login', 'Inicia sesión para aceptar la invitación.', 'info')
    user_id = session['user_id']
    try:
        conn = get_db_connection()
        cursor = conn.cursor()
        cursor.execute(
            "INSERT IGNORE INTO user_lote_roles (user_id, lote_id, role_id, invited_by) VALUES (%s,%s,%s,%s)",
            (user_id, inv['lote_id'], inv['role_id'], inv['invited_by'])
        )
        cursor.execute("UPDATE lote_invitations SET used=TRUE WHERE token=%s", (token,))
        conn.commit()
        cursor.close(); conn.close()
        _set_active_lote_session(user_id, inv['lote_id'])
        return redirect(url_for('dashboard'))
    except Exception as e:
        return render_auth_page('login', f'Error al aceptar invitación: {e}', 'danger')


@app.route('/auth/login', methods=['POST'])
def login():
    email = (request.form.get('email') or '').strip().lower()
    password = request.form.get('password') or ''
    remember = request.form.get('remember_me') == 'on'

    current_form_data = {
        'login': {
            'email': email,
            'remember_me': remember
        }
    }

    if not email or not password:
        return render_auth_page('login', 'Ingresa correo y contrasena.', 'warning', current_form_data)

    try:
        conn = get_db_connection()
        cursor = conn.cursor(dictionary=True)
        cursor.execute(
            """
            SELECT id_user, full_name, password_hash, is_active, email_verified
            FROM users
            WHERE email = %s
            """,
            (email,)
        )
        user = cursor.fetchone()
        cursor.close()
        conn.close()
    except mysql.connector.Error as err:
        return render_auth_page('login', f'Error de acceso: {err}', 'danger', current_form_data)

    if not user or not check_password_hash(user['password_hash'], password):
        return render_auth_page('login', 'Credenciales invalidas.', 'danger', current_form_data)

    if not user['is_active']:
        return render_auth_page('login', 'Tu cuenta esta desactivada.', 'warning', current_form_data)

    # email_verified may be None for pre-existing rows (treat as unverified)
    if not user.get('email_verified'):
        return render_auth_page(
            'login',
            f'Debes verificar tu correo electronico antes de iniciar sesion. '
            f'<a href="/auth/resend-verification?email={email}" class="alert-link">Reenviar correo</a>',
            'warning',
            current_form_data
        )

    session.clear()
    session['user_id'] = user['id_user']
    session['user_name'] = user['full_name']
    session.permanent = remember

    if AUTH_SEND_LOGIN_ALERT:
        try:
            when_text = datetime.now().strftime('%d/%m/%Y %H:%M:%S')
            login_html = render_login_alert_email(user['full_name'], when_text)
            send_email(email, 'Alerta de inicio de sesion - Contabilidad Arroceras', login_html)
        except Exception:
            pass

    # ── Selección de lote post-login ──────────────────────────────────────────
    lotes = _get_user_lotes(user['id_user'])
    if len(lotes) == 0:
        # Usuario nuevo sin lotes → wizard de configuración
        return redirect(url_for('setup_lote_nuevo'))
    elif len(lotes) == 1:
        # Un solo lote → selección automática
        _set_active_lote_session(user['id_user'], lotes[0]['id'])
        return redirect(url_for('dashboard'))
    else:
        # Múltiples lotes → mostrar selector
        return redirect(url_for('select_lote'))


@app.route('/auth/forgot-password', methods=['POST'])
def forgot_password():
    email = (request.form.get('email') or '').strip().lower()

    current_form_data = {'forgot': {'email': email}}

    if not email or not EMAIL_REGEX.match(email):
        return render_auth_page('forgot', 'Ingresa un correo valido para continuar.', 'warning', current_form_data)

    response_message = 'Si el correo esta registrado, recibiras un codigo de 6 digitos para restablecer tu contrasena.'

    try:
        conn = get_db_connection()
        cursor = conn.cursor(dictionary=True)
        cursor.execute("SELECT id_user, full_name FROM users WHERE email = %s", (email,))
        user = cursor.fetchone()

        if user:
            code = generate_6_digit_code()
            save_auth_code(
                email,
                'reset',
                code,
                {'user_id': user['id_user'], 'full_name': user['full_name'], 'email': email}
            )
            html_body = render_reset_code_email(user['full_name'], code)
            send_email(email, 'Codigo de recuperacion - Contabilidad Arroceras', html_body)

        cursor.close()
        conn.close()
    except mysql.connector.Error as err:
        return render_auth_page('forgot', f'Error al procesar solicitud: {err}', 'danger', current_form_data)

    return render_auth_page(
        'reset_code',
        response_message,
        'success',
        {'reset_code': {'email': email, 'code': ''}}
    )


@app.route('/auth/reset-with-code', methods=['POST'])
def reset_with_code():
    email = (request.form.get('email') or '').strip().lower()
    code = re.sub(r'\D', '', request.form.get('code') or '')
    password = request.form.get('password') or ''
    confirm = request.form.get('confirm_password') or ''

    current_form_data = {'reset_code': {'email': email, 'code': code}}

    if not EMAIL_REGEX.match(email):
        return render_auth_page('reset_code', 'Correo no valido.', 'warning', current_form_data)
    if len(code) != 6:
        return render_auth_page('reset_code', 'El codigo debe tener 6 digitos.', 'warning', current_form_data)
    if not password or password != confirm:
        return render_auth_page('reset_code', 'Las contrasenas no coinciden.', 'warning', current_form_data)
    if not is_valid_password(password):
        return render_auth_page(
            'reset_code',
            'La contrasena debe tener minimo 8 caracteres, mayuscula, minuscula y numero.',
            'warning',
            current_form_data
        )

    ok, err_msg, payload = consume_auth_code(email, 'reset', code)
    if not ok:
        return render_auth_page('reset_code', err_msg, 'warning', current_form_data)

    user_id = payload.get('user_id')
    if not user_id:
        return render_auth_page('reset_code', 'No se encontro un usuario valido para este codigo.', 'danger', current_form_data)

    try:
        conn = get_db_connection()
        cursor = conn.cursor()
        cursor.execute(
            "UPDATE users SET password_hash = %s WHERE id_user = %s",
            (generate_password_hash(password), user_id)
        )
        conn.commit()
        cursor.close(); conn.close()
    except mysql.connector.Error as err:
        return render_auth_page('reset_code', f'Error al actualizar contrasena: {err}', 'danger', current_form_data)

    try:
        changed_html = render_password_changed_email(payload.get('full_name') or 'Usuario')
        send_email(email, 'Contrasena actualizada - Contabilidad Arroceras', changed_html)
    except Exception:
        pass

    return render_auth_page('login', 'Contrasena actualizada correctamente. Ya puedes iniciar sesion.', 'success')


@app.route('/auth/logout')
def logout():
    session.clear()
    return auth_redirect('login', 'Sesion cerrada correctamente.', 'success')


# ── Email verification ────────────────────────────────────────────────────────
@app.route('/auth/verify-email/<token>')
def verify_email(token):
    try:
        conn = get_db_connection()
        cursor = conn.cursor(dictionary=True)
        cursor.execute(
            "SELECT id_user, email_verified FROM users WHERE verify_token = %s",
            (token,)
        )
        user = cursor.fetchone()

        if not user:
            cursor.close()
            conn.close()
            return render_auth_page('login',
                                    'Enlace de verificacion invalido o ya utilizado.',
                                    'danger')

        if user['email_verified']:
            cursor.close()
            conn.close()
            return render_auth_page('login',
                                    'Tu correo ya fue verificado. Inicia sesion.',
                                    'info')

        cursor.execute(
            "UPDATE users SET email_verified = TRUE, verify_token = NULL WHERE id_user = %s",
            (user['id_user'],)
        )
        conn.commit()
        cursor.close()
        conn.close()
    except mysql.connector.Error as err:
        return render_auth_page('login', f'Error de verificacion: {err}', 'danger')

    return render_auth_page('login',
                            '¡Correo verificado correctamente! Ya puedes iniciar sesion.',
                            'success')


@app.route('/auth/resend-verification')
def resend_verification():
    email = (request.args.get('email') or '').strip().lower()
    if not email or not EMAIL_REGEX.match(email):
        return auth_redirect('login', 'Correo no valido.', 'warning')

    try:
        conn = get_db_connection()
        cursor = conn.cursor(dictionary=True)
        cursor.execute(
            "SELECT id_user, full_name, email_verified FROM users WHERE email = %s",
            (email,)
        )
        user = cursor.fetchone()

        if user and not user['email_verified']:
            code = generate_6_digit_code()
            payload = {
                'full_name': user['full_name'],
                'email': email,
                'password_hash': '',
            }
            save_auth_code(email, 'signup', code, payload)
            html_body = render_signup_code_email(user['full_name'], code)
            send_email(email, 'Codigo de verificacion - Contabilidad Arroceras', html_body)

        cursor.close()
        conn.close()
    except mysql.connector.Error:
        pass  # Silent — don't reveal user existence

    return render_auth_page(
        'signup_code',
        f'Si la cuenta existe y no esta verificada, se reenviara un codigo de seguridad a {email}.',
        'success',
        {'signup_code': {'email': email, 'code': ''}}
    )


# ── Password reset ────────────────────────────────────────────────────────────
@app.route('/auth/reset-password/<token>', methods=['GET', 'POST'])
def reset_password(token):
    try:
        conn = get_db_connection()
        cursor = conn.cursor(dictionary=True)
        cursor.execute(
            """
            SELECT prt.id, prt.user_id, prt.expires_at, prt.used, u.full_name, u.email
            FROM password_reset_tokens prt
            JOIN users u ON u.id_user = prt.user_id
            WHERE prt.token = %s
            """,
            (token,)
        )
        row = cursor.fetchone()
    except mysql.connector.Error as err:
        return render_template('reset_password.html', error=f'Error de base de datos: {err}', token=token, valid=False)

    if not row:
        cursor.close(); conn.close()
        return render_template('reset_password.html',
                               error='Enlace invalido o ya utilizado.',
                               token=token, valid=False)

    if row['used']:
        cursor.close(); conn.close()
        return render_template('reset_password.html',
                               error='Este enlace ya fue utilizado.',
                               token=token, valid=False)

    if datetime.utcnow() > row['expires_at']:
        cursor.close(); conn.close()
        return render_template('reset_password.html',
                               error='El enlace ha expirado. Solicita uno nuevo.',
                               token=token, valid=False)

    if request.method == 'GET':
        cursor.close(); conn.close()
        return render_template('reset_password.html',
                               token=token, valid=True,
                               full_name=row['full_name'])

    # POST — process new password
    password = request.form.get('password') or ''
    confirm  = request.form.get('confirm_password') or ''

    if not password or password != confirm:
        cursor.close(); conn.close()
        return render_template('reset_password.html',
                               error='Las contrasenas no coinciden.',
                               token=token, valid=True,
                               full_name=row['full_name'])

    if not is_valid_password(password):
        cursor.close(); conn.close()
        return render_template('reset_password.html',
                               error='La contrasena debe tener minimo 8 caracteres, mayuscula, minuscula y numero.',
                               token=token, valid=True,
                               full_name=row['full_name'])

    try:
        cursor.execute(
            "UPDATE users SET password_hash = %s WHERE id_user = %s",
            (generate_password_hash(password), row['user_id'])
        )
        cursor.execute(
            "UPDATE password_reset_tokens SET used = TRUE WHERE id = %s",
            (row['id'],)
        )
        conn.commit()
    except mysql.connector.Error as err:
        cursor.close(); conn.close()
        return render_template('reset_password.html',
                               error=f'Error al actualizar: {err}',
                               token=token, valid=True)

    try:
        changed_html = render_password_changed_email(row['full_name'])
        send_email(row['email'], 'Contrasena actualizada - Contabilidad Arroceras', changed_html)
    except Exception:
        pass

    cursor.close(); conn.close()
    return render_auth_page('login',
                            '¡Contrasena actualizada! Ya puedes iniciar sesion con tu nueva contrasena.',
                            'success')

# =========================
# WORKERS
# =========================
@app.route('/workers/create', methods=['GET', 'POST'])
def create_worker():
    from auth_middleware import has_permission
    if 'user_id' not in session:
        return auth_redirect('login', 'Inicia sesion para acceder al registro de trabajadores.', 'warning')
    if not session.get('lote_id'):
        return redirect(url_for('select_lote'))
    if not has_permission('worker.create'):
        return redirect(url_for('lista_workers'))
    lote_id = session['lote_id']

    message = None
    error = None

    if request.method == 'POST':
        name = request.form.get('name', '').strip()
        lastname = request.form.get('lastname', '').strip()
        cc = request.form.get('cc', '').strip()
        phone_number = request.form.get('phone_number', '').strip()
        email = request.form.get('email') or None
        trabajo = request.form.get('trabajo_desarrolla')
        fecha_ingreso = request.form.get('fecha_ingreso') or None
        activo = request.form.get('activo') == 'on'
        observaciones = request.form.get('observaciones') or None
        alias = request.form.get('alias') or None
        direccion = request.form.get('direccion') or None
        ciudad = request.form.get('ciudad') or None
        concepto_habitual = request.form.get('concepto_habitual') or None
        valor_habitual_raw = request.form.get('valor_habitual') or None
        valor_habitual = float(valor_habitual_raw) if valor_habitual_raw else None

        foto_filename = None
        foto_file = request.files.get('foto')
        if foto_file and foto_file.filename and allowed_file(foto_file.filename):
            ext = foto_file.filename.rsplit('.', 1)[1].lower()
            foto_filename = f"{uuid.uuid4().hex}.{ext}"
            foto_file.save(os.path.join(UPLOAD_FOLDER, foto_filename))

        if not name or not lastname or not cc or not phone_number:
            error = "Completa los campos obligatorios: nombre, apellido, CC/NIT y teléfono."
        else:
            try:
                conn = get_db_connection()
                cursor = conn.cursor()
                cursor.execute("""
                    INSERT INTO workers
                    (name, lastname, cc, phone_number, email, trabajo_desarrolla, fecha_ingreso, activo, observaciones,
                     foto, alias, direccion, ciudad, concepto_habitual, valor_habitual, lote_id)
                    VALUES (%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s)
                """, (name, lastname, cc, phone_number, email, trabajo, fecha_ingreso, activo, observaciones,
                      foto_filename, alias, direccion, ciudad, concepto_habitual, valor_habitual, lote_id))
                conn.commit()
                cursor.close()
                conn.close()
                message = f"Trabajador {name} {lastname} registrado correctamente."
            except Exception as e:
                error = f"Error: {e}"

    return render_template('workers/create.html', message=message, error=error,
                           options=WORKER_OPTIONS, today=date.today().isoformat())

# =========================
# API TRABAJADORES
# =========================
def _get_trabajadores_for_autocomplete():
    """Returns active workers for the current lote in autocomplete-ready format."""
    lote_id = session.get('lote_id')
    try:
        conn = get_db_connection()
        cursor = conn.cursor(dictionary=True)
        if lote_id:
            cursor.execute("""
                SELECT id_worker, name, lastname, cc, phone_number, alias,
                       direccion, ciudad, concepto_habitual, valor_habitual,
                       rol, conceptos_pago
                FROM workers WHERE activo = 1 AND lote_id = %s ORDER BY name
            """, (lote_id,))
        else:
            cursor.execute("""
                SELECT id_worker, name, lastname, cc, phone_number, alias,
                       direccion, ciudad, concepto_habitual, valor_habitual,
                       rol, conceptos_pago
                FROM workers WHERE activo = 1 ORDER BY name
            """)
        rows = cursor.fetchall()
        cursor.close(); conn.close()
    except Exception:
        rows = []
    result = []
    for w in rows:
        alias_str = (w.get('alias') or '').strip()
        alias_list = [a.strip() for a in alias_str.split(',') if a.strip()] if alias_str else []
        cpago = []
        if w.get('conceptos_pago'):
            try:
                cpago = json.loads(w['conceptos_pago'])
            except Exception:
                pass
        result.append({
            'id':                w['id_worker'],
            'nombre':            f"{w['name']} {w['lastname']}",
            'nit':               w.get('cc') or '',
            'alias':             alias_list,
            'telefono':          w.get('phone_number') or '',
            'direccion':         w.get('direccion') or '',
            'ciudad':            w.get('ciudad') or '',
            'rol':               w.get('rol') or '',
            'concepto_habitual': w.get('concepto_habitual') or '',
            'valor_habitual':    float(w['valor_habitual']) if w.get('valor_habitual') else None,
            'conceptos_pago':    cpago,
        })
    return result


@app.route('/api/trabajadores')
def api_trabajadores():
    """Returns workers from DB normalized for autocomplete (nombre, nit, alias[], ...)."""
    q = request.args.get('q', '').strip().lower()
    todos = _get_trabajadores_for_autocomplete()
    if q:
        todos = [t for t in todos if q in t['nombre'].lower()
                 or any(q in a.lower() for a in t.get('alias', []))]
    return jsonify(todos)


def _build_recibo_form_data(recibo):
    lineas = []
    if recibo.get('conceptos_json'):
        try:
            lineas = json.loads(recibo['conceptos_json'])
        except Exception:
            lineas = []
    if not lineas and recibo.get('concepto'):
        lineas = [{
            'concepto': recibo.get('concepto') or '',
            'valor': float(recibo.get('valor_operacion') or recibo.get('neto_a_pagar') or 0),
        }]
    return {
        'serial': recibo.get('serial', ''),
        'fecha': recibo['fecha'].isoformat() if getattr(recibo.get('fecha'), 'isoformat', None) else '',
        'proveedor': recibo.get('proveedor') or '',
        'nit': recibo.get('nit') or '',
        'direccion': recibo.get('direccion') or '',
        'telefono': recibo.get('telefono') or '',
        'ciudad': recibo.get('ciudad') or '',
        'lineas': lineas,
        'rte_fte': float(recibo.get('rte_fte') or 0),
        'neto_a_pagar': float(recibo.get('neto_a_pagar') or 0),
    }

# =========================
# RECIBOS
# =========================
@app.route('/recibos')
def lista_recibos():
    from auth_middleware import require_permission, has_permission
    if 'user_id' not in session:
        return auth_redirect('login', 'Inicia sesion para ver los recibos.', 'warning')
    if not session.get('lote_id'):
        return redirect(url_for('select_lote'))
    if not has_permission('recibo.view'):
        return redirect(url_for('dashboard', msg='Sin permiso para ver recibos.', msg_type='danger'))
    lote_id = session['lote_id']
    error = None
    try:
        conn = get_db_connection()
        cursor = conn.cursor(dictionary=True)
        cursor.execute("SELECT * FROM recibos WHERE lote_id=%s ORDER BY serial", (lote_id,))
        recibos = cursor.fetchall()
        cursor.close(); conn.close()
        for r in recibos:
            r['neto_a_pagar'] = float(r['neto_a_pagar'] or 0)
            r['valor_operacion'] = float(r['valor_operacion'] or 0)
    except Exception as e:
        recibos = []
        error = f"Error al cargar recibos: {e}"
    return render_template('recibos/lista.html', recibos=recibos, error=error,
                           lote_nombre=session.get('lote_nombre', f'Lote {lote_id}'))


@app.route('/recibos/nuevo', methods=['GET', 'POST'])
def nuevo_recibo():
    from auth_middleware import has_permission
    if 'user_id' not in session:
        return auth_redirect('login', 'Inicia sesion para crear recibos.', 'warning')
    if not session.get('lote_id'):
        return redirect(url_for('select_lote'))
    if request.method == 'POST' and not has_permission('recibo.create'):
        return redirect(url_for('lista_recibos'))

    lote_id = session['lote_id']
    next_serial = get_next_serial(lote_id)
    error = None
    warning = None
    success = None
    form_data = {}

    if request.method == 'POST':
        serial = request.form.get('serial', '').strip()
        fecha_str = request.form.get('fecha', '').strip()
        proveedor = request.form.get('proveedor', '').strip()
        nit = request.form.get('nit', '').strip()
        direccion_r = request.form.get('direccion', '').strip()
        telefono = request.form.get('telefono', '').strip()
        ciudad_r = request.form.get('ciudad', '').strip()
        force = request.form.get('force') == 'true'

        # Parse multi-concept lines (up to 7 rows)
        lineas = []
        for i in range(1, 8):
            c_text = request.form.get(f'concepto_{i}', '').strip()
            c_val_raw = request.form.get(f'valor_{i}', '').strip().replace('.', '').replace(',', '')
            if c_text:
                lineas.append({'concepto': c_text, 'valor': float(c_val_raw) if c_val_raw else 0.0})

        concepto = '\n'.join(l['concepto'] for l in lineas)
        valor_op = sum(l['valor'] for l in lineas)
        conceptos_json_str = json.dumps(lineas, ensure_ascii=False)
        rte_raw = request.form.get('rte_fte', '').strip().replace('.', '').replace(',', '')
        rte_fte = float(rte_raw) if rte_raw else 0.0
        neto_raw = request.form.get('neto_a_pagar', '').strip().replace('.', '').replace(',', '')

        form_data = {
            'serial': serial, 'fecha': fecha_str, 'proveedor': proveedor,
            'nit': nit, 'direccion': direccion_r, 'telefono': telefono,
            'ciudad': ciudad_r, 'lineas': lineas, 'rte_fte': rte_fte,
        }

        if not serial or not proveedor or not lineas:
            error = "Serial, proveedor y al menos un concepto son obligatorios."
        elif not serial.isdigit():
            error = "El serial debe ser un número entero."
        elif any(es_concepto_prohibido(l['concepto']) for l in lineas):
            error = "⚠️ No se registran compras de aceite ni ACPM directo. Solo se registra el transporte de ACPM."
        else:
            serial_int = int(serial)
            fecha_obj = None
            if fecha_str:
                try:
                    fecha_obj = date.fromisoformat(fecha_str)
                except ValueError:
                    error = "Formato de fecha inválido."

            if not error:
                # neto = subtotal - rte_fte, or user override
                neto = float(neto_raw) if neto_raw else max(0.0, valor_op - rte_fte)

                # Check serial-date order (only if date provided and not forcing)
                if fecha_obj and not force:
                    try:
                        conn = get_db_connection()
                        cursor = conn.cursor(dictionary=True)
                        cursor.execute("SELECT serial, fecha FROM recibos WHERE serial < %s AND lote_id=%s ORDER BY serial DESC LIMIT 1", (serial_int, lote_id))
                        prev = cursor.fetchone()
                        cursor.execute("SELECT serial, fecha FROM recibos WHERE serial > %s AND lote_id=%s ORDER BY serial ASC LIMIT 1", (serial_int, lote_id))
                        next_r = cursor.fetchone()
                        cursor.close()
                        conn.close()

                        order_ok = True
                        msg_parts = []
                        if prev and prev['fecha'] and prev['fecha'] > fecha_obj:
                            order_ok = False
                            msg_parts.append(f"el recibo anterior (serial {prev['serial']}) tiene fecha {prev['fecha'].strftime('%d/%m/%Y')}")
                        if next_r and next_r['fecha'] and next_r['fecha'] < fecha_obj:
                            order_ok = False
                            msg_parts.append(f"el recibo siguiente (serial {next_r['serial']}) tiene fecha {next_r['fecha'].strftime('%d/%m/%Y')}")

                        if not order_ok:
                            warning = f"⚠️ Incongruencia serial-fecha: {'; '.join(msg_parts)}. ¿Deseas guardar de todas formas?"
                    except Exception:
                        pass  # Don't block if validation fails

                if not warning or force:
                    try:
                        conn = get_db_connection()
                        cursor = conn.cursor()
                        # If serial already exists, shift serials >= serial_int by +1
                        cursor.execute("SELECT serial FROM recibos WHERE serial = %s", (serial_int,))
                        exists = cursor.fetchone()
                        if exists:
                            cursor.execute("UPDATE recibos SET serial = serial + 1 WHERE serial >= %s ORDER BY serial DESC", (serial_int,))

                        cursor.execute("""
                            INSERT INTO recibos (serial, fecha, proveedor, nit, direccion, telefono, ciudad, concepto, valor_operacion, rte_fte, neto_a_pagar, conceptos_json, lote_id)
                            VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s)
                        """, (serial_int, fecha_obj, proveedor, nit, direccion_r, telefono, ciudad_r, concepto, valor_op, rte_fte, neto, conceptos_json_str, lote_id))
                        conn.commit()
                        cursor.close()
                        conn.close()
                        success = f"Recibo serial {serial_int} guardado correctamente."
                        form_data = {}
                        next_serial = get_next_serial(lote_id)
                    except Exception as e:
                        error = f"Error al guardar: {e}"

    trabajadores = _get_trabajadores_for_autocomplete()
    return render_template('recibos/nuevo.html',
                           next_serial=next_serial,
                           trabajadores=trabajadores,
                           error=error,
                           warning=warning,
                           success=success,
                           form_data=form_data,
                           page_title='Nuevo Recibo',
                           page_subtitle='Registra un pago. El serial se asigna según la fecha.',
                           submit_label='Guardar recibo',
                           serial_readonly=False,
                           page_mode='create',
                           today=date.today().isoformat())

# =========================
# WORKERS LIST
# =========================
@app.route('/workers')
def lista_workers():
    from auth_middleware import has_permission
    if 'user_id' not in session:
        return auth_redirect('login', 'Inicia sesion.', 'warning')
    if not session.get('lote_id'):
        return redirect(url_for('select_lote'))
    if not has_permission('worker.view'):
        return redirect(url_for('dashboard', msg='Sin permiso para ver trabajadores.', msg_type='danger'))
    lote_id = session['lote_id']
    conn = get_db_connection()
    cursor = conn.cursor(dictionary=True)
    cursor.execute("SELECT * FROM workers WHERE lote_id=%s ORDER BY name, lastname", (lote_id,))
    workers = cursor.fetchall()
    cursor.close(); conn.close()
    return render_template('workers/lista.html', workers=workers)


@app.route('/workers/<int:wid>/edit', methods=['GET', 'POST'])
def edit_worker(wid):
    from auth_middleware import has_permission
    if 'user_id' not in session:
        return auth_redirect('login', 'Inicia sesion.', 'warning')
    if not session.get('lote_id'):
        return redirect(url_for('select_lote'))
    lote_id = session['lote_id']
    conn = get_db_connection()
    cursor = conn.cursor(dictionary=True)
    # Seguridad: solo workers del lote activo
    cursor.execute("SELECT * FROM workers WHERE id_worker = %s AND lote_id = %s", (wid, lote_id))
    worker = cursor.fetchone()
    if not worker:
        cursor.close(); conn.close()
        return redirect(url_for('lista_workers'))
    message = None
    error = None
    if request.method == 'POST':
        name              = request.form.get('name', '').strip()
        lastname          = request.form.get('lastname', '').strip()
        cc                = request.form.get('cc', '').strip()
        phone_number      = request.form.get('phone_number', '').strip()
        email             = request.form.get('email') or None
        trabajo           = request.form.get('trabajo_desarrolla')
        fecha_ingreso     = request.form.get('fecha_ingreso') or None
        activo            = 1 if request.form.get('activo') == 'on' else 0
        observaciones     = request.form.get('observaciones') or None
        alias             = request.form.get('alias') or None
        direccion         = request.form.get('direccion') or None
        ciudad            = request.form.get('ciudad') or None
        concepto_habitual = request.form.get('concepto_habitual') or None
        vh_raw            = request.form.get('valor_habitual') or None
        valor_habitual    = float(vh_raw.replace('.','').replace(',','')) if vh_raw else None

        foto_filename = worker['foto']
        foto_file = request.files.get('foto')
        if foto_file and foto_file.filename and allowed_file(foto_file.filename):
            # Delete old photo if exists
            if foto_filename:
                old = os.path.join(UPLOAD_FOLDER, foto_filename)
                if os.path.exists(old):
                    os.remove(old)
            ext = foto_file.filename.rsplit('.', 1)[1].lower()
            foto_filename = f"{uuid.uuid4().hex}.{ext}"
            foto_file.save(os.path.join(UPLOAD_FOLDER, foto_filename))

        if not name or not lastname or not cc:
            error = "Nombre, apellido y CC/NIT son obligatorios."
        else:
            try:
                cursor2 = conn.cursor()
                cursor2.execute("""
                    UPDATE workers SET name=%s, lastname=%s, cc=%s, phone_number=%s, email=%s,
                    trabajo_desarrolla=%s, fecha_ingreso=%s, activo=%s, observaciones=%s,
                    foto=%s, alias=%s, direccion=%s, ciudad=%s,
                    concepto_habitual=%s, valor_habitual=%s
                    WHERE id_worker=%s
                """, (name, lastname, cc, phone_number, email, trabajo, fecha_ingreso, activo,
                      observaciones, foto_filename, alias, direccion, ciudad,
                      concepto_habitual, valor_habitual, wid))
                conn.commit()
                cursor2.close()
                message = "Trabajador actualizado correctamente."
                # Reload updated data
                cursor.execute("SELECT * FROM workers WHERE id_worker = %s", (wid,))
                worker = cursor.fetchone()
            except Exception as e:
                error = f"Error al actualizar: {e}"
    cursor.close(); conn.close()
    return render_template('workers/edit.html', worker=worker, message=message, error=error,
                           options=WORKER_OPTIONS)

@app.route('/workers/<int:wid>/toggle', methods=['POST'])
def toggle_worker(wid):
    from auth_middleware import has_permission
    if 'user_id' not in session:
        return auth_redirect('login', 'Inicia sesion.', 'warning')
    lote_id = session.get('lote_id')
    try:
        conn = get_db_connection()
        cursor = conn.cursor()
        # Solo permite toggle si el worker pertenece al lote activo
        if lote_id:
            cursor.execute("UPDATE workers SET activo = NOT activo WHERE id_worker = %s AND lote_id = %s", (wid, lote_id))
        else:
            cursor.execute("UPDATE workers SET activo = NOT activo WHERE id_worker = %s", (wid,))
        conn.commit()
        cursor.close(); conn.close()
    except Exception:
        pass
    return redirect(url_for('lista_workers'))

@app.route('/api/workers')
def api_workers():
    lote_id = session.get('lote_id')
    conn = get_db_connection()
    cursor = conn.cursor(dictionary=True)
    if lote_id:
        cursor.execute("SELECT id_worker, name, lastname, cc, direccion, ciudad, phone_number, concepto_habitual, valor_habitual, trabajo_desarrolla FROM workers WHERE activo = 1 AND lote_id = %s ORDER BY name", (lote_id,))
    else:
        cursor.execute("SELECT id_worker, name, lastname, cc, direccion, ciudad, phone_number, concepto_habitual, valor_habitual, trabajo_desarrolla FROM workers WHERE activo = 1 ORDER BY name")
    workers = cursor.fetchall()
    cursor.close(); conn.close()
    for w in workers:
        if w.get('valor_habitual'):
            w['valor_habitual'] = float(w['valor_habitual'])
    return jsonify(workers)

# =========================
# RECIBO POR LOTE
# =========================
@app.route('/recibos/lote', methods=['GET', 'POST'])
def nuevo_recibo_lote():
    from auth_middleware import has_permission
    if 'user_id' not in session:
        return auth_redirect('login', 'Inicia sesion.', 'warning')
    if not session.get('lote_id'):
        return redirect(url_for('select_lote'))
    lote_id = session['lote_id']

    error = None
    success = None
    next_serial = get_next_serial(lote_id)

    if request.method == 'POST':
        fecha_str = request.form.get('fecha', '').strip()
        direccion_r = request.form.get('direccion', '').strip()
        ciudad_r = request.form.get('ciudad', '').strip()
        valor_por_trabajador = request.form.get('valor_por_trabajador', '').strip().replace('.','').replace(',','')
        serial_inicio = request.form.get('serial_inicio', '').strip()
        worker_ids = request.form.getlist('worker_ids')

        # Parse multi-concept lines
        lineas_lote = []
        for i in range(1, 8):
            c_text = request.form.get(f'concepto_{i}', '').strip()
            c_val_raw = request.form.get(f'valor_{i}', '').strip().replace('.', '').replace(',', '')
            if c_text:
                lineas_lote.append({'concepto': c_text, 'valor': float(c_val_raw) if c_val_raw else 0.0})
        concepto = '\n'.join(l['concepto'] for l in lineas_lote)
        conceptos_json_lote = json.dumps(lineas_lote, ensure_ascii=False)

        if not worker_ids:
            error = "Selecciona al menos un trabajador."
        elif not lineas_lote:
            error = "El concepto es obligatorio."
        elif not serial_inicio or not serial_inicio.isdigit():
            error = "El serial de inicio debe ser un número."
        elif any(es_concepto_prohibido(l['concepto']) for l in lineas_lote):
            error = "⚠️ Las compras de aceite o ACPM directo no se registran."
        else:
            fecha_obj = None
            if fecha_str:
                try:
                    fecha_obj = date.fromisoformat(fecha_str)
                except ValueError:
                    error = "Formato de fecha inválido."

            if not error:
                try:
                    conn = get_db_connection()
                    cursor = conn.cursor(dictionary=True)
                    placeholders = ','.join(['%s'] * len(worker_ids))
                    cursor.execute(f"SELECT id_worker, name, lastname, cc, direccion, ciudad, phone_number as telefono, concepto_habitual, valor_habitual FROM workers WHERE id_worker IN ({placeholders})", worker_ids)
                    selected_workers = cursor.fetchall()

                    serial_int = int(serial_inicio)
                    valor_lote = sum(l['valor'] for l in lineas_lote)
                    valor_float = float(valor_por_trabajador) if valor_por_trabajador else None

                    created = 0
                    for w in selected_workers:
                        wid_str = str(w['id_worker'])
                        proveedor_name = f"{w['name']} {w['lastname']}"
                        nit = w['cc']
                        dir_w = direccion_r or w.get('direccion') or ''
                        ciudad_w = ciudad_r or w.get('ciudad') or ''
                        tel_w = w.get('telefono') or ''

                        # Per-worker value override from individual day inputs
                        pworker_val_raw = request.form.get(f'valor_w_{wid_str}', '').strip()
                        if pworker_val_raw and pworker_val_raw not in ('0', '0.0'):
                            val = float(pworker_val_raw)
                            # Update first line value with individual amount
                            lineas_w = ([{'concepto': lineas_lote[0]['concepto'], 'valor': val}]
                                        + lineas_lote[1:]) if lineas_lote else [{'concepto': concepto, 'valor': val}]
                        elif valor_float:
                            lineas_w = [lineas_lote[0] | {'valor': valor_float}] + lineas_lote[1:]
                            val = valor_float
                        else:
                            lineas_w = lineas_lote
                            val = valor_lote or (float(w['valor_habitual']) if w.get('valor_habitual') else None)
                        cj = json.dumps(lineas_w, ensure_ascii=False)

                        cursor2 = conn.cursor()
                        cursor2.execute("SELECT serial FROM recibos WHERE serial = %s", (serial_int,))
                        if cursor2.fetchone():
                            cursor2.execute("UPDATE recibos SET serial = serial + 1 WHERE serial >= %s ORDER BY serial DESC", (serial_int,))
                        cursor2.execute("""
                            INSERT INTO recibos (serial, fecha, proveedor, nit, direccion, telefono, ciudad, concepto, valor_operacion, rte_fte, neto_a_pagar, conceptos_json, lote_id)
                            VALUES (%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s)
                        """, (serial_int, fecha_obj, proveedor_name, nit, dir_w, tel_w, ciudad_w, concepto, val, 0, val, cj, lote_id))
                        cursor2.close()
                        serial_int += 1
                        created += 1

                    conn.commit()
                    cursor.close(); conn.close()
                    success = f"✅ {created} recibo(s) creados correctamente (seriales {serial_inicio}–{serial_int-1})."
                    next_serial = get_next_serial(lote_id)
                except Exception as e:
                    error = f"Error: {e}"

    try:
        conn = get_db_connection()
        cursor = conn.cursor(dictionary=True)
        lote_id_fetch = session.get('lote_id')
        if lote_id_fetch:
            cursor.execute("SELECT id_worker, name, lastname, cc, trabajo_desarrolla, concepto_habitual, valor_habitual, foto FROM workers WHERE activo=1 AND lote_id=%s ORDER BY trabajo_desarrolla, name", (lote_id_fetch,))
        else:
            cursor.execute("SELECT id_worker, name, lastname, cc, trabajo_desarrolla, concepto_habitual, valor_habitual, foto FROM workers WHERE activo=1 ORDER BY trabajo_desarrolla, name")
        db_workers = cursor.fetchall()
        cursor.close(); conn.close()
        for w in db_workers:
            if w.get('valor_habitual'):
                w['valor_habitual'] = float(w['valor_habitual'])
            w['nombre_completo'] = f"{w.get('name','')} {w.get('lastname','')}".strip()
            w['cargo'] = w.get('trabajo_desarrolla') or 'operario'
            w['foto_path'] = w.get('foto') or ''
    except Exception:
        db_workers = []

    trabajadores_json = load_trabajadores()
    return render_template('recibos/lote.html',
                           db_workers=db_workers,
                           trabajadores_json=trabajadores_json,
                           next_serial=next_serial,
                           error=error, success=success,
                           today=date.today().isoformat())

# =========================
# LABORES ESPECIALES: DESAGÜE / DESPALILLADA
# =========================
TIPOS_LABOR_DIAS = {
    'desague':      {'label': 'Desagüe',         'vdia_default': 60000,  'dias_default': 1, 'icon': 'fa-droplet'},
    'despalillada': {'label': 'Despalillada',     'vdia_default': 60000,  'dias_default': 1, 'icon': 'fa-hands'},
    'bordeada':     {'label': 'Bordeada',         'vdia_default': 80000,  'dias_default': 1, 'icon': 'fa-scissors'},
    'parche_maleza':{'label': 'Parche de maleza', 'vdia_default': 60000,  'dias_default': 1, 'icon': 'fa-leaf'},
    'chapola':      {'label': 'Chapola / rocería','vdia_default': 60000,  'dias_default': 1, 'icon': 'fa-tractor'},
    'otro':         {'label': 'Otro',             'vdia_default': 60000,  'dias_default': 1, 'icon': 'fa-ellipsis'},
}

def _load_workers_for_form():
    try:
        conn = get_db_connection()
        cursor = conn.cursor(dictionary=True)
        lote_id = session.get('lote_id')
        if lote_id:
            cursor.execute("""SELECT id_worker, name, lastname, cc, trabajo_desarrolla,
                             valor_habitual, direccion, ciudad, phone_number as telefono, foto
                             FROM workers WHERE activo=1 AND lote_id=%s ORDER BY trabajo_desarrolla, name""", (lote_id,))
        else:
            cursor.execute("""SELECT id_worker, name, lastname, cc, trabajo_desarrolla,
                             valor_habitual, direccion, ciudad, phone_number as telefono, foto
                             FROM workers WHERE activo=1 ORDER BY trabajo_desarrolla, name""")
        workers = cursor.fetchall()
        cursor.close(); conn.close()
        for w in workers:
            if w.get('valor_habitual'):
                w['valor_habitual'] = float(w['valor_habitual'])
            w['nombre_completo'] = f"{w.get('name','')} {w.get('lastname','')}".strip()
            w['cargo'] = w.get('trabajo_desarrolla') or 'operario'
            w['foto_path'] = w.get('foto') or ''
        return workers
    except Exception:
        return []

@app.route('/recibos/labores/desague', methods=['GET', 'POST'])
def labores_desague():
    if 'user_id' not in session:
        return auth_redirect('login', 'Inicia sesion.', 'warning')
    lote_id = session.get('lote_id')

    error = None
    success = None
    next_serial = get_next_serial(lote_id)

    if request.method == 'POST':
        tipo_labor    = request.form.get('tipo_labor', 'desague')
        fecha_str     = request.form.get('fecha', '').strip()
        serial_ini    = request.form.get('serial_inicio', '').strip()
        direccion_r   = request.form.get('direccion', '').strip()
        ciudad_r      = request.form.get('ciudad', '').strip()
        lote_nombre   = request.form.get('lote_nombre', 'El Mangon').strip()
        concepto_base = request.form.get('concepto_base', '').strip()
        worker_ids    = request.form.getlist('worker_ids')
        labor_info    = TIPOS_LABOR_DIAS.get(tipo_labor, TIPOS_LABOR_DIAS['otro'])

        if not worker_ids:
            error = "Selecciona al menos un trabajador."
        elif not serial_ini or not serial_ini.isdigit():
            error = "El serial de inicio debe ser un número válido."
        else:
            rte_pct = float(request.form.get('rte_fte', '0') or '0')
            fecha_obj = None
            if fecha_str:
                try:
                    fecha_obj = date.fromisoformat(fecha_str)
                except ValueError:
                    error = "Formato de fecha inválido."

            if not error:
                try:
                    conn = get_db_connection()
                    cursor = conn.cursor(dictionary=True)
                    placeholders = ','.join(['%s'] * len(worker_ids))
                    cursor.execute(f"SELECT id_worker, name, lastname, cc, direccion, ciudad, phone_number as telefono FROM workers WHERE id_worker IN ({placeholders}) AND lote_id=%s", (*worker_ids, lote_id))
                    selected_workers = cursor.fetchall()
                    serial_int = int(serial_ini)
                    created = 0

                    for w in selected_workers:
                        wid_str = str(w['id_worker'])
                        try:
                            dias = float(request.form.get(f'dias_w_{wid_str}', str(labor_info['dias_default'])) or labor_info['dias_default'])
                            frac = float(request.form.get(f'frac_w_{wid_str}', '0') or 0)
                            vdia = float((request.form.get(f'vdia_w_{wid_str}', str(labor_info['vdia_default'])) or str(labor_info['vdia_default'])).replace('.','').replace(',',''))
                        except Exception:
                            dias, frac, vdia = 1.0, 0.0, float(labor_info['vdia_default'])

                        total_dias = dias + frac
                        valor = total_dias * vdia
                        rte_amount = round(valor * rte_pct / 100)
                        neto = valor - rte_amount

                        frac_txt = {0.25: ' y ¼ día', 0.5: ' y ½ día', 0.75: ' y ¾ día'}.get(frac, '')
                        dias_int = int(dias)
                        dias_txt = f"{dias_int} {'día' if dias_int == 1 else 'días'}{frac_txt}"
                        concepto_txt = concepto_base or f"{labor_info['label']} {dias_txt} en el lote {lote_nombre}"

                        proveedor_name = f"{w['name']} {w['lastname']}"
                        dir_w    = direccion_r or w.get('direccion') or ''
                        ciudad_w = ciudad_r    or w.get('ciudad')    or ''
                        tel_w    = w.get('telefono') or ''
                        lineas   = [{'concepto': concepto_txt, 'valor': valor}]
                        cj       = json.dumps(lineas, ensure_ascii=False)

                        cursor2 = conn.cursor()
                        cursor2.execute("SELECT serial FROM recibos WHERE serial = %s", (serial_int,))
                        if cursor2.fetchone():
                            cursor2.execute("UPDATE recibos SET serial = serial + 1 WHERE serial >= %s ORDER BY serial DESC", (serial_int,))
                        cursor2.execute("""
                            INSERT INTO recibos (serial, fecha, proveedor, nit, direccion, telefono, ciudad, concepto, valor_operacion, rte_fte, neto_a_pagar, conceptos_json, lote_id)
                            VALUES (%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s)
                        """, (serial_int, fecha_obj, proveedor_name, w['cc'], dir_w, tel_w, ciudad_w, concepto_txt, valor, rte_pct, neto, cj, lote_id))
                        cursor2.close()
                        serial_int += 1
                        created += 1

                    conn.commit()
                    cursor.close(); conn.close()
                    success = f"✅ {created} recibo(s) creados para {labor_info['label']} (seriales {serial_ini}–{serial_int-1})."
                    next_serial = get_next_serial(lote_id)
                except Exception as e:
                    error = f"Error al guardar: {e}"

    db_workers = _load_workers_for_form()
    return render_template('recibos/labores_desague.html',
                           db_workers=db_workers,
                           tipos_labor=TIPOS_LABOR_DIAS,
                           next_serial=next_serial,
                           error=error, success=success,
                           today=date.today().isoformat())


# =========================
# LABORES ESPECIALES: ABONADA / FUMIGACIÓN
# =========================
TIPOS_LABOR_CANTIDAD = {
    'abonada':    {'label': 'Abonada',         'icon': 'fa-seedling',   'unit': 'bultos',     'vpunt_default': 13000},
    'fumigacion': {'label': 'Fumigación',       'icon': 'fa-spray-can', 'unit': 'jornales',   'vpunt_default': 60000},
    'corta':      {'label': 'Corta de arroz',   'icon': 'fa-wheat-awn', 'unit': 'hectáreas',  'vpunt_default': 650000},
    'carga':      {'label': 'Carga de camiones','icon': 'fa-truck',     'unit': 'camiones',   'vpunt_default': 0},
    'otro':       {'label': 'Otro',             'icon': 'fa-ellipsis',  'unit': 'unidades',   'vpunt_default': 0},
}

@app.route('/recibos/labores/abonada', methods=['GET', 'POST'])
def labores_abonada():
    if 'user_id' not in session:
        return auth_redirect('login', 'Inicia sesion.', 'warning')
    lote_id = session.get('lote_id')

    error = None
    success = None
    next_serial = get_next_serial(lote_id)

    if request.method == 'POST':
        tipo_labor    = request.form.get('tipo_labor', 'abonada')
        fecha_str     = request.form.get('fecha', '').strip()
        serial_ini    = request.form.get('serial_inicio', '').strip()
        direccion_r   = request.form.get('direccion', '').strip()
        ciudad_r      = request.form.get('ciudad', '').strip()
        lote_nombre   = request.form.get('lote_nombre', 'El Mangon').strip()
        concepto_base = request.form.get('concepto_base', '').strip()
        worker_ids    = request.form.getlist('worker_ids')
        labor_info    = TIPOS_LABOR_CANTIDAD.get(tipo_labor, TIPOS_LABOR_CANTIDAD['otro'])

        if not worker_ids:
            error = "Selecciona al menos un trabajador."
        elif not serial_ini or not serial_ini.isdigit():
            error = "El serial de inicio debe ser un número válido."
        else:
            rte_pct = float(request.form.get('rte_fte', '0') or '0')
            fecha_obj = None
            if fecha_str:
                try:
                    fecha_obj = date.fromisoformat(fecha_str)
                except ValueError:
                    error = "Formato de fecha inválido."

            if not error:
                try:
                    conn = get_db_connection()
                    cursor = conn.cursor(dictionary=True)
                    placeholders = ','.join(['%s'] * len(worker_ids))
                    cursor.execute(f"SELECT id_worker, name, lastname, cc, direccion, ciudad, phone_number as telefono FROM workers WHERE id_worker IN ({placeholders}) AND lote_id=%s", (*worker_ids, lote_id))
                    selected_workers = cursor.fetchall()
                    serial_int = int(serial_ini)
                    created = 0

                    for w in selected_workers:
                        wid_str = str(w['id_worker'])
                        # Read the pre-calculated total from the hidden input
                        try:
                            valor = float(request.form.get(f'valor_w_{wid_str}', '0') or '0')
                            # Also try to read cant/vpunt for a meaningful concepto
                            cant  = float((request.form.get(f'cant_w_{wid_str}', '0') or '0').replace('.','').replace(',',''))
                            vpunt = float((request.form.get(f'vpunt_w_{wid_str}', str(labor_info['vpunt_default'])) or str(labor_info['vpunt_default'])).replace('.','').replace(',',''))
                        except Exception:
                            valor, cant, vpunt = 0.0, 0.0, float(labor_info['vpunt_default'])

                        rte_amount = round(valor * rte_pct / 100)
                        neto = valor - rte_amount

                        unit_label = labor_info['unit']
                        vpunt_fmt = f"{int(vpunt):,}".replace(',','.')
                        concepto_txt = concepto_base or f"{labor_info['label']} {cant:.0f} {unit_label} × $ {vpunt_fmt} c/u en el lote {lote_nombre}"

                        proveedor_name = f"{w['name']} {w['lastname']}"
                        dir_w    = direccion_r or w.get('direccion') or ''
                        ciudad_w = ciudad_r    or w.get('ciudad')    or ''
                        tel_w    = w.get('telefono') or ''
                        lineas   = [{'concepto': concepto_txt, 'valor': valor}]
                        cj       = json.dumps(lineas, ensure_ascii=False)

                        cursor2 = conn.cursor()
                        cursor2.execute("SELECT serial FROM recibos WHERE serial = %s", (serial_int,))
                        if cursor2.fetchone():
                            cursor2.execute("UPDATE recibos SET serial = serial + 1 WHERE serial >= %s ORDER BY serial DESC", (serial_int,))
                        cursor2.execute("""
                            INSERT INTO recibos (serial, fecha, proveedor, nit, direccion, telefono, ciudad, concepto, valor_operacion, rte_fte, neto_a_pagar, conceptos_json, lote_id)
                            VALUES (%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s)
                        """, (serial_int, fecha_obj, proveedor_name, w['cc'], dir_w, tel_w, ciudad_w, concepto_txt, valor, rte_pct, neto, cj, lote_id))
                        cursor2.close()
                        serial_int += 1
                        created += 1

                    conn.commit()
                    cursor.close(); conn.close()
                    success = f"✅ {created} recibo(s) creados para {labor_info['label']} (seriales {serial_ini}–{serial_int-1})."
                    next_serial = get_next_serial(lote_id)
                except Exception as e:
                    error = f"Error al guardar: {e}"

    db_workers = _load_workers_for_form()
    return render_template('recibos/labores_abonada.html',
                           db_workers=db_workers,
                           tipos_labor=TIPOS_LABOR_CANTIDAD,
                           next_serial=next_serial,
                           error=error, success=success,
                           today=date.today().isoformat())

# =========================
# RECIBO DETALLE / DELETE
# =========================
@app.route('/recibos/<int:serial>')
def detalle_recibo(serial):
    from auth_middleware import has_permission
    if 'user_id' not in session:
        return auth_redirect('login', 'Inicia sesion.', 'warning')
    lote_id = session.get('lote_id')
    conn = get_db_connection()
    cursor = conn.cursor(dictionary=True)
    if lote_id:
        cursor.execute("SELECT * FROM recibos WHERE serial = %s AND lote_id = %s", (serial, lote_id))
    else:
        cursor.execute("SELECT * FROM recibos WHERE serial = %s", (serial,))
    recibo = cursor.fetchone()
    cursor.close(); conn.close()
    if not recibo:
        return redirect(url_for('lista_recibos'))
    if recibo.get('neto_a_pagar') is not None:
        recibo['neto_a_pagar'] = float(recibo['neto_a_pagar'])
    if recibo.get('valor_operacion') is not None:
        recibo['valor_operacion'] = float(recibo['valor_operacion'])
    if recibo.get('rte_fte') is not None:
        recibo['rte_fte'] = float(recibo['rte_fte'])
    # Parse multi-concept lines
    conceptos_lineas = []
    if recibo.get('conceptos_json'):
        try:
            conceptos_lineas = json.loads(recibo['conceptos_json'])
        except Exception:
            pass
    return render_template('recibos/detalle.html', recibo=recibo, conceptos_lineas=conceptos_lineas,
                           can_edit=has_permission('recibo.edit'),
                           can_delete=has_permission('recibo.delete'))


@app.route('/recibos/<int:serial>/editar', methods=['GET', 'POST'])
def editar_recibo(serial):
    from auth_middleware import has_permission
    if 'user_id' not in session:
        return auth_redirect('login', 'Inicia sesion para editar recibos.', 'warning')
    if not session.get('lote_id'):
        return redirect(url_for('select_lote'))
    if not has_permission('recibo.edit'):
        return redirect(url_for('lista_recibos'))

    lote_id = session['lote_id']
    conn = get_db_connection()
    cursor = conn.cursor(dictionary=True)
    cursor.execute("SELECT * FROM recibos WHERE serial = %s AND lote_id = %s", (serial, lote_id))
    recibo = cursor.fetchone()
    if not recibo:
        cursor.close(); conn.close()
        return redirect(url_for('lista_recibos'))

    success = None
    error = None
    warning = None
    form_data = _build_recibo_form_data(recibo)

    if request.method == 'POST':
        fecha_str = request.form.get('fecha', '').strip()
        proveedor = request.form.get('proveedor', '').strip()
        nit = request.form.get('nit', '').strip()
        direccion_r = request.form.get('direccion', '').strip()
        telefono = request.form.get('telefono', '').strip()
        ciudad_r = request.form.get('ciudad', '').strip()

        lineas = []
        for i in range(1, 8):
            c_text = request.form.get(f'concepto_{i}', '').strip()
            c_val_raw = request.form.get(f'valor_{i}', '').strip().replace('.', '').replace(',', '')
            if c_text:
                lineas.append({'concepto': c_text, 'valor': float(c_val_raw) if c_val_raw else 0.0})

        concepto = '\n'.join(l['concepto'] for l in lineas)
        valor_op = sum(l['valor'] for l in lineas)
        conceptos_json_str = json.dumps(lineas, ensure_ascii=False)
        rte_raw = request.form.get('rte_fte', '').strip().replace('.', '').replace(',', '')
        rte_fte = float(rte_raw) if rte_raw else 0.0
        neto_raw = request.form.get('neto_a_pagar', '').strip().replace('.', '').replace(',', '')

        form_data = {
            'serial': serial,
            'fecha': fecha_str,
            'proveedor': proveedor,
            'nit': nit,
            'direccion': direccion_r,
            'telefono': telefono,
            'ciudad': ciudad_r,
            'lineas': lineas,
            'rte_fte': rte_fte,
            'neto_a_pagar': float(neto_raw) if neto_raw else max(0.0, valor_op - rte_fte),
        }

        if not proveedor or not lineas:
            error = 'Proveedor y al menos un concepto son obligatorios.'
        elif any(es_concepto_prohibido(l['concepto']) for l in lineas):
            error = '⚠️ No se registran compras de aceite ni ACPM directo. Solo se registra el transporte de ACPM.'
        else:
            fecha_obj = None
            if fecha_str:
                try:
                    fecha_obj = date.fromisoformat(fecha_str)
                except ValueError:
                    error = 'Formato de fecha inválido.'

            if not error:
                neto = float(neto_raw) if neto_raw else max(0.0, valor_op - rte_fte)
                try:
                    cursor2 = conn.cursor()
                    cursor2.execute("""
                        UPDATE recibos
                        SET fecha=%s, proveedor=%s, nit=%s, direccion=%s, telefono=%s, ciudad=%s,
                            concepto=%s, valor_operacion=%s, rte_fte=%s, neto_a_pagar=%s, conceptos_json=%s
                        WHERE serial=%s AND lote_id=%s
                    """, (fecha_obj, proveedor, nit, direccion_r, telefono, ciudad_r,
                          concepto, valor_op, rte_fte, neto, conceptos_json_str, serial, lote_id))
                    conn.commit()
                    cursor2.close()
                    cursor.execute("SELECT * FROM recibos WHERE serial = %s AND lote_id = %s", (serial, lote_id))
                    recibo = cursor.fetchone()
                    form_data = _build_recibo_form_data(recibo)
                    success = f'Recibo #{serial} actualizado correctamente.'
                except Exception as e:
                    error = f'Error al actualizar: {e}'

    cursor.close(); conn.close()
    trabajadores = _get_trabajadores_for_autocomplete()
    return render_template('recibos/nuevo.html',
                           next_serial=serial,
                           trabajadores=trabajadores,
                           error=error,
                           warning=warning,
                           success=success,
                           form_data=form_data,
                           page_title=f'Editar Recibo #{serial}',
                           page_subtitle='Corrige los datos del recibo sin depender de la base manual.',
                           submit_label='Guardar cambios',
                           serial_readonly=True,
                           page_mode='edit',
                           form_action=url_for('editar_recibo', serial=serial),
                           back_url=url_for('detalle_recibo', serial=serial),
                           back_label='Ver detalle',
                           today=date.today().isoformat())

@app.route('/recibos/<int:serial>/eliminar', methods=['POST'])
def eliminar_recibo(serial):
    from auth_middleware import has_permission
    if 'user_id' not in session:
        return auth_redirect('login', 'Inicia sesion.', 'warning')
    lote_id = session.get('lote_id')
    if lote_id and not has_permission('recibo.delete'):
        return redirect(url_for('lista_recibos'))
    conn = get_db_connection()
    cursor = conn.cursor()
    if lote_id:
        cursor.execute("DELETE FROM recibos WHERE serial = %s AND lote_id = %s", (serial, lote_id))
    else:
        cursor.execute("DELETE FROM recibos WHERE serial = %s", (serial,))
    conn.commit(); cursor.close(); conn.close()
    return redirect(url_for('lista_recibos'))


@app.route('/recibos/conciliacion')
def conciliacion_recibos():
    from auth_middleware import has_permission
    if 'user_id' not in session:
        return auth_redirect('login', 'Inicia sesion para conciliar recibos.', 'warning')
    if not has_permission('recibo.view'):
        return redirect(url_for('dashboard', msg='Sin permiso para ver recibos.', msg_type='danger'))

    q = request.args.get('q', '').strip().lower()
    lote_sel = request.args.get('lote_id', '').strip()

    conn = get_db_connection()
    cursor = conn.cursor(dictionary=True)

    if session.get('is_superadmin'):
        cursor.execute("SELECT id, nombre FROM lotes ORDER BY nombre")
        lotes = cursor.fetchall()
    else:
        lotes = _get_user_lotes(session['user_id'])
        lotes = [{'id': l.get('id'), 'nombre': l.get('nombre')} for l in lotes]
        if not lotes and session.get('lote_id'):
            lotes = [{'id': session['lote_id'], 'nombre': session.get('lote_nombre', f"Lote {session['lote_id']}")}]

    lote_ids = [int(l['id']) for l in lotes if l.get('id') is not None]
    if not lote_ids:
        cursor.close(); conn.close()
        return render_template('recibos/conciliacion.html', recibos=[], lotes=[], resumen_lotes=[],
                               selected_lote='', search=q, total_recibos=0, total_neto=0)

    placeholders = ','.join(['%s'] * len(lote_ids))
    params = list(lote_ids)
    sql = f"""
        SELECT r.serial, r.fecha, r.proveedor, r.nit, r.concepto, r.neto_a_pagar,
               r.lote_id, l.nombre AS lote_nombre
        FROM recibos r
        LEFT JOIN lotes l ON l.id = r.lote_id
        WHERE r.lote_id IN ({placeholders})
    """
    if lote_sel and lote_sel.isdigit() and int(lote_sel) in lote_ids:
        sql += " AND r.lote_id = %s"
        params.append(int(lote_sel))
    if q:
        like = f'%{q}%'
        sql += " AND (LOWER(r.proveedor) LIKE %s OR LOWER(IFNULL(r.nit,'')) LIKE %s OR LOWER(IFNULL(r.concepto,'')) LIKE %s OR CAST(r.serial AS CHAR) LIKE %s)"
        params.extend([like, like, like, like])
    sql += " ORDER BY r.lote_id, r.serial DESC"
    cursor.execute(sql, tuple(params))
    recibos = cursor.fetchall()

    resumen = {}
    for row in recibos:
        row['neto_a_pagar'] = float(row.get('neto_a_pagar') or 0)
        lote_key = row['lote_id']
        item = resumen.setdefault(lote_key, {
            'lote_id': lote_key,
            'lote_nombre': row.get('lote_nombre') or f'Lote {lote_key}',
            'total_recibos': 0,
            'total_neto': 0.0,
            'min_serial': row['serial'],
            'max_serial': row['serial'],
        })
        item['total_recibos'] += 1
        item['total_neto'] += row['neto_a_pagar']
        item['min_serial'] = min(item['min_serial'], row['serial'])
        item['max_serial'] = max(item['max_serial'], row['serial'])

    resumen_lotes = sorted(resumen.values(), key=lambda item: (item['lote_nombre'] or '', item['lote_id']))
    total_neto = sum(row['neto_a_pagar'] for row in recibos)

    cursor.close(); conn.close()
    return render_template('recibos/conciliacion.html',
                           recibos=recibos,
                           lotes=lotes,
                           resumen_lotes=resumen_lotes,
                           selected_lote=lote_sel,
                           search=q,
                           total_recibos=len(recibos),
                           total_neto=total_neto)

# =========================
# REPORTES
# =========================
@app.route('/reportes')
def reportes():
    if 'user_id' not in session:
        return auth_redirect('login', 'Inicia sesion.', 'warning')
    if not session.get('lote_id'):
        return redirect(url_for('select_lote'))
    lote_id = session['lote_id']

    conn = get_db_connection()
    cursor = conn.cursor(dictionary=True)

    cursor.execute("SELECT COALESCE(SUM(neto_a_pagar),0) as total FROM recibos WHERE lote_id=%s", (lote_id,))
    total_gastado = float(cursor.fetchone()['total'])

    cursor.execute("""
        SELECT proveedor, SUM(neto_a_pagar) as total
        FROM recibos WHERE neto_a_pagar IS NOT NULL AND lote_id=%s
        GROUP BY proveedor ORDER BY total DESC LIMIT 10
    """, (lote_id,))
    por_trabajador = cursor.fetchall()

    cursor.execute("""
        SELECT
            YEARWEEK(fecha, 1)               AS semana_key,
            MIN(fecha)                        AS semana_inicio,
            SUM(neto_a_pagar)                 AS total
        FROM recibos
        WHERE fecha IS NOT NULL AND neto_a_pagar IS NOT NULL AND lote_id=%s
        GROUP BY semana_key
        ORDER BY semana_key DESC
        LIMIT 16
    """, (lote_id,))
    por_semana_raw = cursor.fetchall()

    cursor.execute("SELECT SUM(cargas) as total_cargas, SUM(kg_total) as total_kg FROM cosechas WHERE lote_id=%s", (lote_id,))
    produccion = cursor.fetchone()
    total_cargas = int(produccion['total_cargas'] or 0)
    total_kg = float(produccion['total_kg'] or 0)

    cursor.execute("SELECT * FROM recibos WHERE lote_id=%s ORDER BY serial DESC LIMIT 5", (lote_id,))
    ultimos_recibos = cursor.fetchall()

    # Lote config for limits
    cursor.execute("SELECT hectareas, meta_cargas_ha FROM lotes WHERE id=%s", (lote_id,))
    lote_row = cursor.fetchone() or {}
    lote_ha = float(lote_row.get('hectareas') or session.get('lote_ha') or TOTAL_HA)
    meta_cargas = int(lote_ha * (lote_row.get('meta_cargas_ha') or MIN_CARGAS_POR_HA))

    # Workers list for filter selector
    cursor.execute("SELECT id_worker, CONCAT(name,' ',lastname) AS nombre_completo, cc AS nit, trabajo_desarrolla AS cargo FROM workers WHERE lote_id=%s AND activo=1 ORDER BY name", (lote_id,))
    workers_list = cursor.fetchall()

    # Presupuesto info
    cursor.execute("SELECT COALESCE(SUM(monto),0) as ti FROM presupuesto_recargas WHERE lote_id=%s", (lote_id,))
    total_ingresado = float(cursor.fetchone()['ti'])
    pres_saldo = total_ingresado - total_gastado
    pres_pct   = round(total_gastado / total_ingresado * 100, 1) if total_ingresado > 0 else 0

    cursor.close(); conn.close()

    ledger = build_budget_movements(lote_id)
    movimientos_reporte = ledger['movimientos']

    max_gasto = lote_ha * MAX_GASTO_POR_HA
    for p in por_trabajador:
        p['total'] = float(p['total'])

    por_semana = []
    for row in reversed(por_semana_raw):
        row['total'] = float(row['total'])
        ini = row['semana_inicio']
        if hasattr(ini, 'strftime'):
            row['semana'] = 'Sem. ' + ini.strftime('%d/%m')
        else:
            row['semana'] = str(ini)
        por_semana.append(row)

    pct_gasto = min(100, round(total_gastado / max_gasto * 100, 1)) if max_gasto else 0
    pct_produccion = min(100, round(total_cargas / meta_cargas * 100, 1)) if meta_cargas else 0

    return render_template('reportes/index.html',
        total_gastado=total_gastado,
        max_gasto=max_gasto,
        pct_gasto=pct_gasto,
        por_trabajador=por_trabajador,
        por_semana=por_semana,
        total_cargas=total_cargas,
        total_kg=total_kg,
        min_cargas=meta_cargas,
        pct_produccion=pct_produccion,
        ultimos_recibos=ultimos_recibos,
        total_ha=lote_ha,
        max_gasto_ha=MAX_GASTO_POR_HA,
        workers_list=workers_list,
        total_ingresado=total_ingresado,
        pres_saldo=pres_saldo,
        pres_pct=pres_pct,
        movimientos_reporte=movimientos_reporte,
        saldo_inicial_reporte=ledger['saldo_inicial'],
        saldo_final_reporte=ledger['saldo_final'],
    )

@app.route('/presupuesto', methods=['GET', 'POST'])
def presupuesto_view():
    if 'user_id' not in session:
        return auth_redirect('login', 'Inicia sesion.', 'warning')
    if not session.get('lote_id'):
        return redirect(url_for('select_lote'))
    lote_id = session['lote_id']

    conn = get_db_connection()
    cursor = conn.cursor(dictionary=True)
    error = None
    success = None

    if request.method == 'POST':
        monto_raw = (request.form.get('monto') or '').replace('.', '').replace(',', '').strip()
        descripcion = (request.form.get('descripcion') or '').strip()[:255]
        fecha = (request.form.get('fecha') or '').strip() or date.today().isoformat()
        try:
            monto = float(monto_raw)
            if monto <= 0:
                raise ValueError()
        except (ValueError, TypeError):
            error = 'Ingresa un monto válido mayor a cero.'
        else:
            cursor.execute(
                "INSERT INTO presupuesto_recargas (lote_id, monto, descripcion, fecha) VALUES (%s,%s,%s,%s)",
                (lote_id, monto, descripcion, fecha))
            conn.commit()
            success = f'Recarga de $ {monto:,.0f} registrada correctamente.'.replace(',', '.')

    cursor.execute("SELECT COALESCE(SUM(monto),0) as ti FROM presupuesto_recargas WHERE lote_id=%s", (lote_id,))
    total_ingresado = float(cursor.fetchone()['ti'])
    cursor.execute("SELECT COALESCE(SUM(neto_a_pagar),0) as tg FROM recibos WHERE lote_id=%s", (lote_id,))
    total_gastado = float(cursor.fetchone()['tg'])
    saldo = total_ingresado - total_gastado
    pct_usado = round(total_gastado / total_ingresado * 100, 1) if total_ingresado > 0 else 0
    alerta = total_ingresado > 0 and saldo < total_ingresado * 0.15

    cursor.execute("SELECT * FROM presupuesto_recargas WHERE lote_id=%s ORDER BY fecha DESC, id DESC", (lote_id,))
    recargas = cursor.fetchall()
    cursor.close(); conn.close()

    _PRES_TMPL = r"""{% extends 'base.html' %}
{% block title %}Presupuesto{% endblock %}
{% block content %}
<div class="container py-4" style="max-width:900px">

  <div class="d-flex align-items-center gap-3 mb-4">
    <div class="rounded-3 p-3" style="background:#1B4332">
      <i class="bi bi-wallet2 fs-3 text-white"></i>
    </div>
    <div>
      <h3 class="mb-0 fw-bold" style="color:#1B4332">Presupuesto</h3>
      <small class="text-muted">{{ lote_nombre }}</small>
    </div>
    <a href="{{ url_for('reportes') }}" class="btn btn-sm btn-outline-secondary ms-auto">
      <i class="bi bi-bar-chart me-1"></i>Reportes
    </a>
  </div>

  {% if error %}
  <div class="alert alert-warning alert-dismissible fade show"><i class="bi bi-exclamation-triangle me-2"></i>{{ error }}
    <button type="button" class="btn-close" data-bs-dismiss="alert"></button></div>
  {% endif %}
  {% if success %}
  <div class="alert alert-success alert-dismissible fade show"><i class="bi bi-check-circle me-2"></i>{{ success }}
    <button type="button" class="btn-close" data-bs-dismiss="alert"></button></div>
  {% endif %}

  {# ── Balance Cards ── #}
  <div class="row g-3 mb-4">
    <div class="col-md-4">
      <div class="card border-0 shadow-sm h-100" style="border-top:4px solid #2D6A4F!important">
        <div class="card-body text-center">
          <div class="text-muted small mb-1">Total Ingresado</div>
          <div class="fw-bold fs-4" style="color:#2D6A4F">$ {{ '{:,.0f}'.format(total_ingresado).replace(',','.') }}</div>
        </div>
      </div>
    </div>
    <div class="col-md-4">
      <div class="card border-0 shadow-sm h-100" style="border-top:4px solid #E9A800!important">
        <div class="card-body text-center">
          <div class="text-muted small mb-1">Total Gastado</div>
          <div class="fw-bold fs-4" style="color:#8B6000">$ {{ '{:,.0f}'.format(total_gastado).replace(',','.') }}</div>
        </div>
      </div>
    </div>
    <div class="col-md-4">
      <div class="card border-0 shadow-sm h-100" style="border-top:4px solid {% if alerta %}#DC3545{% elif saldo < 0 %}#DC3545{% else %}#1B4332{% endif %}!important">
        <div class="card-body text-center">
          <div class="text-muted small mb-1">Saldo Disponible</div>
          <div class="fw-bold fs-4" style="color:{% if saldo < 0 %}#DC3545{% else %}#1B4332{% endif %}">
            $ {{ '{:,.0f}'.format(saldo).replace(',','.') }}
          </div>
          {% if alerta %}<div class="badge bg-danger mt-1">¡Saldo bajo!</div>{% endif %}
        </div>
      </div>
    </div>
  </div>

  {# ── Progress bar ── #}
  {% if total_ingresado > 0 %}
  <div class="card border-0 shadow-sm mb-4">
    <div class="card-body">
      <div class="d-flex justify-content-between mb-1">
        <small class="fw-semibold">Uso del presupuesto</small>
        <small class="fw-bold {% if pct_usado > 85 %}text-danger{% elif pct_usado > 60 %}text-warning{% else %}text-success{% endif %}">{{ pct_usado }}%</small>
      </div>
      <div class="progress" style="height:14px;border-radius:8px">
        <div class="progress-bar {% if pct_usado > 85 %}bg-danger{% elif pct_usado > 60 %}bg-warning{% else %}bg-success{% endif %}"
          style="width:{{ [pct_usado,100]|min }}%;border-radius:8px"></div>
      </div>
    </div>
  </div>
  {% endif %}

  <div class="row g-4">
    {# ── Agregar recarga ── #}
    <div class="col-md-5">
      <div class="card border-0 shadow-sm">
        <div class="card-header border-0 fw-bold" style="background:#1B4332;color:#fff">
          <i class="bi bi-plus-circle me-2"></i>Agregar Presupuesto
        </div>
        <div class="card-body">
          <form method="POST" action="{{ url_for('presupuesto_view') }}">
            <div class="mb-3">
              <label class="form-label fw-semibold">Monto ($)</label>
              <input type="number" name="monto" class="form-control" placeholder="Ej: 40000000" min="1" step="1000" required>
            </div>
            <div class="mb-3">
              <label class="form-label fw-semibold">Descripción</label>
              <input type="text" name="descripcion" class="form-control" placeholder="Ej: Capital inicial lote">
            </div>
            <div class="mb-3">
              <label class="form-label fw-semibold">Fecha</label>
              <input type="date" name="fecha" class="form-control" value="{{ today }}">
            </div>
            <button type="submit" class="btn w-100 text-white fw-bold" style="background:#2D6A4F">
              <i class="bi bi-plus-lg me-2"></i>Registrar Recarga
            </button>
          </form>
        </div>
      </div>
    </div>

    {# ── Historial recargas ── #}
    <div class="col-md-7">
      <div class="card border-0 shadow-sm">
        <div class="card-header border-0 fw-bold" style="background:#2D6A4F;color:#fff">
          <i class="bi bi-clock-history me-2"></i>Historial de Recargas
        </div>
        <div class="card-body p-0">
          {% if recargas %}
          <div class="table-responsive">
            <table class="table table-sm table-hover mb-0 align-middle">
              <thead style="background:#F0F8F4">
                <tr>
                  <th class="px-3">Fecha</th>
                  <th>Descripción</th>
                  <th class="text-end pe-3">Monto</th>
                </tr>
              </thead>
              <tbody>
                {% for r in recargas %}
                <tr>
                  <td class="px-3 text-muted small">{{ r.fecha.strftime('%d/%m/%Y') if r.fecha else '-' }}</td>
                  <td class="small">{{ r.descripcion or '—' }}</td>
                  <td class="text-end pe-3 fw-semibold text-success">$ {{ '{:,.0f}'.format(r.monto|float).replace(',','.') }}</td>
                </tr>
                {% endfor %}
              </tbody>
              <tfoot style="background:#1B4332;color:#fff">
                <tr>
                  <td colspan="2" class="px-3 fw-bold">Total Ingresado</td>
                  <td class="text-end pe-3 fw-bold">$ {{ '{:,.0f}'.format(total_ingresado).replace(',','.') }}</td>
                </tr>
              </tfoot>
            </table>
          </div>
          {% else %}
          <div class="text-center py-5 text-muted">
            <i class="bi bi-wallet2 fs-1 d-block mb-2" style="color:#ccc"></i>
            Aún no hay recargas registradas.<br>
            <small>Agrega tu presupuesto inicial para comenzar el seguimiento.</small>
          </div>
          {% endif %}
        </div>
      </div>
    </div>
  </div>

</div>
{% endblock %}"""

    # Ensure template file exists for future use
    _tmpl_dir = os.path.join(os.path.dirname(__file__), 'templates', 'presupuesto')
    os.makedirs(_tmpl_dir, exist_ok=True)
    _tmpl_path = os.path.join(_tmpl_dir, 'index.html')
    if not os.path.exists(_tmpl_path):
        with open(_tmpl_path, 'w', encoding='utf-8') as _tf:
            _tf.write(_PRES_TMPL)

    return render_template_string(_PRES_TMPL,
        total_ingresado=total_ingresado,
        total_gastado=total_gastado,
        saldo=saldo,
        pct_usado=pct_usado,
        alerta=alerta,
        recargas=recargas,
        error=error,
        success=success,
        today=date.today().isoformat(),
        lote_nombre=session.get('lote_nombre', ''),
    )


@app.route('/reportes/semana')
def reporte_semana():
    if 'user_id' not in session:
        return auth_redirect('login', 'Inicia sesion.', 'warning')
    lote_id = session.get('lote_id')
    fecha_str = request.args.get('fecha', date.today().isoformat())
    try:
        fecha_ref = date.fromisoformat(fecha_str)
    except:
        fecha_ref = date.today()
    inicio = fecha_ref - timedelta(days=fecha_ref.weekday())
    fin = inicio + timedelta(days=6)

    conn = get_db_connection()
    cursor = conn.cursor(dictionary=True)
    cursor.execute("SELECT * FROM recibos WHERE fecha BETWEEN %s AND %s AND lote_id=%s ORDER BY serial", (inicio, fin, lote_id))
    recibos = cursor.fetchall()
    cursor.execute("SELECT COALESCE(SUM(neto_a_pagar),0) as total FROM recibos WHERE fecha BETWEEN %s AND %s AND lote_id=%s", (inicio, fin, lote_id))
    total = float(cursor.fetchone()['total'])
    cursor.execute("SELECT COALESCE(SUM(monto),0) as total FROM presupuesto_recargas WHERE fecha BETWEEN %s AND %s AND lote_id=%s", (inicio, fin, lote_id))
    ingresos_semana = float(cursor.fetchone()['total'])
    cursor.close(); conn.close()
    for r in recibos:
        r['neto_a_pagar'] = float(r['neto_a_pagar'] or 0)
        r['valor_operacion'] = float(r['valor_operacion'] or 0)

    ledger = build_budget_movements(lote_id, inicio, fin)
    movimientos_por_serial = {
        movimiento['serial']: movimiento
        for movimiento in ledger['movimientos']
        if movimiento.get('origen') == 'recibo' and movimiento.get('serial') is not None
    }

    return render_template('reportes/semana.html',
        recibos=recibos, total=total,
        ingresos_semana=ingresos_semana,
        movimientos_semana=ledger['movimientos'],
        movimientos_por_serial=movimientos_por_serial,
        saldo_inicial_semana=ledger['saldo_inicial'],
        saldo_final_semana=ledger['saldo_final'],
        inicio=inicio, fin=fin, fecha_str=fecha_str)

@app.route('/reportes/exportar_txt')
def exportar_txt():
    # TXT export deprecated — redirect to full PDF report
    return redirect(url_for('generar_pdf', tipo='completo'))

# =========================
# PRODUCCIÓN
# =========================
@app.route('/produccion')
def lista_produccion():
    if 'user_id' not in session:
        return auth_redirect('login', 'Inicia sesion.', 'warning')
    if not session.get('lote_id'):
        return redirect(url_for('select_lote'))
    lote_id = session['lote_id']
    conn = get_db_connection()
    cursor = conn.cursor(dictionary=True)
    cursor.execute("SELECT * FROM cosechas WHERE lote_id=%s ORDER BY fecha DESC", (lote_id,))
    cosechas = cursor.fetchall()
    cursor.execute("SELECT COALESCE(SUM(cargas),0) as tc, COALESCE(SUM(kg_total),0) as tk, COALESCE(SUM(valor_total),0) as tv FROM cosechas WHERE lote_id=%s", (lote_id,))
    totales = cursor.fetchone()
    cursor.execute("SELECT hectareas, meta_cargas_ha FROM lotes WHERE id=%s", (lote_id,))
    lote_row = cursor.fetchone() or {}
    cursor.close(); conn.close()
    for c in cosechas:
        for k in ['kg_total','precio_carga','valor_total','bultos_ha','total_bultos']:
            if c.get(k): c[k] = float(c[k])
    total_cargas = int(totales['tc'] or 0)
    lote_ha = float(lote_row.get('hectareas') or TOTAL_HA)
    meta_cargas = int(lote_ha * (lote_row.get('meta_cargas_ha') or MIN_CARGAS_POR_HA))
    pct_produccion = min(100, round(total_cargas / meta_cargas * 100, 1)) if meta_cargas else 0
    return render_template('produccion/index.html', cosechas=cosechas,
        totales=totales, min_cargas=meta_cargas, min_cargas_ha=MIN_CARGAS_POR_HA,
        kg_por_carga=KG_POR_CARGA, total_ha=lote_ha,
        total_cargas=total_cargas, pct_produccion=pct_produccion)


@app.route('/produccion/nueva', methods=['GET', 'POST'])
def nueva_cosecha():
    if 'user_id' not in session:
        return auth_redirect('login', 'Inicia sesion.', 'warning')
    if not session.get('lote_id'):
        return redirect(url_for('select_lote'))
    lote_id = session['lote_id']
    error = None; success = None
    # Get lote name for default
    try:
        conn = get_db_connection()
        cursor = conn.cursor(dictionary=True)
        cursor.execute("SELECT nombre, hectareas FROM lotes WHERE id=%s", (lote_id,))
        lote_row = cursor.fetchone() or {}
        cursor.close(); conn.close()
        nombre_lote = lote_row.get('nombre', 'El Mangon')
        lote_ha = float(lote_row.get('hectareas') or TOTAL_HA)
    except:
        nombre_lote = 'El Mangon'
        lote_ha = TOTAL_HA
    if request.method == 'POST':
        fecha_str       = request.form.get('fecha','').strip()
        lote            = request.form.get('lote', nombre_lote).strip()
        hectareas       = request.form.get('hectareas', str(lote_ha)).strip()
        cargas_raw      = request.form.get('cargas','').strip()
        precio_carga_raw= request.form.get('precio_carga','').strip().replace('.','').replace(',','')
        fase            = request.form.get('fase', 'cosecha').strip()
        variedad_semilla= request.form.get('variedad_semilla','').strip() or None
        origen_semilla  = request.form.get('origen_semilla','').strip() or None
        bultos_ha_raw   = request.form.get('bultos_ha','').strip()
        metodo_siembra  = request.form.get('metodo_siembra','al_voleo').strip()
        fecha_siembra_str= request.form.get('fecha_siembra','').strip() or None
        observaciones   = request.form.get('observaciones') or None

        if not fecha_str:
            error = "La fecha es obligatoria."
        elif fase == 'cosecha' and not cargas_raw:
            error = "El número de cargas es obligatorio para registrar una cosecha."
        else:
            try:
                fecha_obj   = date.fromisoformat(fecha_str)
                ha          = float(hectareas) if hectareas else lote_ha
                cargas      = int(cargas_raw) if cargas_raw else 0
                kg_total    = cargas * KG_POR_CARGA if cargas else None
                precio      = float(precio_carga_raw) if precio_carga_raw else None
                valor_total = cargas * precio if (cargas and precio) else None
                bultos_ha   = float(bultos_ha_raw) if bultos_ha_raw else None
                total_bultos= round(bultos_ha * ha, 2) if (bultos_ha and ha) else None
                fecha_siembra_obj = date.fromisoformat(fecha_siembra_str) if fecha_siembra_str else None

                conn   = get_db_connection()
                cursor = conn.cursor()
                cursor.execute("""
                    INSERT INTO cosechas
                        (fecha, lote, hectareas, cargas, kg_total,
                         precio_carga, valor_total, observaciones,
                         fase, variedad_semilla, origen_semilla,
                         bultos_ha, total_bultos, metodo_siembra, fecha_siembra, lote_id)
                    VALUES (%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s)
                """, (fecha_obj, lote, ha,
                      cargas if cargas else None, kg_total,
                      precio, valor_total, observaciones,
                      fase, variedad_semilla, origen_semilla,
                      bultos_ha, total_bultos, metodo_siembra, fecha_siembra_obj, lote_id))
                conn.commit(); cursor.close(); conn.close()
                if fase == 'cosecha':
                    success = f"Cosecha registrada: {cargas} cargas ({kg_total:,.0f} kg)."
                else:
                    success = f"Siembra registrada: {variedad_semilla or 'variedad sin especificar'} — {total_bultos or '?'} bultos totales."
            except Exception as e:
                error = f"Error: {e}"

    return render_template('produccion/nueva.html', error=error, success=success,
                           today=date.today().isoformat(), total_ha=lote_ha,
                           min_cargas=MIN_CARGAS, min_cargas_ha=MIN_CARGAS_POR_HA)

# =========================
# CONFIGURACIÓN
# =========================
@app.route('/config', methods=['GET', 'POST'])
def config_app():
    from auth_middleware import has_permission
    if 'user_id' not in session:
        return auth_redirect('login', 'Inicia sesion para acceder a la configuracion.', 'warning')
    if not session.get('lote_id'):
        return redirect(url_for('select_lote'))
    if not has_permission('config.manage'):
        return redirect(url_for('dashboard', msg='Sin permiso para configurar.', msg_type='danger'))
    lote_id = session['lote_id']

    message = None
    error = None

    if request.method == 'POST':
        serial_inicial = request.form.get('serial_inicial', '').strip()
        if not serial_inicial or not serial_inicial.isdigit():
            error = "El serial inicial debe ser un número entero positivo."
        else:
            try:
                conn = get_db_connection()
                cursor = conn.cursor()
                cursor.execute("""
                    INSERT INTO config (clave, valor, lote_id) VALUES ('serial_inicial', %s, %s)
                    ON DUPLICATE KEY UPDATE valor = VALUES(valor)
                """, (serial_inicial, lote_id))
                conn.commit()
                cursor.close()
                conn.close()
                message = f"Serial inicial actualizado a {serial_inicial}."
            except Exception as e:
                error = f"Error: {e}"

    try:
        conn = get_db_connection()
        cursor = conn.cursor(dictionary=True)
        cursor.execute("SELECT clave, valor FROM config WHERE lote_id=%s", (lote_id,))
        config_rows = {row['clave']: row['valor'] for row in cursor.fetchall()}
        cursor.execute("SELECT COUNT(*) as total FROM recibos WHERE lote_id=%s", (lote_id,))
        total_recibos = cursor.fetchone()['total']
        next_serial_effective = get_next_serial(lote_id)
        cursor.close()
        conn.close()
    except:
        config_rows = {'serial_inicial': '1'}
        total_recibos = 0
        next_serial_effective = 1

    return render_template('config/index.html', config=config_rows, total_recibos=total_recibos,
                           next_serial_effective=next_serial_effective,
                           lote_nombre=session.get('lote_nombre', 'Lote activo'),
                           message=message, error=error)

# =========================
# PDF REPORTS
# =========================
@app.route('/reportes/pdf')
def generar_pdf():
    if 'user_id' not in session:
        return auth_redirect('login', 'Inicia sesion.', 'warning')
    if not session.get('lote_id'):
        return redirect(url_for('select_lote'))
    lote_id     = session['lote_id']
    lote_nombre = session.get('lote_nombre', 'Arrocera')
    from fpdf import FPDF
    tipo      = request.args.get('tipo', 'gastos')
    desde_str = request.args.get('desde', '')
    hasta_str = request.args.get('hasta', '')

    # ── Palette ──────────────────────────────────────────────────────────────
    C_DEEP   = (27,  67,  50)
    C_FOREST = (45, 106,  79)
    C_SAGE   = (82, 183, 136)
    C_SMINT  = (238,247, 242)
    C_MINT   = (216,243, 220)
    C_GOLD   = (233,168,   0)
    C_WHITE  = (255,255, 255)
    C_DARK   = ( 27,  45,  30)
    C_MUTED  = (100,100, 100)

    TIPO_LABELS = {
        'gastos':       'Reporte de Gastos',
        'trabajadores': 'Reporte por Trabajador',
        'produccion':   'Reporte de Produccion',
        'semana':       'Reporte Semanal',
        'completo':     'Recibos Detallados',
        'rango':        'Reporte por Rango de Fechas',
    }

    def fmt_cop(v):
        try:   return '$ {:,.0f}'.format(float(v or 0)).replace(',', '.')
        except: return '$ 0'

    class ReportePDF(FPDF):
        PAGE_W = 210
        MARGIN = 10

        def header(self):
            if self.page_no() == 1:
                return
            self.set_fill_color(*C_DEEP)
            self.rect(0, 0, self.PAGE_W, 16, 'F')
            self.set_fill_color(*C_GOLD)
            self.rect(0, 16, self.PAGE_W, 2, 'F')
            self.set_xy(self.MARGIN, 3)
            self.set_font('Helvetica', 'B', 10)
            self.set_text_color(*C_WHITE)
            self.cell(130, 10, lote_nombre + '  -  Contabilidad Interna', align='L')
            self.set_xy(self.PAGE_W - 60, 3)
            self.set_font('Helvetica', '', 8)
            self.set_text_color(216, 243, 220)
            self.cell(55, 10, f'Gen: {datetime.now().strftime("%d/%m/%Y")}', align='R')
            self.set_y(22)

        def footer(self):
            if self.page_no() == 1:
                return
            self.set_y(-13)
            self.set_font('Helvetica', '', 8)
            self.set_text_color(*C_MUTED)
            self.cell(0, 10, f'Pag. {self.page_no()}  |  {lote_nombre} - Contabilidad Interna', align='C')

        def cover(self, tipo_label, desde_s='', hasta_s=''):
            self.set_fill_color(*C_DEEP)
            self.rect(0, 0, self.PAGE_W, 297, 'F')
            # Subtle dot pattern
            self.set_fill_color(45, 106, 79)
            for row in range(30):
                for col in range(22):
                    self.rect(col * 10 + 4.5, row * 10 + 4.5, 1, 1, 'F')
            # Gold bottom accent
            self.set_fill_color(*C_GOLD)
            self.rect(0, 292, self.PAGE_W, 5, 'F')
            # Card shadow
            cw, ch = 160, 90
            cxs = (self.PAGE_W - cw) / 2
            cys = 98
            self.set_fill_color(15, 40, 28)
            self.rect(cxs + 3, cys + 3, cw, ch, 'F')
            # White card
            self.set_fill_color(*C_WHITE)
            self.rect(cxs, cys, cw, ch, 'F')
            # Gold top bar on card
            self.set_fill_color(*C_GOLD)
            self.rect(cxs, cys, cw, 5, 'F')
            # Farm name
            self.set_xy(cxs, cys + 9)
            self.set_font('Helvetica', 'B', 20)
            self.set_text_color(*C_DEEP)
            self.cell(cw, 12, lote_nombre, align='C')
            # Sage divider
            self.set_draw_color(*C_SAGE)
            self.set_line_width(0.6)
            self.line(cxs + 25, cys + 23, cxs + cw - 25, cys + 23)
            # Report type
            self.set_xy(cxs, cys + 26)
            self.set_font('Helvetica', 'B', 14)
            self.set_text_color(*C_FOREST)
            self.cell(cw, 10, tipo_label, align='C')
            # Date range
            if desde_s and hasta_s:
                self.set_xy(cxs, cys + 38)
                self.set_font('Helvetica', '', 10)
                self.set_text_color(*C_MUTED)
                self.cell(cw, 8, f'Periodo: {desde_s}  al  {hasta_s}', align='C')
            # Year
            self.set_xy(cxs, cys + 52)
            self.set_font('Helvetica', '', 9)
            self.set_text_color(150, 150, 150)
            self.cell(cw, 7, str(datetime.now().year), align='C')
            # Footer disclaimer
            self.set_xy(0, 268)
            self.set_font('Helvetica', 'I', 9)
            self.set_text_color(*C_MINT)
            self.cell(self.PAGE_W, 8, 'Documento de uso exclusivo administrativo  -  Confidencial', align='C')

        def section_title(self, text):
            self.set_font('Helvetica', 'B', 13)
            self.set_text_color(*C_DEEP)
            self.cell(0, 9, text, new_x='LMARGIN', new_y='NEXT')
            self.set_draw_color(*C_SAGE)
            self.set_line_width(0.8)
            self.line(self.MARGIN, self.get_y(), self.PAGE_W - self.MARGIN, self.get_y())
            self.ln(4)
            self.set_text_color(*C_DARK)
            self.set_draw_color(*C_FOREST)
            self.set_line_width(0.3)

        def stat_row(self, stats):
            sw  = (self.PAGE_W - 2 * self.MARGIN) / len(stats)
            y0  = self.get_y()
            bgs = [C_DEEP, C_FOREST, C_SAGE]
            txs = [C_WHITE, C_WHITE, C_WHITE]
            for i, (lbl, val) in enumerate(stats):
                x = self.MARGIN + i * sw
                self.set_fill_color(*bgs[i % 3])
                self.rect(x, y0, sw - 2, 22, 'F')
                self.set_xy(x, y0 + 2)
                self.set_font('Helvetica', 'B', 14)
                self.set_text_color(*txs[i % 3])
                self.cell(sw - 2, 9, str(val), align='C')
                self.set_xy(x, y0 + 13)
                self.set_font('Helvetica', '', 7.5)
                self.cell(sw - 2, 6, lbl, align='C')
            self.set_y(y0 + 27)
            self.set_text_color(*C_DARK)

        def tbl_header(self, headers, widths):
            self.set_fill_color(*C_FOREST)
            self.set_text_color(*C_WHITE)
            self.set_font('Helvetica', 'B', 8)
            self.set_draw_color(*C_DEEP)
            self.set_line_width(0.3)
            for h, w in zip(headers, widths):
                self.cell(w, 7, h, border=1, fill=True)
            self.ln()

        def tbl_row(self, vals, widths, idx=0):
            self.set_fill_color(*(C_SMINT if idx % 2 == 0 else C_WHITE))
            self.set_text_color(*C_DARK)
            self.set_font('Helvetica', '', 7.5)
            self.set_draw_color(200, 220, 208)
            self.set_line_width(0.2)
            for v, w in zip(vals, widths):
                self.cell(w, 6, str(v), border=1, fill=True)
            self.ln()

        def tbl_total(self, label, valor, label_w, valor_w):
            self.set_fill_color(*C_DEEP)
            self.set_text_color(*C_WHITE)
            self.set_font('Helvetica', 'B', 8.5)
            self.set_draw_color(*C_DEEP)
            self.cell(label_w, 7, label, border=1, fill=True)
            self.cell(valor_w, 7, valor, border=1, fill=True, align='R')
            self.ln(3)

    conn   = get_db_connection()
    cursor = conn.cursor(dictionary=True)
    tipo_label = TIPO_LABELS.get(tipo, 'Reporte')

    pdf = ReportePDF(orientation='P', unit='mm', format='A4')
    pdf.set_auto_page_break(auto=True, margin=18)

    # Cover page
    pdf.add_page()
    pdf.cover(tipo_label, desde_str, hasta_str)

    # Data page(s)
    pdf.add_page()

    if tipo == 'gastos':
        if desde_str and hasta_str:
            cursor.execute("SELECT * FROM recibos WHERE lote_id=%s AND fecha BETWEEN %s AND %s ORDER BY serial",
                           (lote_id, desde_str, hasta_str))
            subtitle = f'Periodo: {desde_str}  al  {hasta_str}'
        else:
            cursor.execute("SELECT * FROM recibos WHERE lote_id=%s ORDER BY serial", (lote_id,))
            subtitle = 'Todos los registros'
        recibos = cursor.fetchall()
        total   = sum(float(r.get('neto_a_pagar') or 0) for r in recibos)

        pdf.section_title('Reporte de Gastos')
        pdf.set_font('Helvetica', '', 9)
        pdf.set_text_color(*C_MUTED)
        pdf.cell(0, 6, subtitle, new_x='LMARGIN', new_y='NEXT')
        pdf.ln(2)
        pdf.stat_row([('Total Recibos', str(len(recibos))), ('Total Gastado', fmt_cop(total))])

        hdrs = ['#', 'Fecha', 'Proveedor', 'NIT', 'Dirección / Ciudad', 'Concepto', 'Neto']
        wids = [14, 22, 38, 22, 44, 26, 24]
        pdf.tbl_header(hdrs, wids)
        for i, r in enumerate(recibos):
            neto = float(r.get('neto_a_pagar') or 0)
            fecha_fmt = r['fecha'].strftime('%d/%m/%Y') if r.get('fecha') else '-'
            dir_raw  = (r.get('direccion') or '').strip()
            ciudad_r = (r.get('ciudad') or '').strip()
            dir_ciudad = (dir_raw + (', ' + ciudad_r if ciudad_r else '')).strip(', ') or '-'
            pdf.tbl_row([str(r['serial']), fecha_fmt,
                         (r.get('proveedor') or '')[:22], (r.get('nit') or ''),
                         dir_ciudad[:26],
                         (r.get('concepto') or '')[:16], fmt_cop(neto)], wids, i)
        pdf.tbl_total(f'  TOTAL -- {len(recibos)} recibos', fmt_cop(total),
                      sum(wids[:-1]), wids[-1])

    elif tipo == 'trabajadores':
        workers_param = request.args.get('workers', '')
        nits_filter = [n.strip() for n in workers_param.split(',') if n.strip()] if workers_param else []

        if nits_filter:
            fmt_in = ','.join(['%s'] * len(nits_filter))
            cursor.execute(f"""
                SELECT proveedor, nit, COUNT(*) as num_recibos, SUM(neto_a_pagar) as total_pagado
                FROM recibos WHERE lote_id=%s AND nit IN ({fmt_in})
                GROUP BY proveedor, nit ORDER BY total_pagado DESC
            """, (lote_id, *nits_filter))
        else:
            cursor.execute("""
                SELECT proveedor, nit, COUNT(*) as num_recibos, SUM(neto_a_pagar) as total_pagado
                FROM recibos WHERE lote_id=%s GROUP BY proveedor, nit ORDER BY total_pagado DESC
            """, (lote_id,))
        rows  = cursor.fetchall()
        grand = sum(float(r.get('total_pagado') or 0) for r in rows)

        pdf.section_title('Reporte por Trabajador / Proveedor')
        pdf.set_font('Helvetica', '', 9)
        pdf.set_text_color(*C_MUTED)
        subtitle_w = f'Filtrado: {len(nits_filter)} trabajador(es) seleccionado(s)' if nits_filter else 'Consolidado total de pagos por persona'
        pdf.cell(0, 6, subtitle_w, new_x='LMARGIN', new_y='NEXT')
        pdf.ln(2)
        pdf.stat_row([('Trabajadores', str(len(rows))), ('Total General', fmt_cop(grand))])

        hdrs = ['Proveedor / Trabajador', 'NIT', 'Recibos', 'Total Pagado']
        wids = [90, 30, 20, 50]
        pdf.tbl_header(hdrs, wids)
        for i, r in enumerate(rows):
            t = float(r.get('total_pagado') or 0)
            pdf.tbl_row([(r.get('proveedor') or '')[:43], (r.get('nit') or ''),
                          str(r['num_recibos']), fmt_cop(t)], wids, i)
        pdf.tbl_total(f'  TOTAL -- {len(rows)} proveedores', fmt_cop(grand),
                      sum(wids[:-1]), wids[-1])

        # ── Detalle por trabajador ──────────────────────────────────────────
        for w in rows:
            nit_w = w.get('nit') or ''
            nombre_w = (w.get('proveedor') or 'Proveedor sin nombre')[:60]
            if nit_w:
                cursor.execute("""
                    SELECT serial, fecha, concepto, direccion, ciudad, telefono,
                           valor_operacion AS subtotal, rte_fte AS deducciones, neto_a_pagar
                    FROM recibos WHERE lote_id=%s AND nit=%s ORDER BY fecha, serial
                """, (lote_id, nit_w))
            else:
                cursor.execute("""
                    SELECT serial, fecha, concepto, direccion, ciudad, telefono,
                           valor_operacion AS subtotal, rte_fte AS deducciones, neto_a_pagar
                    FROM recibos WHERE lote_id=%s AND proveedor=%s AND (nit IS NULL OR nit='')
                    ORDER BY fecha, serial
                """, (lote_id, w.get('proveedor') or ''))
            det = cursor.fetchall()
            if not det:
                continue

            # Gather address from first receipt that has one
            dir_w     = next((d.get('direccion') or '' for d in det if d.get('direccion')), '')
            ciudad_w  = next((d.get('ciudad')    or '' for d in det if d.get('ciudad')),    '')
            tel_w     = next((d.get('telefono')  or '' for d in det if d.get('telefono')),  '')

            pdf.add_page()
            # Worker header card (2 lines: name+NIT, then address)
            pdf.set_fill_color(*C_FOREST)
            pdf.set_text_color(*C_WHITE)
            pdf.set_font('Helvetica', 'B', 12)
            pdf.rect(pdf.MARGIN, pdf.get_y(), 190, 10, 'F')
            pdf.set_xy(pdf.MARGIN + 2, pdf.get_y() + 1.5)
            pdf.cell(186, 7, f'{nombre_w}  |  NIT: {nit_w or "-"}', align='L')
            pdf.ln(11)
            # Address sub-bar
            pdf.set_fill_color(*C_SMINT)
            pdf.rect(pdf.MARGIN, pdf.get_y(), 190, 6, 'F')
            pdf.set_xy(pdf.MARGIN + 2, pdf.get_y() + 0.8)
            pdf.set_font('Helvetica', '', 8)
            pdf.set_text_color(*C_MUTED)
            addr_parts = []
            if dir_w:    addr_parts.append('Dir: ' + dir_w)
            if ciudad_w: addr_parts.append('Ciudad: ' + ciudad_w)
            if tel_w:    addr_parts.append('Tel: ' + tel_w)
            pdf.cell(186, 5, '   '.join(addr_parts) or 'Sin datos de contacto', align='L')
            pdf.ln(8)

            total_w = sum(float(d.get('neto_a_pagar') or 0) for d in det)
            total_sub = sum(float(d.get('subtotal') or 0) for d in det)
            total_ded = sum(float(d.get('deducciones') or 0) for d in det)
            pdf.stat_row([('Recibos', str(len(det))),
                          ('Subtotal', fmt_cop(total_sub)),
                          ('Neto Total', fmt_cop(total_w))])

            hdrs_d = ['Serial', 'Fecha', 'Concepto', 'Subtotal', 'Deduc.', 'Neto']
            wids_d = [18, 24, 68, 28, 24, 28]
            pdf.tbl_header(hdrs_d, wids_d)
            for i, d in enumerate(det):
                fecha_d = d['fecha'].strftime('%d/%m/%Y') if d.get('fecha') else '-'
                pdf.tbl_row([str(d['serial']), fecha_d,
                             (d.get('concepto') or '-')[:35],
                             fmt_cop(d.get('subtotal')), fmt_cop(d.get('deducciones')),
                             fmt_cop(d.get('neto_a_pagar'))], wids_d, i)
            pdf.tbl_total(f'  TOTAL -- {len(det)} recibos', fmt_cop(total_w),
                          sum(wids_d[:-1]), wids_d[-1])

    elif tipo == 'produccion':
        cursor.execute("SELECT * FROM cosechas WHERE lote_id=%s ORDER BY fecha DESC", (lote_id,))
        cosechas   = cursor.fetchall()
        tot_cargas = sum(int(c.get('cargas') or 0) for c in cosechas)
        tot_val    = sum(float(c.get('valor_total') or 0) for c in cosechas)
        cosechas_count = sum(1 for c in cosechas if (c.get('fase') or 'cosecha') == 'cosecha')
        siembras_count = sum(1 for c in cosechas if (c.get('fase') or 'cosecha') == 'siembra')

        pdf.section_title('Reporte de Produccion')
        pdf.set_font('Helvetica', '', 9)
        pdf.set_text_color(*C_MUTED)
        pdf.cell(0, 6, 'Historial de cosechas y siembras registradas', new_x='LMARGIN', new_y='NEXT')
        pdf.ln(2)
        pdf.stat_row([('Cosechas', str(cosechas_count)),
                      ('Siembras', str(siembras_count)),
                      ('Total Cargas', str(tot_cargas)),
                      ('Valor Total', fmt_cop(tot_val))])

        MET = {'al_voleo': 'Al voleo', 'sembradora': 'Sembradora',
               'labranza_minima': 'Labranza min.', 'otro': 'Otro'}
        hdrs = ['Fase', 'Fecha', 'Lote', 'Variedad', 'Metodo', 'Bultos', 'Cargas', 'Valor']
        wids = [18, 22, 28, 30, 24, 18, 18, 32]
        pdf.tbl_header(hdrs, wids)
        for i, c in enumerate(cosechas):
            fase_label = 'Cosecha' if (c.get('fase') or 'cosecha') == 'cosecha' else 'Siembra'
            cargas = int(c.get('cargas') or 0)
            val    = float(c.get('valor_total') or 0)
            fecha_fmt = c['fecha'].strftime('%d/%m/%Y') if c.get('fecha') else '-'
            met = MET.get(c.get('metodo_siembra') or '', '-')
            bultos = str(float(c.get('total_bultos') or 0) or '-')
            pdf.tbl_row([fase_label, fecha_fmt, (c.get('lote') or '')[:12],
                         (c.get('variedad_semilla') or '-')[:14],
                         met, bultos,
                         str(cargas) if cargas else '-', fmt_cop(val)], wids, i)
        pdf.tbl_total(f'  TOTALES -- {len(cosechas)} registros', fmt_cop(tot_val),
                      sum(wids[:-1]), wids[-1])

    elif tipo == 'semana':
        fecha_ref_str = request.args.get('fecha', date.today().isoformat())
        try:
            fecha_ref = date.fromisoformat(fecha_ref_str)
        except Exception:
            fecha_ref = date.today()
        inicio = fecha_ref - timedelta(days=fecha_ref.weekday())
        fin    = inicio + timedelta(days=6)
        cursor.execute("SELECT * FROM recibos WHERE lote_id=%s AND fecha BETWEEN %s AND %s ORDER BY serial",
                       (lote_id, inicio, fin))
        recibos_s = cursor.fetchall()
        total_s   = sum(float(r.get('neto_a_pagar') or 0) for r in recibos_s)

        pdf.section_title('Reporte Semanal')
        pdf.set_font('Helvetica', '', 9)
        pdf.set_text_color(*C_MUTED)
        pdf.cell(0, 6, f'Semana: {inicio.strftime("%d/%m/%Y")}  al  {fin.strftime("%d/%m/%Y")}',
                 new_x='LMARGIN', new_y='NEXT')
        pdf.ln(2)
        pdf.stat_row([('Recibos', str(len(recibos_s))), ('Total Semana', fmt_cop(total_s))])

        hdrs = ['#', 'Fecha', 'Proveedor', 'NIT', 'Dirección / Ciudad', 'Concepto', 'Neto']
        wids = [14, 22, 38, 22, 44, 26, 24]
        pdf.tbl_header(hdrs, wids)
        for i, r in enumerate(recibos_s):
            neto = float(r.get('neto_a_pagar') or 0)
            fecha_fmt = r['fecha'].strftime('%d/%m/%Y') if r.get('fecha') else '-'
            dir_raw  = (r.get('direccion') or '').strip()
            ciudad_s = (r.get('ciudad') or '').strip()
            dir_ciudad = (dir_raw + (', ' + ciudad_s if ciudad_s else '')).strip(', ') or '-'
            pdf.tbl_row([str(r['serial']), fecha_fmt,
                         (r.get('proveedor') or '')[:22], (r.get('nit') or ''),
                         dir_ciudad[:26],
                         (r.get('concepto') or '')[:16], fmt_cop(neto)], wids, i)
        pdf.tbl_total(f'  TOTAL SEMANA -- {len(recibos_s)} recibos', fmt_cop(total_s),
                      sum(wids[:-1]), wids[-1])

    elif tipo == 'completo':
        # Detailed per-receipt report with all concept lines
        if desde_str and hasta_str:
            cursor.execute(
                "SELECT * FROM recibos WHERE lote_id=%s AND fecha BETWEEN %s AND %s ORDER BY serial",
                (lote_id, desde_str, hasta_str))
            subtitle = 'Periodo: ' + desde_str + '  al  ' + hasta_str
        else:
            cursor.execute("SELECT * FROM recibos WHERE lote_id=%s ORDER BY serial", (lote_id,))
            subtitle = 'Todos los recibos registrados'
        recibos_c = cursor.fetchall()
        grand_total = sum(float(r.get('neto_a_pagar') or 0) for r in recibos_c)

        pdf.section_title('Recibos Detallados - Conceptos Completos')
        pdf.set_font('Helvetica', '', 9)
        pdf.set_text_color(*C_MUTED)
        pdf.cell(0, 6, subtitle, new_x='LMARGIN', new_y='NEXT')
        pdf.ln(2)
        pdf.stat_row([('Total Recibos', str(len(recibos_c))),
                      ('Total Neto Pagado', fmt_cop(grand_total))])

        W = 190
        for r in recibos_c:
            if pdf.get_y() > 240:
                pdf.add_page()
            neto      = float(r.get('neto_a_pagar') or 0)
            rte       = float(r.get('rte_fte') or 0)
            subtot    = neto + rte
            fecha_fmt = r['fecha'].strftime('%d/%m/%Y') if r.get('fecha') else '-'
            lineas = []
            if r.get('conceptos_json'):
                try:
                    lineas = json.loads(r['conceptos_json'])
                except Exception:
                    pass
            if not lineas and r.get('concepto'):
                lineas = [{'concepto': r['concepto'],
                           'valor': float(r.get('valor_operacion') or 0)}]
            # Header bar
            y0 = pdf.get_y()
            pdf.set_fill_color(*C_DEEP)
            pdf.rect(pdf.MARGIN, y0, W, 7, 'F')
            pdf.set_xy(pdf.MARGIN + 1, y0)
            pdf.set_font('Helvetica', 'B', 8)
            pdf.set_text_color(*C_WHITE)
            pdf.cell(18, 7, '# ' + str(r.get('serial', '')), fill=False)
            pdf.cell(26, 7, fecha_fmt, fill=False)
            pdf.cell(80, 7, (r.get('proveedor') or '')[:45], fill=False)
            pdf.cell(30, 7, 'NIT: ' + str(r.get('nit') or '-'), fill=False)
            pdf.set_text_color(*C_GOLD)
            pdf.cell(0, 7, 'Neto: ' + fmt_cop(neto), align='R', fill=False)
            pdf.ln(7)
            # Info sub-row
            pdf.set_fill_color(*C_SMINT)
            pdf.rect(pdf.MARGIN, pdf.get_y(), W, 5, 'F')
            pdf.set_xy(pdf.MARGIN + 1, pdf.get_y())
            pdf.set_font('Helvetica', '', 7)
            pdf.set_text_color(*C_MUTED)
            pdf.cell(0, 5,
                     '  Dir: '    + (r.get('direccion') or '-') +
                     '   Ciudad: ' + (r.get('ciudad') or '-') +
                     '   Tel: '   + (r.get('telefono') or '-'),
                     fill=False)
            pdf.ln(5)
            # Concept table header
            pdf.set_font('Helvetica', 'B', 7)
            pdf.set_fill_color(*C_FOREST)
            pdf.set_text_color(*C_WHITE)
            pdf.set_draw_color(*C_DEEP)
            pdf.set_line_width(0.2)
            pdf.cell(W - 30, 5, '  Concepto', border=1, fill=True)
            pdf.cell(30,     5, 'Valor',      border=1, fill=True, align='R')
            pdf.ln(5)
            for li, ln in enumerate(lineas):
                pdf.set_fill_color(*(C_SMINT if li % 2 == 0 else C_WHITE))
                pdf.set_text_color(*C_DARK)
                pdf.set_font('Helvetica', '', 7)
                pdf.cell(W - 30, 5,
                         '  ' + str(ln.get('concepto') or '')[:85],
                         border=1, fill=True)
                pdf.cell(30, 5,
                         fmt_cop(float(ln.get('valor') or 0)),
                         border=1, fill=True, align='R')
                pdf.ln(5)
            # Footer summary
            pdf.set_font('Helvetica', '', 7)
            pdf.set_text_color(*C_MUTED)
            pdf.cell(W - 30, 4.5, '')
            pdf.cell(30, 4.5, 'Subtotal: ' + fmt_cop(subtot), align='R')
            pdf.ln(4.5)
            if rte:
                pdf.cell(W - 30, 4.5, '')
                pdf.cell(30, 4.5, 'RTE/FTE: -' + fmt_cop(rte), align='R')
                pdf.ln(4.5)
            pdf.set_fill_color(*C_MINT)
            pdf.set_text_color(*C_DEEP)
            pdf.set_font('Helvetica', 'B', 7.5)
            pdf.cell(W - 30, 5, '', fill=False)
            pdf.cell(30, 5, 'NETO: ' + fmt_cop(neto), fill=True, align='R')
            pdf.ln(7)

    elif tipo == 'rango':
        # ── Comprehensive date-range statistical report ───────────────────────
        if not (desde_str and hasta_str):
            desde_str = date.today().replace(day=1).isoformat()
            hasta_str = date.today().isoformat()

        cursor.execute(
            "SELECT * FROM recibos WHERE lote_id=%s AND fecha BETWEEN %s AND %s ORDER BY serial",
            (lote_id, desde_str, hasta_str))
        recibos_r = cursor.fetchall()

        # Per-worker totals
        worker_totals = {}
        for r in recibos_r:
            key = (r.get('proveedor') or 'Sin nombre', r.get('nit') or '')
            if key not in worker_totals:
                worker_totals[key] = {'count': 0, 'total': 0.0}
            worker_totals[key]['count'] += 1
            worker_totals[key]['total'] += float(r.get('neto_a_pagar') or 0)
        worker_rows = sorted(worker_totals.items(), key=lambda x: -x[1]['total'])

        # Per-week totals
        from collections import defaultdict
        week_totals = defaultdict(float)
        for r in recibos_r:
            if r.get('fecha'):
                fd = r['fecha']
                week_start = fd - timedelta(days=fd.weekday())
                week_totals[week_start] += float(r.get('neto_a_pagar') or 0)

        grand_total = sum(float(r.get('neto_a_pagar') or 0) for r in recibos_r)
        avg_recibo  = grand_total / len(recibos_r) if recibos_r else 0
        rte_total   = sum(float(r.get('rte_fte') or 0) for r in recibos_r)

        # ── Summary stats ────────────────────────────────────────────────────
        pdf.section_title('Estadisticas Generales')
        pdf.set_font('Helvetica', '', 9)
        pdf.set_text_color(*C_MUTED)
        pdf.cell(0, 6,
                 'Periodo: ' + desde_str + '  al  ' + hasta_str,
                 new_x='LMARGIN', new_y='NEXT')
        pdf.ln(2)
        pdf.stat_row([
            ('Recibos', str(len(recibos_r))),
            ('Trabajadores', str(len(worker_rows))),
            ('Total Neto', fmt_cop(grand_total)),
            ('Promedio/recibo', fmt_cop(avg_recibo)),
        ])
        if rte_total:
            pdf.set_font('Helvetica', '', 8)
            pdf.set_text_color(*C_MUTED)
            pdf.ln(2)
            pdf.cell(0, 5,
                     '  RTE/FTE total descontado en el periodo: ' + fmt_cop(rte_total),
                     new_x='LMARGIN', new_y='NEXT')
        pdf.ln(4)

        # ── Breakdown by worker ──────────────────────────────────────────────
        pdf.section_title('Gastos por Trabajador / Proveedor')
        hdrs_w = ['Trabajador / Proveedor', 'NIT', 'Recibos', 'Total Pagado', '%']
        wids_w = [74, 28, 18, 38, 32]
        pdf.tbl_header(hdrs_w, wids_w)
        for i, ((nombre, nit), vals) in enumerate(worker_rows):
            pct = (vals['total'] / grand_total * 100) if grand_total else 0
            bar_chars = int(pct / 5)
            bar_str   = '#' * bar_chars + '.' * (20 - bar_chars)
            pdf.tbl_row([
                nombre[:38],
                nit[:16],
                str(vals['count']),
                fmt_cop(vals['total']),
                '{:.1f}% {}'.format(pct, bar_str[:10]),
            ], wids_w, i)
        pdf.tbl_total(
            '  TOTAL -- ' + str(len(worker_rows)) + ' proveedores',
            fmt_cop(grand_total),
            sum(wids_w[:-1]), wids_w[-1])

        # ── Breakdown by week ────────────────────────────────────────────────
        if week_totals:
            pdf.ln(6)
            if pdf.get_y() > 220:
                pdf.add_page()
            pdf.section_title('Evolucion Semanal de Gastos')
            hdrs_wk = ['Semana (inicio)', 'Total Pagado', 'Barra de progreso']
            wids_wk = [45, 40, 105]
            pdf.tbl_header(hdrs_wk, wids_wk)
            max_week = max(week_totals.values()) if week_totals else 1
            for i, (wdate, wtotal) in enumerate(sorted(week_totals.items())):
                bar_len = int(wtotal / max_week * 50) if max_week else 0
                bar_str = '|' * bar_len
                pdf.tbl_row([
                    wdate.strftime('%d/%m/%Y'),
                    fmt_cop(wtotal),
                    bar_str,
                ], wids_wk, i)
            pdf.tbl_total(
                '  TOTAL DEL PERIODO',
                fmt_cop(grand_total),
                sum(wids_wk[:-1]), wids_wk[-1])

        # ── Detailed receipts table ──────────────────────────────────────────
        if recibos_r:
            pdf.ln(6)
            if pdf.get_y() > 200:
                pdf.add_page()
            pdf.section_title('Detalle de Recibos')
            hdrs_d = ['Serial', 'Fecha', 'Proveedor', 'NIT', 'Dirección / Ciudad', 'Concepto', 'RTE/FTE', 'Neto']
            wids_d = [14, 20, 36, 20, 38, 22, 18, 22]
            pdf.tbl_header(hdrs_d, wids_d)
            for i, r in enumerate(recibos_r):
                neto_r    = float(r.get('neto_a_pagar') or 0)
                rte_r     = float(r.get('rte_fte') or 0)
                fecha_fmt = r['fecha'].strftime('%d/%m/%Y') if r.get('fecha') else '-'
                concepto_str = ''
                if r.get('conceptos_json'):
                    try:
                        cjs = json.loads(r['conceptos_json'])
                        if cjs:
                            concepto_str = str(cjs[0].get('concepto') or '')
                    except Exception:
                        pass
                if not concepto_str:
                    concepto_str = r.get('concepto') or ''
                dir_raw   = (r.get('direccion') or '').strip()
                ciudad_rr = (r.get('ciudad') or '').strip()
                dir_ciudad_r = (dir_raw + (', ' + ciudad_rr if ciudad_rr else '')).strip(', ') or '-'
                pdf.tbl_row([
                    str(r.get('serial', '')),
                    fecha_fmt,
                    (r.get('proveedor') or '')[:20],
                    (r.get('nit') or '')[:12],
                    dir_ciudad_r[:22],
                    concepto_str[:14],
                    fmt_cop(rte_r) if rte_r else '-',
                    fmt_cop(neto_r),
                ], wids_d, i)
            pdf.tbl_total(
                '  TOTAL -- ' + str(len(recibos_r)) + ' recibos',
                fmt_cop(grand_total),
                sum(wids_d[:-1]), wids_d[-1])

    cursor.close(); conn.close()
    nombre_archivo = 'reporte_' + tipo + '_' + date.today().isoformat() + '.pdf'
    from flask import Response as FlaskResp
    pdf_bytes = pdf.output()
    return FlaskResp(bytes(pdf_bytes), mimetype='application/pdf',
                     headers={'Content-Disposition': f'inline; filename={nombre_archivo}'})


# =========================
# EXCEL REPORTS
# =========================
@app.route('/reportes/excel')
def generar_excel():
    if 'user_id' not in session:
        return auth_redirect('login', 'Inicia sesion.', 'warning')
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    except ImportError:
        return "openpyxl no instalado. Ejecuta: pip install openpyxl", 500

    if not session.get('lote_id'):
        return redirect(url_for('select_lote'))
    lote_id_xl = session['lote_id']

    tipo = request.args.get('tipo', 'recibos')
    desde_str = request.args.get('desde', '')
    hasta_str = request.args.get('hasta', '')

    wb = openpyxl.Workbook()
    ws = wb.active

    clr_header = PatternFill('solid', fgColor='2D6A4F')
    clr_alt    = PatternFill('solid', fgColor='F0F8F4')
    fnt_header = Font(bold=True, color='FFFFFF', size=10)
    fnt_title  = Font(bold=True, color='1B4332', size=13)
    fnt_bold   = Font(bold=True, size=9)
    fnt_norm   = Font(size=9)
    aln_center = Alignment(horizontal='center', vertical='center')
    aln_right  = Alignment(horizontal='right', vertical='center')
    thin_side  = Side(style='thin', color='CCCCCC')
    thin_border = Border(left=thin_side, right=thin_side, top=thin_side, bottom=thin_side)

    def hdr_cell(cell, text):
        cell.value = text; cell.font = fnt_header; cell.fill = clr_header
        cell.alignment = aln_center; cell.border = thin_border

    def data_cell(cell, text, alt=False, bold=False, align='left'):
        cell.value = text
        cell.font = fnt_bold if bold else fnt_norm
        if alt: cell.fill = clr_alt
        cell.alignment = Alignment(horizontal=align, vertical='center')
        cell.border = thin_border

    conn = get_db_connection()
    cursor = conn.cursor(dictionary=True)

    if tipo == 'recibos':
        ws.title = 'Recibos'
        ws.merge_cells('A1:J1')
        ws['A1'].value = 'Contabilidad Arroceras - Listado de Recibos'
        ws['A1'].font = fnt_title
        ws['A1'].alignment = aln_center
        if desde_str and hasta_str:
            cursor.execute("SELECT * FROM recibos WHERE lote_id=%s AND fecha BETWEEN %s AND %s ORDER BY serial", (lote_id_xl, desde_str, hasta_str))
        else:
            cursor.execute("SELECT * FROM recibos WHERE lote_id=%s ORDER BY serial", (lote_id_xl,))
        recibos = cursor.fetchall()
        headers = ['Serial','Fecha','Proveedor','NIT','Dirección','Teléfono','Ciudad','Concepto','Valor Operación','Neto a Pagar']
        for col, h in enumerate(headers, 1):
            hdr_cell(ws.cell(3, col), h)
        ws.row_dimensions[3].height = 20
        for i, r in enumerate(recibos, 4):
            alt = (i % 2 == 0)
            fecha_val = r['fecha'].strftime('%d/%m/%Y') if r.get('fecha') else ''
            row = [r['serial'], fecha_val, r.get('proveedor',''), r.get('nit',''),
                   r.get('direccion',''), r.get('telefono',''), r.get('ciudad',''),
                   r.get('concepto',''), float(r.get('valor_operacion') or 0), float(r.get('neto_a_pagar') or 0)]
            for col, val in enumerate(row, 1):
                data_cell(ws.cell(i, col), val, alt=alt, align='right' if col >= 9 else 'left')
        # Total row
        tot = sum(float(r.get('neto_a_pagar') or 0) for r in recibos)
        trow = len(recibos) + 4
        ws.cell(trow, 8).value = f'TOTAL ({len(recibos)} recibos)'; ws.cell(trow,8).font = fnt_bold
        ws.cell(trow, 10).value = tot; ws.cell(trow,10).font = fnt_bold; ws.cell(trow,10).fill = clr_header; ws.cell(trow,10).font = Font(bold=True, color='FFFFFF', size=9)
        ws.column_dimensions['C'].width = 28; ws.column_dimensions['H'].width = 40
        for c in ['A','B','D','E','F','G','I','J']: ws.column_dimensions[c].width = 16

    elif tipo == 'trabajadores':
        workers_param_xl = request.args.get('workers', '')
        nits_filter_xl = [n.strip() for n in workers_param_xl.split(',') if n.strip()] if workers_param_xl else []

        ws.title = 'Por Trabajador'
        ws.merge_cells('A1:D1')
        ws['A1'].value = 'Gastos por Trabajador / Proveedor'
        ws['A1'].font = fnt_title; ws['A1'].alignment = aln_center

        if nits_filter_xl:
            fmt_in_xl = ','.join(['%s'] * len(nits_filter_xl))
            cursor.execute(f"""SELECT proveedor, nit, COUNT(*) as num_recibos, SUM(neto_a_pagar) as total_pagado
                              FROM recibos WHERE lote_id=%s AND nit IN ({fmt_in_xl})
                              GROUP BY proveedor, nit ORDER BY total_pagado DESC""",
                           (lote_id_xl, *nits_filter_xl))
        else:
            cursor.execute("""SELECT proveedor, nit, COUNT(*) as num_recibos, SUM(neto_a_pagar) as total_pagado
                              FROM recibos WHERE lote_id=%s GROUP BY proveedor, nit ORDER BY total_pagado DESC""",
                           (lote_id_xl,))
        rows = cursor.fetchall()
        for col, h in enumerate(['Proveedor','NIT','Num. Recibos','Total Pagado'], 1):
            hdr_cell(ws.cell(3, col), h)
        grand = 0
        for i, r in enumerate(rows, 4):
            t = float(r.get('total_pagado') or 0); grand += t
            for col, val in enumerate([r.get('proveedor',''), r.get('nit',''), r['num_recibos'], t], 1):
                data_cell(ws.cell(i, col), val, alt=(i%2==0), align='right' if col >= 3 else 'left')
        trow2 = len(rows) + 4
        ws.cell(trow2, 3).value = 'TOTAL'; ws.cell(trow2,3).font = fnt_bold
        ws.cell(trow2, 4).value = grand; ws.cell(trow2,4).font = Font(bold=True, color='FFFFFF', size=9); ws.cell(trow2,4).fill = clr_header
        ws.column_dimensions['A'].width = 35; ws.column_dimensions['B'].width = 18
        ws.column_dimensions['C'].width = 15; ws.column_dimensions['D'].width = 20

        # Hoja de detalle por trabajador
        for w in rows:
            nit_w_xl = w.get('nit') or ''
            nombre_w_xl = (w.get('proveedor') or 'Sin nombre')[:30]
            if nit_w_xl:
                cursor.execute("""SELECT serial, fecha, concepto,
                                         valor_operacion AS subtotal, rte_fte AS deducciones, neto_a_pagar
                                   FROM recibos WHERE lote_id=%s AND nit=%s ORDER BY fecha, serial""",
                               (lote_id_xl, nit_w_xl))
            else:
                cursor.execute("""SELECT serial, fecha, concepto,
                                         valor_operacion AS subtotal, rte_fte AS deducciones, neto_a_pagar
                                   FROM recibos WHERE lote_id=%s AND proveedor=%s AND (nit IS NULL OR nit='')
                                   ORDER BY fecha, serial""",
                               (lote_id_xl, w.get('proveedor') or ''))
            det_xl = cursor.fetchall()
            if not det_xl:
                continue
            ws2 = wb.create_sheet(title=nombre_w_xl[:28])
            ws2.merge_cells('A1:F1')
            ws2['A1'].value = nombre_w_xl + (f'  |  NIT: {nit_w_xl}' if nit_w_xl else '')
            ws2['A1'].font = fnt_title; ws2['A1'].alignment = aln_center
            for col, h in enumerate(['Serial','Fecha','Concepto','Subtotal','Deducciones','Neto a Pagar'], 1):
                hdr_cell(ws2.cell(3, col), h)
            tot_w = 0
            for i, d in enumerate(det_xl, 4):
                neto_v = float(d.get('neto_a_pagar') or 0); tot_w += neto_v
                fecha_d = d['fecha'].strftime('%d/%m/%Y') if d.get('fecha') else ''
                row_d = [d['serial'], fecha_d, d.get('concepto','')[:40],
                         float(d.get('subtotal') or 0), float(d.get('deducciones') or 0), neto_v]
                for col, val in enumerate(row_d, 1):
                    data_cell(ws2.cell(i, col), val, alt=(i%2==0), align='right' if col >= 4 else 'left')
            trow_w = len(det_xl) + 4
            ws2.cell(trow_w,5).value = 'NETO TOTAL'; ws2.cell(trow_w,5).font = fnt_bold
            ws2.cell(trow_w,6).value = tot_w; ws2.cell(trow_w,6).font = Font(bold=True, color='FFFFFF',size=9); ws2.cell(trow_w,6).fill = clr_header
            ws2.column_dimensions['C'].width = 40
            for c2 in ['A','B','D','E','F']: ws2.column_dimensions[c2].width = 18

    elif tipo == 'produccion':
        ws.title = 'Produccion'
        ws.merge_cells('A1:G1')
        ws['A1'].value = 'Registro de Cosechas y Producción'
        ws['A1'].font = fnt_title; ws['A1'].alignment = aln_center
        cursor.execute("SELECT * FROM cosechas WHERE lote_id=%s ORDER BY fecha DESC", (lote_id_xl,))
        cosechas = cursor.fetchall()
        for col, h in enumerate(['Fecha','Lote','Hectáreas','Cargas','Kg Total','Precio/Carga','Valor Total','Observaciones'], 1):
            hdr_cell(ws.cell(3, col), h)
        tot_c = tot_k = tot_v = 0
        for i, c in enumerate(cosechas, 4):
            cargas = int(c.get('cargas') or 0); kg = float(c.get('kg_total') or 0); val = float(c.get('valor_total') or 0)
            tot_c += cargas; tot_k += kg; tot_v += val
            fecha_val = c['fecha'].strftime('%d/%m/%Y') if c.get('fecha') else ''
            row_c = [fecha_val, c.get('lote',''), float(c.get('hectareas') or 20), cargas, kg,
                     float(c.get('precio_carga') or 0), val, c.get('observaciones','')]
            for col, val_c in enumerate(row_c, 1):
                data_cell(ws.cell(i, col), val_c, alt=(i%2==0), align='right' if col in [3,4,5,6,7] else 'left')
        trow3 = len(cosechas) + 4
        ws.cell(trow3,3).value = 'TOTALES'; ws.cell(trow3,3).font = fnt_bold
        for col, val_t in zip([4,5,7], [tot_c, tot_k, tot_v]):
            ws.cell(trow3, col).value = val_t; ws.cell(trow3,col).font = Font(bold=True, color='FFFFFF',size=9); ws.cell(trow3,col).fill = clr_header
        for c2 in ['A','B','C','D','E','F','G']: ws.column_dimensions[c2].width = 18
        ws.column_dimensions['H'].width = 35

    cursor.close(); conn.close()

    import io
    buf = io.BytesIO()
    wb.save(buf); buf.seek(0)
    nombre = f'reporte_{tipo}_{date.today().isoformat()}.xlsx'
    from flask import Response as FlaskResp, send_file
    return send_file(buf, as_attachment=True, download_name=nombre,
                     mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')


def init_templates():
    """Create template directories and write all HTML files on first run."""
    BASE = os.path.dirname(os.path.abspath(__file__))
    dirs = [
        os.path.join(BASE, 'data'),
        os.path.join(BASE, 'uploads'),
        os.path.join(BASE, 'templates', 'recibos'),
        os.path.join(BASE, 'templates', 'reportes'),
        os.path.join(BASE, 'templates', 'produccion'),
        os.path.join(BASE, 'templates', 'config'),
        os.path.join(BASE, 'templates', 'workers'),
    ]
    for d in dirs:
        os.makedirs(d, exist_ok=True)

    def w(path, content):
        full = os.path.join(BASE, path)
        if not os.path.exists(full) or os.path.getsize(full) == 0:
            with open(full, 'w', encoding='utf-8') as f:
                f.write(content)
            print(f'  created: {path}')

    # ── data/trabajadores.json ───────────────────────────────
    w('data/trabajadores.json', '[{"key":"elias_castano","nombre":"Elias Casta\u00f1o","alias":["paisa"],"nit":"5962798","direccion":"El Tambo","telefono":"3173041545","ciudad":"Natagaima","concepto_habitual":"Pago por cuidar el motor x semana en el lote El Mangon (motor lister de 38 hp a diesel)","valor_habitual":350000},{"key":"eliodoro_aroca","nombre":"Eliodoro Aroca Alape","alias":["delio"],"nit":"93345400","direccion":"El Tambo","telefono":"","ciudad":"Natagaima","concepto_habitual":"Pago de moje corrido x semana en el lote El Mangon en 20 hectareas a motor","valor_habitual":370000},{"key":"alexander_botache","nombre":"Alexander Botache Yara","alias":["alex"],"nit":"93470959","direccion":"El Tambo","telefono":"3157051538","ciudad":"Natagaima","concepto_habitual":"Pago de transporte de 24 bidones de ACPM de Angostura a El Lote El Mangon","valor_habitual":100000},{"key":"jorge_delvasto","nombre":"Jorge Enrique Delvasto","alias":[],"nit":"93471018","direccion":"El Tambo","telefono":"3164936224","ciudad":"Natagaima","concepto_habitual":"Arrego de via + transporte canal de la Cevedo","valor_habitual":null},{"key":"silvio_zorrillo","nombre":"Silvio Zorrillo Carrillo","alias":[],"nit":"93470637","direccion":"El Tambo","telefono":"","ciudad":"Natagaima","concepto_habitual":"Despalillada x jornales x semana","valor_habitual":null},{"key":"gilberto_guarnizo","nombre":"Gilberto Guarnizo Rojas","alias":[],"nit":"93152907","direccion":"Cra 9A #11B Abonanza","telefono":"3123978471","ciudad":"Salda\u00f1a","concepto_habitual":"Carga de camiones de arroz paddy en Salda\u00f1a en corta del Tambo","valor_habitual":null},{"key":"joany_cuevas","nombre":"Joany Andres Cuevas Vanegas","alias":[],"nit":"1081182208","direccion":"Guasimal","telefono":"3001392542","ciudad":"Natagaima","concepto_habitual":"Pago de corta de arroz en el Lote El Mangon","valor_habitual":null},{"key":"humberto_yara","nombre":"Humberto Yara Tique","alias":[],"nit":"11293591","direccion":"El Tambo","telefono":"3166422455","ciudad":"Natagaima","concepto_habitual":"Asofructo parcela Lote El Mangon","valor_habitual":null},{"key":"jose_medina","nombre":"Jose Maria Medina Santos","alias":["chepe"],"nit":"93151067","direccion":"Carrera 10 #11","telefono":"3132429292","ciudad":"Natagaima","concepto_habitual":"Reparaciones motor lister de 38 hp a diesel","valor_habitual":null},{"key":"vicente_andrade","nombre":"Vicente Andrade Lozano","alias":["agronomo"],"nit":"","direccion":"El Tambo","telefono":"","ciudad":"Natagaima","concepto_habitual":"Pago porcentaje agr\u00f3nomo final de corta de arroz","valor_habitual":null},{"key":"fernando_lozano","nombre":"Fernando Lozano","alias":[],"nit":"","direccion":"El Tambo","telefono":"","ciudad":"Natagaima","concepto_habitual":"Bordeada a motosierra en el lote El Mangon","valor_habitual":880000},{"key":"raquel_cumaco","nombre":"Raquel Cumaco Tacuma","alias":[],"nit":"65789347","direccion":"El Tambo","telefono":"3154910540","ciudad":"Natagaima","concepto_habitual":"Cancelaci\u00f3n de asofructo","valor_habitual":null},{"key":"cabildo_tambo","nombre":"Cabildo de Tambo","alias":["cabildo"],"nit":"809005671","direccion":"El Tambo","telefono":"3123525731","ciudad":"Natagaima","concepto_habitual":"Cancelaci\u00f3n de asofructo","valor_habitual":null}]')

    # ── templates/recibos/lista.html ─────────────────────────
    w('templates/recibos/lista.html', """\
{%% extends "base.html" %%}
{%% block title %%}Recibos | Contabilidad Arroceras{%% endblock %%}
{%% block content %%}
<div class="page-hero">
  <div class="container page-hero__inner">
    <div class="page-hero__icon"><i class="fa-solid fa-list-check"></i></div>
    <div>
      <h1 class="page-hero__title">Recibos</h1>
      <p class="page-hero__sub">{{{ recibos|length }}} recibo(s) registrado(s).</p>
    </div>
    <div class="page-hero__actions">
      <a href="{{{ url_for('nuevo_recibo') }}}" class="button button--primary"><i class="fa-solid fa-plus"></i> Nuevo recibo</a>
      <a href="{{{ url_for('nuevo_recibo_lote') }}}" class="button button--ghost"><i class="fa-solid fa-users-between-lines"></i> Por lote</a>
      <a href="{{{ url_for('exportar_txt') }}}" class="button button--ghost"><i class="fa-solid fa-file-export"></i> Exportar TXT</a>
    </div>
  </div>
</div>
<section class="section"><div class="container">
  {%% if recibos %%}
  <div class="lista-toolbar">
    <div class="lista-search-wrap">
      <i class="fa-solid fa-magnifying-glass lista-search-icon"></i>
      <input type="text" id="searchInput" placeholder="Buscar proveedor, NIT, concepto\u2026" class="lista-search-input">
    </div>
    <div class="lista-summary">Total: <strong id="totalVisible">{{{ recibos|length }}}</strong> | Suma: <strong id="sumaVisible">$ {{{ "{:,.0f}".format(recibos|sum(attribute='neto_a_pagar')).replace(",",".") }}}</strong></div>
  </div>
  <div class="table-wrap">
    <table class="data-table">
      <thead><tr><th>Serial</th><th>Fecha</th><th>Proveedor</th><th>NIT</th><th>Concepto</th><th>Neto a pagar</th></tr></thead>
      <tbody id="recibosBody">
        {%% for r in recibos %%}
        <tr class="recibo-row" style="cursor:pointer" onclick="location.href='{{{ url_for('detalle_recibo', serial=r.serial) }}}'"
            data-proveedor="{{{ r.proveedor|lower }}}" data-nit="{{{ r.nit or '' }}}" data-concepto="{{{ r.concepto|lower }}}" data-valor="{{{ r.neto_a_pagar or 0 }}}">
          <td><span class="serial-badge">{{{ r.serial }}}</span></td>
          <td class="txt-muted">{{{ r.fecha.strftime('%d/%m/%Y') if r.fecha else '\u2014' }}}</td>
          <td><strong>{{{ r.proveedor }}}</strong></td>
          <td class="txt-muted small">{{{ r.nit or '\u2014' }}}</td>
          <td class="concepto-cell small">{{{ r.concepto }}}</td>
          <td class="valor-cell">{%% if r.neto_a_pagar %%}<strong>$ {{{ "{:,.0f}".format(r.neto_a_pagar).replace(",",".") }}}</strong>{%% else %%}\u2014{%% endif %%}</td>
        </tr>
        {%% endfor %%}
      </tbody>
      <tfoot><tr class="table-foot"><td colspan="5" style="text-align:right;font-weight:600">Total visible:</td><td class="valor-cell" id="footTotal">$ {{{ "{:,.0f}".format(recibos|sum(attribute='neto_a_pagar')).replace(",",".") }}}</td></tr></tfoot>
    </table>
  </div>
  <p id="noResultsMsg" style="display:none;text-align:center;padding:2rem;color:var(--txt-muted)"><i class="fa-solid fa-magnifying-glass"></i> Sin resultados.</p>
  {%% else %%}
  <div class="empty-state"><i class="fa-solid fa-file-circle-plus empty-state__icon"></i><h3>No hay recibos a\u00fan</h3><p>Crea el primer recibo.</p><a href="{{{ url_for('nuevo_recibo') }}}" class="button button--primary"><i class="fa-solid fa-plus"></i> Crear primer recibo</a></div>
  {%% endif %%}
</div></section>
{%% endblock %%}
{%% block scripts %%}
<script>
(function(){
  const inp=document.getElementById('searchInput'),rows=document.querySelectorAll('.recibo-row'),noMsg=document.getElementById('noResultsMsg'),ft=document.getElementById('footTotal'),sv=document.getElementById('sumaVisible'),tv=document.getElementById('totalVisible');
  if(!inp)return;
  inp.addEventListener('input',function(){
    const q=this.value.toLowerCase().trim();let vis=0,sum=0;
    rows.forEach(r=>{const ok=!q||[r.dataset.proveedor,r.dataset.nit,r.dataset.concepto].join(' ').includes(q);r.style.display=ok?'':'none';if(ok){vis++;sum+=parseFloat(r.dataset.valor)||0;}});
    noMsg.style.display=vis===0?'block':'none';tv.textContent=vis;
    const fmt=sum.toLocaleString('es-CO',{minimumFractionDigits:0,maximumFractionDigits:0});
    if(ft)ft.textContent='$ '+fmt;if(sv)sv.textContent='$ '+fmt;
  });
})();
</script>
{%% endblock %%}
""".replace('{{{','{{').replace('}}}','}}').replace('%%','{').replace('%%','}').replace('{%','{%').replace('%}','%}'))

    # ── templates/recibos/nuevo.html ─────────────────────────
    w('templates/recibos/nuevo.html', _tpl_nuevo_recibo())

    # ── templates/recibos/lote.html ──────────────────────────
    w('templates/recibos/lote.html', _tpl_lote())

    # ── templates/recibos/detalle.html ───────────────────────
    w('templates/recibos/detalle.html', _tpl_detalle())

    # ── templates/reportes/index.html ────────────────────────
    w('templates/reportes/index.html', _tpl_reportes())

    # ── templates/reportes/semana.html ───────────────────────
    w('templates/reportes/semana.html', _tpl_semana())

    # ── templates/produccion/index.html ──────────────────────
    w('templates/produccion/index.html', _tpl_produccion_lista())

    # ── templates/produccion/nueva.html ──────────────────────
    w('templates/produccion/nueva.html', _tpl_produccion_nueva())

    # ── templates/config/index.html ──────────────────────────
    w('templates/config/index.html', _tpl_config())

    print('Templates OK.')


# ── Template helpers ──────────────────────────────────────────────────────
def _j(s):
    """Replace placeholder brackets so Jinja2 syntax survives Python strings."""
    return s.replace('[[', '{%').replace(']]', '%}').replace('<<<', '{{').replace('>>>', '}}')


def _tpl_nuevo_recibo():
    return _j("""\
[[ extends "base.html" ]]
[[ block title ]]Nuevo Recibo | Contabilidad Arroceras[[ endblock ]]
[[ block head ]]<script src="<<<url_for('static', filename='css/js/autocomplete.js')>>>" defer></script>[[ endblock ]]
[[ block content ]]
<div class="page-hero">
  <div class="container page-hero__inner">
    <div class="page-hero__icon"><i class="fa-solid fa-file-invoice-dollar"></i></div>
    <div>
      <h1 class="page-hero__title">Nuevo Recibo</h1>
      <p class="page-hero__sub">Registra un pago. Serial asignado según la fecha.</p>
    </div>
    <div class="page-hero__actions">
      <a href="<<<url_for('lista_recibos')>>>" class="button button--ghost"><i class="fa-solid fa-list"></i> Ver recibos</a>
      <a href="<<<url_for('nuevo_recibo_lote')>>>" class="button button--ghost"><i class="fa-solid fa-users-between-lines"></i> Por lote</a>
    </div>
  </div>
</div>
<section class="section"><div class="container" style="max-width:860px">

  [[ if success ]]<div class="alert alert--success"><i class="fa-solid fa-circle-check"></i> <<<success>>></div>[[ endif ]]
  [[ if error ]]<div class="alert alert--danger"><i class="fa-solid fa-circle-xmark"></i> <<<error>>></div>[[ endif ]]
  [[ if warning ]]
  <div class="alert alert--warning" id="warningAlert">
    <i class="fa-solid fa-triangle-exclamation"></i> <<<warning>>>
    <div style="margin-top:.6rem;display:flex;gap:.6rem;flex-wrap:wrap">
      <button type="button" class="button button--sm button--primary" id="btnForce"><i class="fa-solid fa-floppy-disk"></i> Guardar de todas formas</button>
      <button type="button" class="button button--sm button--ghost-dark" onclick="document.getElementById('warningAlert').style.display='none'">Cancelar</button>
    </div>
  </div>
  [[ endif ]]

  <div class="trabajador-search-box">
    <label class="search-label"><i class="fa-solid fa-magnifying-glass"></i> Buscar trabajador por nombre o apodo</label>
    <div class="search-input-wrap">
      <input type="text" id="buscarTrabajador" placeholder="Ej: paisa, delio, Silvio\u2026" autocomplete="off" class="search-input">
      <ul id="trabajadorSuggestions" class="suggestions-list"></ul>
    </div>
    <p class="field-hint">Al seleccionar se autocompletan NIT, dirección, ciudad, teléfono, concepto y valor.</p>
  </div>

  <form method="POST" id="reciboForm" class="form-card">
    <input type="hidden" name="force" id="forceField" value="false">
    <div class="form-section">
      <h3 class="form-section__title"><i class="fa-solid fa-hashtag"></i> Identificación</h3>
      <div class="form-grid form-grid--3">
        <div class="form-group">
          <label for="serial">Serial <span class="req-star">*</span></label>
          <input type="number" id="serial" name="serial" value="<<<form_data.get('serial', next_serial)>>>" min="1" required>
          <p class="field-hint">Próximo sugerido: <strong><<<next_serial>>></strong></p>
        </div>
        <div class="form-group">
          <label for="fecha">Fecha</label>
          <input type="date" id="fecha" name="fecha" value="<<<form_data.get('fecha', today)>>>">
        </div>
      </div>
    </div>
    <div class="form-section">
      <h3 class="form-section__title"><i class="fa-solid fa-user-tie"></i> Proveedor / Trabajador</h3>
      <div class="form-grid form-grid--2">
        <div class="form-group">
          <label for="proveedor">Nombre completo <span class="req-star">*</span></label>
          <input type="text" id="proveedor" name="proveedor" value="<<<form_data.get('proveedor','')>>>" placeholder="Nombre real (no apodo)" required>
        </div>
        <div class="form-group">
          <label for="nit">NIT / Cédula</label>
          <input type="text" id="nit" name="nit" value="<<<form_data.get('nit','')>>>" placeholder="Ej: 93470637">
        </div>
        <div class="form-group">
          <label for="direccion">Dirección</label>
          <input type="text" id="direccion" name="direccion" value="<<<form_data.get('direccion','')>>>" placeholder="Ej: El Tambo">
        </div>
        <div class="form-group">
          <label for="telefono">Teléfono</label>
          <input type="text" id="telefono" name="telefono" value="<<<form_data.get('telefono','')>>>" placeholder="Ej: 3001234567">
        </div>
        <div class="form-group">
          <label for="ciudad">Ciudad</label>
          <input type="text" id="ciudad" name="ciudad" value="<<<form_data.get('ciudad','')>>>" placeholder="Ej: Natagaima">
        </div>
      </div>
    </div>
    <div class="form-section">
      <h3 class="form-section__title"><i class="fa-solid fa-file-lines"></i> Concepto y valor</h3>
      <div class="form-group">
        <label for="concepto">Concepto <span class="req-star">*</span></label>
        <textarea id="concepto" name="concepto" rows="3" required placeholder="Descripción detallada del trabajo o pago\u2026"><<<form_data.get('concepto','')>>></textarea>
        <p class="field-hint"><i class="fa-solid fa-circle-xmark" style="color:#c0392b"></i> No registrar compras de aceite ni ACPM directo. Solo el <em>transporte</em> de ACPM.</p>
      </div>
      <div class="form-grid form-grid--2">
        <div class="form-group">
          <label for="valor_operacion">Valor operación (COP)</label>
          <input type="text" id="valor_operacion" name="valor_operacion" value="<<<form_data.get('valor_operacion','')>>>" placeholder="Ej: 350.000" inputmode="numeric" oninput="syncNeto(this)">
        </div>
        <div class="form-group">
          <label for="neto_a_pagar">Neto a pagar (COP)</label>
          <input type="text" id="neto_a_pagar" name="neto_a_pagar" value="<<<form_data.get('neto_a_pagar','')>>>" placeholder="Igual al valor si no hay descuentos" inputmode="numeric">
          <p class="field-hint">Si se deja vacío se usa el valor de operación.</p>
        </div>
      </div>
    </div>

    <div class="form-section--collapsible">
      <button type="button" class="form-section__toggle" onclick="toggleCalc()">
        <i class="fa-solid fa-calculator"></i> Calculadora de jornales
        <i class="fa-solid fa-chevron-down toggle-icon" id="calcIcon"></i>
      </button>
      <div class="calc-body" id="calcBody" style="display:none">
        <div class="form-grid form-grid--3">
          <div class="form-group"><label>N° trabajadores</label><input type="number" id="c_trab" min="1" value="1" oninput="calcJ()"></div>
          <div class="form-group"><label>Días trabajados</label><input type="number" id="c_dias" min="1" value="1" oninput="calcJ()"></div>
          <div class="form-group"><label>Valor por jornal/día (COP)</label><input type="text" id="c_vdia" value="60.000" oninput="calcJ()"></div>
        </div>
        <div class="calc-result" id="calcResult" style="display:none">
          <span class="calc-result__label">Resultado:</span>
          <span class="calc-result__valor" id="calcValor">\u2014</span>
          <button type="button" class="button button--sm button--primary" onclick="applyCalc()"><i class="fa-solid fa-arrow-right"></i> Aplicar</button>
        </div>
      </div>
    </div>

    <div class="form-actions">
      <button type="submit" class="button button--primary"><i class="fa-solid fa-floppy-disk"></i> Guardar recibo</button>
      <a href="<<<url_for('lista_recibos')>>>" class="button button--ghost-dark"><i class="fa-solid fa-list"></i> Ver todos</a>
    </div>
  </form>
</div></section>
[[ endblock ]]
[[ block scripts ]]
<script>
const TRABAJADORES = <<<trabajadores|tojson>>>;
const btnForce = document.getElementById('btnForce');
if(btnForce){btnForce.addEventListener('click',()=>{document.getElementById('forceField').value='true';document.getElementById('reciboForm').submit();});}
initTrabajadorAutocomplete(TRABAJADORES,{
  searchInput:document.getElementById('buscarTrabajador'),
  suggestionsList:document.getElementById('trabajadorSuggestions'),
  fields:{proveedor:document.getElementById('proveedor'),nit:document.getElementById('nit'),direccion:document.getElementById('direccion'),telefono:document.getElementById('telefono'),ciudad:document.getElementById('ciudad'),concepto:document.getElementById('concepto'),valor_operacion:document.getElementById('valor_operacion'),neto_a_pagar:document.getElementById('neto_a_pagar')}
});
function syncNeto(el){const n=document.getElementById('neto_a_pagar');if(!n.value||!n.dataset.edited)n.value=el.value;}
document.getElementById('neto_a_pagar').addEventListener('input',function(){this.dataset.edited='1';});
function toggleCalc(){const b=document.getElementById('calcBody'),i=document.getElementById('calcIcon'),o=b.style.display==='none';b.style.display=o?'block':'none';i.style.transform=o?'rotate(180deg)':'';}
function pCOP(s){return parseFloat((s||'0').replace(/\\./g,'').replace(',','.'))||0;}
function fCOP(n){return Math.round(n).toLocaleString('es-CO');}
function calcJ(){const t=parseInt(document.getElementById('c_trab').value)||0,d=parseInt(document.getElementById('c_dias').value)||0,v=pCOP(document.getElementById('c_vdia').value),r=t*d*v,res=document.getElementById('calcResult');document.getElementById('calcValor').textContent=r>0?'$ '+fCOP(r):'\u2014';res.style.display=r>0?'flex':'none';}
function applyCalc(){const t=parseInt(document.getElementById('c_trab').value)||0,d=parseInt(document.getElementById('c_dias').value)||0,v=pCOP(document.getElementById('c_vdia').value),r=t*d*v;if(r>0){const fmt=fCOP(r);document.getElementById('valor_operacion').value=fmt;document.getElementById('neto_a_pagar').value=fmt;const c=document.getElementById('concepto');if(!c.value)c.value=t+' trabajador(es) × '+d+' día(s) × $ '+fCOP(v)+'/día';}}
</script>
[[ endblock ]]
""")


def _tpl_lote():
    return _j("""\
[[ extends "base.html" ]]
[[ block title ]]Recibo por Lote | Contabilidad Arroceras[[ endblock ]]
[[ block content ]]
<div class="page-hero">
  <div class="container page-hero__inner">
    <div class="page-hero__icon"><i class="fa-solid fa-users-between-lines"></i></div>
    <div>
      <h1 class="page-hero__title">Recibo por Lote</h1>
      <p class="page-hero__sub">Mismo concepto para varios trabajadores — un recibo por cada uno.</p>
    </div>
    <div class="page-hero__actions">
      <a href="<<<url_for('lista_recibos')>>>" class="button button--ghost"><i class="fa-solid fa-list"></i> Ver recibos</a>
      <a href="<<<url_for('nuevo_recibo')>>>" class="button button--ghost"><i class="fa-solid fa-file-invoice-dollar"></i> Individual</a>
    </div>
  </div>
</div>
<section class="section"><div class="container">
  [[ if success ]]<div class="alert alert--success"><i class="fa-solid fa-circle-check"></i> <<<success>>></div>[[ endif ]]
  [[ if error ]]<div class="alert alert--danger"><i class="fa-solid fa-circle-xmark"></i> <<<error>>></div>[[ endif ]]

  <form method="POST" id="loteForm">
    <div class="form-card" style="margin-bottom:var(--s6)">
      <div class="form-section">
        <h3 class="form-section__title"><span class="step-badge">1</span> Datos comunes del lote</h3>
        <div class="form-grid form-grid--3">
          <div class="form-group">
            <label for="serial_inicio">Serial de inicio <span class="req-star">*</span></label>
            <input type="number" id="serial_inicio" name="serial_inicio" value="<<<next_serial>>>" min="1" required>
            <p class="field-hint">Los seriales se asignan consecutivamente.</p>
          </div>
          <div class="form-group">
            <label for="fecha">Fecha</label>
            <input type="date" id="fecha" name="fecha" value="<<<today>>>">
          </div>
          <div class="form-group">
            <label for="valor_por_trabajador">Valor por trabajador (COP)</label>
            <input type="text" id="valor_por_trabajador" name="valor_por_trabajador" placeholder="Ej: 60.000" inputmode="numeric">
            <p class="field-hint">Vacío = valor habitual de cada uno.</p>
          </div>
          <div class="form-group" style="grid-column:1/-1">
            <label for="concepto">Concepto <span class="req-star">*</span></label>
            <textarea id="concepto" name="concepto" rows="2" required placeholder="Ej: Despalillada 3 jornales x semana en el lote El Mangon"></textarea>
          </div>
          <div class="form-group">
            <label for="direccion">Dirección (opcional)</label>
            <input type="text" id="direccion" name="direccion" placeholder="Ej: El Tambo">
          </div>
          <div class="form-group">
            <label for="ciudad">Ciudad (opcional)</label>
            <input type="text" id="ciudad" name="ciudad" placeholder="Ej: Natagaima">
          </div>
        </div>
        <details class="calc-inline" style="margin-top:1rem">
          <summary class="calc-inline__toggle"><i class="fa-solid fa-calculator"></i> Calculadora de jornales</summary>
          <div class="calc-inline__body">
            <div class="form-grid form-grid--3" style="margin-top:.8rem">
              <div class="form-group"><label>Días trabajados</label><input type="number" id="lc_dias" min="1" value="1" oninput="calcLote()"></div>
              <div class="form-group"><label>Valor por jornal/día (COP)</label><input type="text" id="lc_vdia" value="60.000" oninput="calcLote()"></div>
              <div class="form-group"><label>Resultado</label><input type="text" id="lc_res" readonly style="background:var(--clr-50);font-weight:600"></div>
            </div>
            <button type="button" class="button button--sm button--primary" onclick="applyLote()"><i class="fa-solid fa-arrow-right"></i> Aplicar al valor</button>
          </div>
        </details>
      </div>
    </div>

    <div class="form-card">
      <div class="form-section">
        <div class="worker-selector-header">
          <h3 class="worker-selector-title"><span class="step-badge">2</span> Seleccionar trabajadores</h3>
          <span class="selection-counter" id="selCounter">0 seleccionados</span>
        </div>
        <div class="cargo-filters" id="cargoFilters">
          <button type="button" class="cargo-filter-btn is-active" data-cargo="todos">Todos</button>
          [[ set ns = namespace(cargos=[]) ]]
          [[ for w in db_workers ]]
            [[ if w.trabajo_desarrolla and w.trabajo_desarrolla not in ns.cargos ]]
              [[ set ns.cargos = ns.cargos + [w.trabajo_desarrolla] ]]
              <button type="button" class="cargo-filter-btn" data-cargo="<<<w.trabajo_desarrolla>>>"><<<w.trabajo_desarrolla|replace('_',' ')|title>>></button>
            [[ endif ]]
          [[ endfor ]]
        </div>
        <div style="margin:.6rem 0 1rem">
          <input type="text" id="workerSearch" placeholder="Filtrar por nombre\u2026" class="lista-search-input" style="max-width:320px" oninput="filterW()">
        </div>
        [[ if db_workers ]]
        <div class="workers-grid" id="workersGrid">
          [[ for w in db_workers ]]
          <div class="worker-card-sel" data-id="<<<w.id_worker>>>" data-cargo="<<<w.trabajo_desarrolla or ''>>>" data-nombre="<<<(w.name~' '~w.lastname)|lower>>>" onclick="toggleW(this)">
            <i class="fa-solid fa-check worker-card-sel__check"></i>
            <div class="worker-card-sel__name"><<<w.name>>> <<<w.lastname>>></div>
            <span class="worker-card-sel__cargo"><<<(w.trabajo_desarrolla or 'sin cargo')|replace('_',' ')>>></span>
          </div>
          [[ endfor ]]
        </div>
        <div id="hiddenInputs"></div>
        <div id="selResumen" style="display:none;margin-top:1rem" class="alert alert--info">
          <i class="fa-solid fa-circle-info"></i>
          Se crearán <strong id="numR">0</strong> recibos (seriales <strong id="sRange">\u2014</strong>).
        </div>
        [[ else ]]
        <div class="empty-state" style="padding:var(--s12)">
          <i class="fa-solid fa-user-plus empty-state__icon"></i><h3>No hay trabajadores registrados</h3>
          <a href="<<<url_for('create_worker')>>>" class="button button--primary">Registrar trabajador</a>
        </div>
        [[ endif ]]
      </div>
      <div class="form-actions">
        <button type="submit" class="button button--primary" id="btnLote" disabled><i class="fa-solid fa-floppy-disk"></i> Crear recibos del lote</button>
        <a href="<<<url_for('lista_recibos')>>>" class="button button--ghost-dark"><i class="fa-solid fa-list"></i> Ver recibos</a>
      </div>
    </div>
  </form>
</div></section>
[[ endblock ]]
[[ block scripts ]]
<script>
let sel=new Set(),activeCargo='todos',sterm='';
function toggleW(c){const id=c.dataset.id;if(sel.has(id)){sel.delete(id);c.classList.remove('is-selected');}else{sel.add(id);c.classList.add('is-selected');}updCounter();updHidden();updRes();}
function updCounter(){const n=sel.size;document.getElementById('selCounter').textContent=n+' seleccionado'+(n!==1?'s':'');document.getElementById('btnLote').disabled=n===0;}
function updHidden(){const c=document.getElementById('hiddenInputs');c.innerHTML='';sel.forEach(id=>{const i=document.createElement('input');i.type='hidden';i.name='worker_ids';i.value=id;c.appendChild(i);});}
function updRes(){const n=sel.size,res=document.getElementById('selResumen');if(!res)return;if(n===0){res.style.display='none';return;}const ini=parseInt(document.getElementById('serial_inicio').value)||1;res.style.display='block';document.getElementById('numR').textContent=n;document.getElementById('sRange').textContent=n===1?ini:ini+' – '+(ini+n-1);}
document.getElementById('serial_inicio').addEventListener('input',updRes);
document.getElementById('cargoFilters').addEventListener('click',function(e){const b=e.target.closest('.cargo-filter-btn');if(!b)return;document.querySelectorAll('.cargo-filter-btn').forEach(x=>x.classList.remove('is-active'));b.classList.add('is-active');activeCargo=b.dataset.cargo;applyFilters();});
function filterW(){sterm=document.getElementById('workerSearch').value.toLowerCase();applyFilters();}
function applyFilters(){document.querySelectorAll('.worker-card-sel').forEach(c=>{const co=activeCargo==='todos'||c.dataset.cargo===activeCargo,so=!sterm||c.dataset.nombre.includes(sterm);c.style.display=(co&&so)?'':'none';});}
function pCOP(s){return parseFloat((s||'0').replace(/\\./g,'').replace(',','.'))||0;}
function fCOP(n){return Math.round(n).toLocaleString('es-CO');}
function calcLote(){const d=parseInt(document.getElementById('lc_dias').value)||0,v=pCOP(document.getElementById('lc_vdia').value),r=d*v;document.getElementById('lc_res').value=r>0?'$ '+fCOP(r):'';}
function applyLote(){const d=parseInt(document.getElementById('lc_dias').value)||0,v=pCOP(document.getElementById('lc_vdia').value),r=d*v;if(r>0)document.getElementById('valor_por_trabajador').value=fCOP(r);}
</script>
[[ endblock ]]
""")


def _tpl_detalle():
    return _j("""\
[[ extends "base.html" ]]
[[ block title ]]Recibo #<<<recibo.serial>>> | Contabilidad Arroceras[[ endblock ]]
[[ block content ]]
<div class="page-hero">
  <div class="container page-hero__inner">
    <div class="page-hero__icon"><i class="fa-solid fa-file-invoice"></i></div>
    <div>
      <h1 class="page-hero__title">Recibo #<<<recibo.serial>>></h1>
      <p class="page-hero__sub"><<<recibo.fecha.strftime('%d/%m/%Y') if recibo.fecha else 'Sin fecha'>>> &middot; <<<recibo.proveedor>>></p>
    </div>
    <div class="page-hero__actions">
      <a href="<<<url_for('lista_recibos')>>>" class="button button--ghost"><i class="fa-solid fa-arrow-left"></i> Volver</a>
      <button type="button" class="button button--ghost" onclick="window.print()"><i class="fa-solid fa-print"></i> Imprimir</button>
    </div>
  </div>
</div>
<section class="section"><div class="container" style="max-width:700px">
  <div class="recibo-card print-area">
    <div class="recibo-card__header">
      <div class="recibo-card__logo"><i class="fa-solid fa-wheat-awn"></i></div>
      <div class="recibo-card__empresa">
        <span class="recibo-card__empresa-nombre">Arrocera El Mangón</span>
        <span class="recibo-card__empresa-sub">Contabilidad Interna</span>
      </div>
      <div class="recibo-card__serial">
        <span class="recibo-card__serial-label">RECIBO</span>
        <span class="recibo-card__serial-num">#<<<recibo.serial>>></span>
      </div>
    </div>
    <dl class="recibo-dl" style="padding:var(--s5) var(--s6)">
      <div class="recibo-dl__row"><dt>Fecha</dt><dd><<<recibo.fecha.strftime('%d/%m/%Y') if recibo.fecha else '\u2014'>>></dd></div>
      <div class="recibo-dl__row"><dt>Proveedor</dt><dd><strong><<<recibo.proveedor>>></strong></dd></div>
      <div class="recibo-dl__row"><dt>NIT / Cédula</dt><dd><<<recibo.nit or '\u2014'>>></dd></div>
      <div class="recibo-dl__row"><dt>Dirección</dt><dd><<<recibo.direccion or '\u2014'>>></dd></div>
      <div class="recibo-dl__row"><dt>Teléfono</dt><dd><<<recibo.telefono or '\u2014'>>></dd></div>
      <div class="recibo-dl__row"><dt>Ciudad</dt><dd><<<recibo.ciudad or '\u2014'>>></dd></div>
      <div class="recibo-dl__row recibo-dl__row--full"><dt>Concepto</dt><dd><<<recibo.concepto or '\u2014'>>></dd></div>
      <div class="recibo-dl__row"><dt>Valor operación</dt><dd class="valor-cell">[[ if recibo.valor_operacion ]]$ <<<"{:,.0f}".format(recibo.valor_operacion).replace(",",".")>>>[[ else ]]\u2014[[ endif ]]</dd></div>
      <div class="recibo-dl__row recibo-dl__row--highlight"><dt>Neto a pagar</dt><dd class="valor-cell">[[ if recibo.neto_a_pagar ]]<strong>$ <<<"{:,.0f}".format(recibo.neto_a_pagar).replace(",",".")>>></strong>[[ else ]]\u2014[[ endif ]]</dd></div>
    </dl>
    <div class="recibo-card__footer no-print">
      <div class="recibo-firma"><div class="recibo-firma__linea"></div><span>Firma del trabajador</span></div>
      <div class="recibo-firma"><div class="recibo-firma__linea"></div><span>Firma del empleador</span></div>
    </div>
  </div>
  <div class="form-actions no-print" style="margin-top:1.5rem">
    <a href="<<<url_for('lista_recibos')>>>" class="button button--ghost-dark"><i class="fa-solid fa-arrow-left"></i> Volver</a>
    <form method="POST" action="<<<url_for('eliminar_recibo', serial=recibo.serial)>>>"
          onsubmit="return confirm('¿Eliminar el recibo #<<<recibo.serial>>>? Esta acción no se puede deshacer.')"
          style="margin:0">
      <button type="submit" class="button button--danger"><i class="fa-solid fa-trash"></i> Eliminar recibo</button>
    </form>
  </div>
</div></section>
[[ endblock ]]
""")


def _tpl_reportes():
    return _j("""\
[[ extends "base.html" ]]
[[ block title ]]Reportes | Contabilidad Arroceras[[ endblock ]]
[[ block head ]]<script src="https://cdn.jsdelivr.net/npm/chart.js@4.4.0/dist/chart.umd.min.js"></script>[[ endblock ]]
[[ block content ]]
<div class="page-hero">
  <div class="container page-hero__inner">
    <div class="page-hero__icon"><i class="fa-solid fa-chart-line"></i></div>
    <div>
      <h1 class="page-hero__title">Reportes y Estadísticas</h1>
      <p class="page-hero__sub">Control financiero &middot; <<<total_ha>>> hectáreas</p>
    </div>
    <div class="page-hero__actions">
      <a href="<<<url_for('exportar_txt')>>>" class="button button--ghost"><i class="fa-solid fa-file-export"></i> Exportar TXT</a>
    </div>
  </div>
</div>
<section class="section"><div class="container">
  <div class="stats-grid">
    <div class="stat-card [[ if pct_gasto > 90 ]]stat-card--danger[[ elif pct_gasto > 70 ]]stat-card--warning[[ else ]]stat-card--success[[ endif ]]">
      <span class="stat-card__label"><i class="fa-solid fa-peso-sign"></i> Total gastado</span>
      <span class="stat-card__value">$ <<<"{:,.0f}".format(total_gastado).replace(",",".")>>></span>
      <span class="stat-card__sub">de $ <<<"{:,.0f}".format(max_gasto).replace(",",".")>>> máximo</span>
    </div>
    <div class="stat-card [[ if pct_gasto > 90 ]]stat-card--danger[[ elif pct_gasto > 70 ]]stat-card--warning[[ else ]]stat-card--success[[ endif ]]">
      <span class="stat-card__label"><i class="fa-solid fa-percent"></i> Presupuesto usado</span>
      <span class="stat-card__value"><<<pct_gasto>>>%</span>
      <span class="stat-card__sub">Máx. $ <<<"{:,.0f}".format(max_gasto_ha).replace(",",".")>>>/ha</span>
    </div>
    <div class="stat-card [[ if pct_produccion >= 100 ]]stat-card--success[[ elif pct_produccion >= 60 ]]stat-card--warning[[ else ]]stat-card--danger[[ endif ]]">
      <span class="stat-card__label"><i class="fa-solid fa-wheat-awn"></i> Cargas cosechadas</span>
      <span class="stat-card__value"><<<"{:,}".format(total_cargas).replace(",",".")>>></span>
      <span class="stat-card__sub">mínimo <<<"{:,}".format(min_cargas).replace(",",".")>>> cargas</span>
    </div>
    <div class="stat-card [[ if pct_produccion >= 100 and pct_gasto <= 90 ]]stat-card--success[[ else ]]stat-card--warning[[ endif ]]">
      <span class="stat-card__label"><i class="fa-solid fa-seedling"></i> Rentabilidad</span>
      <span class="stat-card__value">[[ if pct_produccion >= 100 and pct_gasto <= 90 ]]&#x2705; Rentable[[ elif pct_gasto > 90 ]]&#x1F534; Riesgo[[ else ]]&#x26A0;&#xFE0F; En curso[[ endif ]]</span>
      <span class="stat-card__sub">Producción: <<<pct_produccion>>>%</span>
    </div>
  </div>
  <div class="gauge-section">
    <div class="gauge-header">
      <span class="gauge-title"><i class="fa-solid fa-peso-sign"></i> Gasto total vs presupuesto máximo</span>
      <span class="gauge-amount"><strong>$ <<<"{:,.0f}".format(total_gastado).replace(",",".")>>></strong> / $ <<<"{:,.0f}".format(max_gasto).replace(",",".")>>></span>
    </div>
    <div class="gauge-bar-wrap">
      <div class="gauge-bar [[ if pct_gasto > 90 ]]gauge-bar--red[[ elif pct_gasto > 70 ]]gauge-bar--yellow[[ else ]]gauge-bar--green[[ endif ]]" style="width:<<<[pct_gasto,100]|min>>>%"><<<pct_gasto>>>%</div>
    </div>
    <p class="gauge-footnote">Máximo: $\u00a0<<<"{:,.0f}".format(max_gasto_ha).replace(",",".")>>> &times; <<<total_ha>>>\u00a0ha = $\u00a0<<<"{:,.0f}".format(max_gasto).replace(",",".")>>></p>
  </div>
  <div class="gauge-section">
    <div class="gauge-header">
      <span class="gauge-title"><i class="fa-solid fa-wheat-awn"></i> Producción vs mínimo requerido</span>
      <span class="gauge-amount"><strong><<<"{:,}".format(total_cargas).replace(",",".")>>> cargas</strong> / <<<"{:,}".format(min_cargas).replace(",",".")>>> mínimo</span>
    </div>
    <div class="gauge-bar-wrap">
      <div class="gauge-bar [[ if pct_produccion >= 100 ]]gauge-bar--green[[ elif pct_produccion >= 60 ]]gauge-bar--yellow[[ else ]]gauge-bar--red[[ endif ]]" style="width:<<<[pct_produccion,100]|min>>>%"><<<pct_produccion>>>%</div>
    </div>
    <p class="gauge-footnote">Mínimo: <<<min_cargas>>> cargas (100 bultos/ha &times; <<<total_ha>>>\u00a0ha). Cada carga = 62.5\u00a0kg.</p>
  </div>
  <div class="charts-row">
    <div class="chart-card"><h3 class="chart-card__title"><i class="fa-solid fa-chart-bar"></i> Gastos por mes</h3><div class="chart-canvas-wrap"><canvas id="chartMes"></canvas></div></div>
    <div class="chart-card"><h3 class="chart-card__title"><i class="fa-solid fa-users"></i> Top trabajadores</h3><div class="chart-canvas-wrap"><canvas id="chartTrab"></canvas></div></div>
  </div>
  <div class="grid-3" style="margin-top:var(--s6)">
    <a href="<<<url_for('reporte_semana')>>>" class="card" style="text-decoration:none"><div class="card__icon"><i class="fa-solid fa-calendar-week"></i></div><h3>Reporte semanal</h3><p>Ver pagos de la semana actual.</p></a>
    <a href="<<<url_for('exportar_txt')>>>" class="card" style="text-decoration:none"><div class="card__icon"><i class="fa-solid fa-file-export"></i></div><h3>Exportar TXT</h3><p>Descargar todos los recibos.</p></a>
    <a href="<<<url_for('nueva_cosecha')>>>" class="card" style="text-decoration:none"><div class="card__icon"><i class="fa-solid fa-wheat-awn"></i></div><h3>Registrar cosecha</h3><p>Añadir producción de cargas y kg.</p></a>
  </div>

  <!-- ═══ EXPORTAR PDF / EXCEL ═══════════════════════════════════════ -->
  <div style="margin-top:var(--s8)">
    <h2 style="font-size:1.15rem;font-weight:700;color:var(--clr-700);margin-bottom:var(--s4)">
      <i class="fa-solid fa-file-arrow-down"></i> Exportar reportes
    </h2>

    <!-- Filtro de fechas para gastos -->
    <div class="form-card" style="margin-bottom:var(--s5);padding:var(--s4) var(--s5)">
      <p style="font-weight:600;margin-bottom:.6rem"><i class="fa-solid fa-filter"></i> Filtro por fecha (opcional — aplica a Gastos y Excel recibos)</p>
      <div style="display:flex;gap:var(--s3);align-items:center;flex-wrap:wrap">
        <div class="form-group" style="margin-bottom:0;flex:1;min-width:140px">
          <label style="font-size:.82rem">Desde</label>
          <input type="date" id="filtroDesde" style="width:100%">
        </div>
        <div class="form-group" style="margin-bottom:0;flex:1;min-width:140px">
          <label style="font-size:.82rem">Hasta</label>
          <input type="date" id="filtroHasta" style="width:100%">
        </div>
      </div>
    </div>

    <div class="export-grid">
      <!-- PDF Gastos -->
      <a id="linkPdfGastos" href="<<<url_for('generar_pdf', tipo='gastos')>>>"
         class="export-card export-card--pdf" target="_blank">
        <span class="export-card__icon"><i class="fa-solid fa-file-pdf"></i></span>
        <span class="export-card__title">PDF — Gastos</span>
        <span class="export-card__desc">Todos los recibos con totales. Usa el filtro de fecha si aplica.</span>
      </a>
      <!-- PDF Trabajadores -->
      <a href="<<<url_for('generar_pdf', tipo='trabajadores')>>>"
         class="export-card export-card--pdf" target="_blank">
        <span class="export-card__icon"><i class="fa-solid fa-file-pdf"></i></span>
        <span class="export-card__title">PDF — Por Trabajador</span>
        <span class="export-card__desc">Total pagado por cada proveedor / trabajador.</span>
      </a>
      <!-- PDF Producción -->
      <a href="<<<url_for('generar_pdf', tipo='produccion')>>>"
         class="export-card export-card--pdf" target="_blank">
        <span class="export-card__icon"><i class="fa-solid fa-file-pdf"></i></span>
        <span class="export-card__title">PDF — Producción</span>
        <span class="export-card__desc">Registro de cosechas, cargas y kg cosechados.</span>
      </a>
      <!-- PDF Semana -->
      <a id="linkPdfSemana" href="<<<url_for('generar_pdf', tipo='semana')>>>"
         class="export-card export-card--pdf" target="_blank">
        <span class="export-card__icon"><i class="fa-solid fa-file-pdf"></i></span>
        <span class="export-card__title">PDF — Semana actual</span>
        <span class="export-card__desc">Recibos de la semana en curso.</span>
      </a>
      <!-- Excel Recibos -->
      <a id="linkXlsRecibos" href="<<<url_for('generar_excel', tipo='recibos')>>>"
         class="export-card export-card--xlsx">
        <span class="export-card__icon"><i class="fa-solid fa-file-excel"></i></span>
        <span class="export-card__title">Excel — Recibos</span>
        <span class="export-card__desc">Hoja de cálculo con todos los recibos y totales.</span>
      </a>
      <!-- Excel Trabajadores -->
      <a href="<<<url_for('generar_excel', tipo='trabajadores')>>>"
         class="export-card export-card--xlsx">
        <span class="export-card__icon"><i class="fa-solid fa-file-excel"></i></span>
        <span class="export-card__title">Excel — Trabajadores</span>
        <span class="export-card__desc">Gastos agrupados por proveedor.</span>
      </a>
      <!-- Excel Producción -->
      <a href="<<<url_for('generar_excel', tipo='produccion')>>>"
         class="export-card export-card--xlsx">
        <span class="export-card__icon"><i class="fa-solid fa-file-excel"></i></span>
        <span class="export-card__title">Excel — Producción</span>
        <span class="export-card__desc">Cosechas, cargas, kg y valores.</span>
      </a>
    </div>
  </div>

</div></section>
[[ endblock ]]
[[ block scripts ]]
<script>
const dM=<<<por_mes|tojson>>>,dT=<<<por_trabajador|tojson>>>;
const G='#2D6A4F',D='#1B4332',Gs=['#1B4332','#2D6A4F','#40916C','#52B788','#74C69D','#95D5B2','#B7E4C7','#D8F3DC'];
const cM=document.getElementById('chartMes');
if(cM&&dM.length>0){new Chart(cM,{type:'bar',data:{labels:dM.map(d=>d.mes),datasets:[{label:'Gasto',data:dM.map(d=>d.total),backgroundColor:G,borderColor:D,borderWidth:1,borderRadius:6}]},options:{responsive:true,maintainAspectRatio:false,plugins:{legend:{display:false}},scales:{y:{ticks:{callback:v=>'$ '+(v/1000000).toFixed(1)+'M'}}}}})}
else if(cM)cM.parentElement.innerHTML='<p style="text-align:center;color:var(--txt-muted);padding:2rem">Sin datos de gastos aún.</p>';
const cT=document.getElementById('chartTrab');
if(cT&&dT.length>0){new Chart(cT,{type:'bar',data:{labels:dT.map(d=>d.proveedor.split(' ')[0]),datasets:[{label:'Total',data:dT.map(d=>d.total),backgroundColor:Gs,borderWidth:0,borderRadius:6}]},options:{indexAxis:'y',responsive:true,maintainAspectRatio:false,plugins:{legend:{display:false}},scales:{x:{ticks:{callback:v=>'$ '+(v/1000000).toFixed(1)+'M'}}}}})}
else if(cT)cT.parentElement.innerHTML='<p style="text-align:center;color:var(--txt-muted);padding:2rem">Sin datos aún.</p>';
// Filtro de fechas en links de exportación
(function(){
  const desde=document.getElementById('filtroDesde');
  const hasta=document.getElementById('filtroHasta');
  const linkPdfG=document.getElementById('linkPdfGastos');
  const linkXlsR=document.getElementById('linkXlsRecibos');
  function upd(){
    const d=desde.value,h=hasta.value;
    const qs=d&&h?`?tipo=gastos&desde=${d}&hasta=${h}`:'?tipo=gastos';
    if(linkPdfG)linkPdfG.href=qs;
    const qs2=d&&h?`?tipo=recibos&desde=${d}&hasta=${h}`:'?tipo=recibos';
    if(linkXlsR)linkXlsR.href=qs2;
  }
  desde.addEventListener('change',upd);
  hasta.addEventListener('change',upd);
})();
</script>
[[ endblock ]]
""")


def _tpl_semana():
    return _j("""\
[[ extends "base.html" ]]
[[ block title ]]Reporte Semanal | Contabilidad Arroceras[[ endblock ]]
[[ block content ]]
<div class="page-hero">
  <div class="container page-hero__inner">
    <div class="page-hero__icon"><i class="fa-solid fa-calendar-week"></i></div>
    <div>
      <h1 class="page-hero__title">Reporte Semanal</h1>
      <p class="page-hero__sub">[[ if inicio and fin ]]<<<inicio.strftime('%d/%m/%Y')>>> al <<<fin.strftime('%d/%m/%Y')>>>[[ else ]]Selecciona una fecha[[ endif ]]</p>
    </div>
    <div class="page-hero__actions"><a href="<<<url_for('reportes')>>>" class="button button--ghost"><i class="fa-solid fa-chart-line"></i> Dashboard</a></div>
  </div>
</div>
<section class="section"><div class="container">
  <form method="GET" class="form-card" style="max-width:400px;margin-bottom:var(--s6)">
    <div class="form-group"><label for="fecha">Semana que contiene esta fecha</label><input type="date" id="fecha" name="fecha" value="<<<fecha_str or ''>>>"></div>
    <div class="form-actions" style="margin-top:0"><button type="submit" class="button button--primary"><i class="fa-solid fa-search"></i> Ver semana</button></div>
  </form>
  [[ if recibos ]]
  <div class="table-wrap">
    <table class="data-table">
      <thead><tr><th>Serial</th><th>Fecha</th><th>Proveedor</th><th>Concepto</th><th>Neto a pagar</th></tr></thead>
      <tbody>
        [[ for r in recibos ]]
        <tr style="cursor:pointer" onclick="location.href='<<<url_for('detalle_recibo', serial=r.serial)>>>'">
          <td><span class="serial-badge"><<<r.serial>>></span></td>
          <td class="txt-muted"><<<r.fecha.strftime('%d/%m/%Y') if r.fecha else '\u2014'>>></td>
          <td><strong><<<r.proveedor>>></strong></td>
          <td class="concepto-cell small"><<<r.concepto>>></td>
          <td class="valor-cell">[[ if r.neto_a_pagar ]]<strong>$ <<<"{:,.0f}".format(r.neto_a_pagar).replace(",",".")>>></strong>[[ else ]]\u2014[[ endif ]]</td>
        </tr>
        [[ endfor ]]
      </tbody>
      <tfoot><tr class="table-foot"><td colspan="4" style="text-align:right;font-weight:600">Total semana:</td><td class="valor-cell"><strong>$ <<<"{:,.0f}".format(recibos|sum(attribute='neto_a_pagar')).replace(",",".")>>></strong></td></tr></tfoot>
    </table>
  </div>
  [[ elif fecha_str ]]
  <div class="empty-state"><i class="fa-solid fa-calendar-xmark empty-state__icon"></i><h3>Sin recibos esa semana</h3></div>
  [[ endif ]]
</div></section>
[[ endblock ]]
""")


def _tpl_produccion_lista():
    return _j("""\
[[ extends "base.html" ]]
[[ block title ]]Producción | Contabilidad Arroceras[[ endblock ]]
[[ block content ]]
<div class="page-hero">
  <div class="container page-hero__inner">
    <div class="page-hero__icon"><i class="fa-solid fa-wheat-awn"></i></div>
    <div>
      <h1 class="page-hero__title">Producción / Cosechas</h1>
      <p class="page-hero__sub"><<<total_cargas>>> cargas de <<<min_cargas>>> requeridas (<<<pct_produccion>>>%)</p>
    </div>
    <div class="page-hero__actions"><a href="<<<url_for('nueva_cosecha')>>>" class="button button--primary"><i class="fa-solid fa-plus"></i> Registrar cosecha</a></div>
  </div>
</div>
<section class="section"><div class="container">
  <div class="gauge-section" style="margin-bottom:var(--s6)">
    <div class="gauge-header">
      <span class="gauge-title"><i class="fa-solid fa-wheat-awn"></i> Avance de producción</span>
      <span class="gauge-amount"><strong><<<total_cargas>>> cargas</strong> / <<<min_cargas>>> mínimo</span>
    </div>
    <div class="gauge-bar-wrap">
      <div class="gauge-bar [[ if pct_produccion >= 100 ]]gauge-bar--green[[ elif pct_produccion >= 60 ]]gauge-bar--yellow[[ else ]]gauge-bar--red[[ endif ]]" style="width:<<<[pct_produccion,100]|min>>>%"><<<pct_produccion>>>%</div>
    </div>
  </div>
  [[ if cosechas ]]
  <div class="table-wrap">
    <table class="data-table">
      <thead><tr><th>Fecha</th><th>Lote</th><th>Cargas</th><th>Kg totales</th><th>Notas</th></tr></thead>
      <tbody>
        [[ for c in cosechas ]]
        <tr>
          <td class="txt-muted"><<<c.fecha.strftime('%d/%m/%Y') if c.fecha else '\u2014'>>></td>
          <td><<<c.lote or '\u2014'>>></td>
          <td><strong><<<c.cargas>>></strong></td>
          <td class="txt-muted"><<<"{:,.1f}".format(c.kg_total).replace(",",".") if c.kg_total else '\u2014'>>></td>
          <td class="small txt-muted"><<<c.notas or ''>>></td>
        </tr>
        [[ endfor ]]
      </tbody>
      <tfoot><tr class="table-foot"><td colspan="2" style="text-align:right;font-weight:600">Totales:</td><td><strong><<<total_cargas>>></strong></td><td class="txt-muted"><<<"{:,.1f}".format(cosechas|sum(attribute='kg_total')).replace(",",".") if cosechas else 0>>> kg</td><td></td></tr></tfoot>
    </table>
  </div>
  [[ else ]]
  <div class="empty-state"><i class="fa-solid fa-seedling empty-state__icon"></i><h3>Sin cosechas registradas</h3><a href="<<<url_for('nueva_cosecha')>>>" class="button button--primary"><i class="fa-solid fa-plus"></i> Registrar cosecha</a></div>
  [[ endif ]]
</div></section>
[[ endblock ]]
""")


def _tpl_produccion_nueva():
    return _j("""\
[[ extends "base.html" ]]
[[ block title ]]Nueva Cosecha | Contabilidad Arroceras[[ endblock ]]
[[ block content ]]
<div class="page-hero">
  <div class="container page-hero__inner">
    <div class="page-hero__icon"><i class="fa-solid fa-seedling"></i></div>
    <div><h1 class="page-hero__title">Registrar Cosecha</h1><p class="page-hero__sub">Ingresa los bultos / cargas recolectados.</p></div>
    <div class="page-hero__actions"><a href="<<<url_for('lista_produccion')>>>" class="button button--ghost"><i class="fa-solid fa-wheat-awn"></i> Ver producción</a></div>
  </div>
</div>
<section class="section"><div class="container" style="max-width:600px">
  [[ if success ]]<div class="alert alert--success"><i class="fa-solid fa-circle-check"></i> <<<success>>></div>[[ endif ]]
  [[ if error ]]<div class="alert alert--danger"><i class="fa-solid fa-circle-xmark"></i> <<<error>>></div>[[ endif ]]
  <form method="POST" class="form-card">
    <div class="form-section">
      <div class="form-grid form-grid--2">
        <div class="form-group"><label for="fecha">Fecha <span class="req-star">*</span></label><input type="date" id="fecha" name="fecha" required value="<<<today>>>"></div>
        <div class="form-group"><label for="lote">Lote / Parcela</label><input type="text" id="lote" name="lote" placeholder="Ej: El Mangón"></div>
        <div class="form-group"><label for="cargas">Cargas (bultos) <span class="req-star">*</span></label><input type="number" id="cargas" name="cargas" min="1" required oninput="calcKg()"><p class="field-hint">1 carga = 62.5 kg</p></div>
        <div class="form-group"><label for="kg_totales">Kg totales</label><input type="text" id="kg_totales" name="kg_totales" readonly style="background:var(--clr-50)"></div>
        <div class="form-group" style="grid-column:1/-1"><label for="observaciones">Notas</label><textarea id="observaciones" name="observaciones" rows="2" placeholder="Observaciones opcionales\u2026"></textarea></div>
      </div>
    </div>
    <div class="form-actions">
      <button type="submit" class="button button--primary"><i class="fa-solid fa-floppy-disk"></i> Guardar cosecha</button>
      <a href="<<<url_for('lista_produccion')>>>" class="button button--ghost-dark">Cancelar</a>
    </div>
  </form>
</div></section>
[[ endblock ]]
[[ block scripts ]]<script>function calcKg(){const c=parseFloat(document.getElementById('cargas').value)||0;document.getElementById('kg_totales').value=c>0?(c*62.5).toFixed(1):'';}  </script>[[ endblock ]]
""")


def _tpl_config():
    return _j("""\
[[ extends "base.html" ]]
[[ block title ]]Configuración | Contabilidad Arroceras[[ endblock ]]
[[ block content ]]
<div class="page-hero">
  <div class="container page-hero__inner">
    <div class="page-hero__icon"><i class="fa-solid fa-gear"></i></div>
    <div><h1 class="page-hero__title">Configuración</h1><p class="page-hero__sub">Parámetros del sistema</p></div>
  </div>
</div>
<section class="section"><div class="container" style="max-width:580px">
  [[ if message ]]<div class="alert alert--success"><i class="fa-solid fa-circle-check"></i> <<<message>>></div>[[ endif ]]
  [[ if error ]]<div class="alert alert--danger"><<<error>>></div>[[ endif ]]
  <div class="form-card">
    <div class="form-section">
      <h3 class="form-section__title"><i class="fa-solid fa-hashtag"></i> Seriales</h3>
      <div class="config-info-box">
        <i class="fa-solid fa-circle-info"></i>
        <div><strong>Total de recibos:</strong> <<<total_recibos>>><br><strong>Serial inicial configurado:</strong> <<<config.get('serial_inicial', '1')>>></div>
      </div>
      <form method="POST">
        <div class="form-group" style="max-width:280px;margin-top:1.5rem">
          <label for="serial_inicial">Serial inicial para nuevos recibos</label>
          <input type="number" id="serial_inicial" name="serial_inicial" value="<<<config.get('serial_inicial', '1')>>>" min="1" required>
          <p class="field-hint">Define desde qué número empezarán los seriales cuando no haya recibos previos.[[ if total_recibos > 0 ]] <strong>Nota:</strong> ya hay <<<total_recibos>>> recibo(s); el próximo serial se calcula automáticamente.[[ endif ]]</p>
        </div>
        <div class="form-actions"><button type="submit" class="button button--primary"><i class="fa-solid fa-floppy-disk"></i> Guardar configuración</button></div>
      </form>
    </div>
  </div>
</div></section>
[[ endblock ]]
""")


# =========================
# MAIN
# =========================
if __name__ == '__main__':
    init_templates()
    init_database()
    app.run(debug=True)